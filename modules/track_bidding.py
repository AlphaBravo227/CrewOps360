# modules/track_bidding.py
"""
Track Bidding module — lets staff bid on shifts for a future track cycle.
Admin controls: create bid tracks, toggle bidding, set capacity, manage per-staff
bid access, add/update/remove bids on staff members' behalf, promote to active.
"""

import streamlit as st
import pandas as pd
import altair as alt
from datetime import datetime, date, timedelta
import pytz
import json
import io

_eastern_tz = pytz.timezone('America/New_York')

from modules.db_utils import (
    initialize_database,
    get_active_track_config,
    get_bidding_track_config,
    get_track_config_by_name,
    get_all_track_configs,
    create_track_config,
    update_track_config,
    toggle_bidding,
    get_track_capacity,
    get_track_capacity_by_weekday,
    get_weekday_capacity_overrides,
    set_weekday_capacity_override,
    promote_bid_to_active,
    save_bid_track_to_db,
    get_bid_track_from_db,
    save_bid_draft,
    get_bid_draft,
    delete_bid_draft,
    get_all_bid_tracks,
    get_track_from_db,
    delete_track_config,
    delete_bid,
    wipe_all_bids,
    get_bid_access,
    set_bid_access,
    get_all_bid_access_configs,
    get_all_bid_access_details,
    log_bid_progression_event,
    get_bid_progression_log,
)
from modules.security import check_admin_access
from modules.shift_definitions import day_shifts, night_shifts


# ──────────────────────────────────────────────
# Shared data loading (staff roster + Excel files used by the bidding editor)
# ──────────────────────────────────────────────

def _get_preassignment_day_columns(path):
    """
    Read the ordered list of day-pattern columns (e.g. "Sun A 1" ... "Sat C 6")
    straight from the preassignments file's header row.

    Read independently of load_preassignments() (which may collapse the file into
    a plain dict when duplicate staff names are present) so the day schema is
    always derived directly from the file's own columns, never from Tracks.xlsx.
    """
    header_df = pd.read_excel(path, nrows=0)
    cols = list(header_df.columns)
    staff_col = cols[0]
    for col in cols:
        if isinstance(col, str) and "name" in col.lower() and "staff" in col.lower():
            staff_col = col
            break
    return [c for c in cols if c != staff_col]


def _load_bidding_data_files():
    """
    Load and column-map the Excel data files used by the bidding interface
    (same files/pattern as the clinical track hub). Shared by the staff-facing
    flow and the admin "Add/Update Selection" / "Manage Bid Access" tabs.

    Returns:
        tuple: (ctx, error) — ctx is a dict with everything the bidding staff
        interface needs (minus the selected staff and track-specific capacity),
        or None if a required file/column is missing, in which case error is a
        message describing what's wrong.
    """
    import os
    from modules.column_mapper import auto_detect_columns
    from modules.track_management.preassignment import load_preassignments

    excel_files = {
        "preferences": None,
        "current_tracks": None,
        "requirements": None,
        "preassignments": None,
    }
    upload_dir = "upload files"
    if os.path.exists(upload_dir):
        for f in os.listdir(upload_dir):
            fl = f.lower()
            fp = os.path.join(upload_dir, f)
            if 'preference' in fl and fl.endswith('.xlsx'):
                excel_files['preferences'] = fp
            elif 'track' in fl and fl.endswith('.xlsx'):
                excel_files['current_tracks'] = fp
            elif 'requirement' in fl and fl.endswith('.xlsx'):
                excel_files['requirements'] = fp
            elif 'preassignment' in fl and fl.endswith('.xlsx'):
                excel_files['preassignments'] = fp

    if not excel_files['preferences']:
        return None, "Preferences file not found in 'upload files' folder."

    if not excel_files['current_tracks']:
        return None, "Current tracks file not found in 'upload files' folder."

    if not excel_files['preassignments']:
        return None, "Preassignments file not found in 'upload files' folder."

    def load_excel(path):
        return pd.read_excel(path) if path else None

    preferences_df = load_excel(excel_files['preferences'])
    current_tracks_df = load_excel(excel_files['current_tracks'])
    requirements_df = load_excel(excel_files['requirements'])

    # NOTE: preassignments must be loaded via load_preassignments() rather than a bare
    # load_excel() — get_staff_preassignments() looks staff up via
    # preassignment_df.loc[staff_name], which requires the staff-name index (and
    # duplicate-row handling) that only load_preassignments() sets up.
    preassignment_df = load_preassignments()

    if preferences_df is None:
        return None, "Could not load preferences file."

    # Detect columns
    detection = auto_detect_columns(preferences_df, current_tracks_df)
    mappings = detection['column_mappings']
    staff_col_prefs = mappings['staff_col_prefs']
    staff_col_tracks = mappings.get('staff_col_tracks')
    role_col = mappings['role_col']
    # The bid cycle's day schema comes from Preassignments.xlsx (the file authored
    # per bid cycle), not from Tracks.xlsx (last cycle's active roster) — Tracks.xlsx
    # is still used below only as reference data (capacity/options + "Your Active
    # Track" comparison), never as the source of which day columns exist.
    days = _get_preassignment_day_columns(excel_files['preassignments'])
    no_matrix_col = mappings.get('no_matrix_col')
    reduced_rest_col = mappings.get('reduced_rest_col')
    seniority_col = mappings.get('seniority_col')

    if not staff_col_prefs or not days:
        return None, "Could not detect required columns."

    staff_names = sorted(preferences_df[staff_col_prefs].dropna().unique().tolist())

    role_mapping = {}
    seniority_mapping = {}
    no_matrix_mapping = {}
    for _, row in preferences_df.iterrows():
        name = row.get(staff_col_prefs)
        if pd.notna(name):
            name = str(name).strip()
            if role_col:
                role_mapping[name] = row.get(role_col, 'Unknown')
            if seniority_col:
                seniority_val = row.get(seniority_col)
                try:
                    seniority_mapping[name] = int(seniority_val)
                except (TypeError, ValueError):
                    seniority_mapping[name] = seniority_val
            if no_matrix_col:
                try:
                    no_matrix_mapping[name] = int(row.get(no_matrix_col)) == 1
                except (TypeError, ValueError):
                    no_matrix_mapping[name] = False

    return {
        'preferences_df': preferences_df,
        'current_tracks_df': current_tracks_df,
        'requirements_df': requirements_df,
        'preassignment_df': preassignment_df,
        'days': days,
        'staff_col_prefs': staff_col_prefs,
        'staff_col_tracks': staff_col_tracks,
        'role_col': role_col,
        'no_matrix_col': no_matrix_col,
        'reduced_rest_col': reduced_rest_col,
        'seniority_col': seniority_col,
        'staff_names': staff_names,
        'role_mapping': role_mapping,
        'seniority_mapping': seniority_mapping,
        'no_matrix_mapping': no_matrix_mapping,
    }, None


# ──────────────────────────────────────────────
# Automatic Bid Access & Notification — when enabled for a track cycle, submitting a
# bid grants bid access to the next staff member in seniority rank order (same role)
# and emails them that their bid is now open, instead of an admin doing it by hand.
# ──────────────────────────────────────────────

def _load_requirements_map(requirements_df):
    """
    Parse Requirements.xlsx into {staff_name: {shifts_per_pay_period, night_minimum,
    weekend_minimum, weekend_group, email}}.

    Column layout is positional (STAFF NAME, SHIFTS PER PAY PERIOD, NIGHT MINIMUM,
    WEEKEND MINIMUM, WEEKEND GROUP, EMAIL), matching the per-staff parsing in
    _display_bidding_staff_interface. Staff with a blank SHIFTS PER PAY PERIOD are
    management who don't bid on tracks — shifts_per_pay_period stays None for them,
    which is what callers use to exclude them from the bidding roster.
    """
    result = {}
    if requirements_df is None or requirements_df.empty:
        return result

    cols = requirements_df.columns
    name_col = cols[0]
    for _, row in requirements_df.iterrows():
        name = row.get(name_col)
        if pd.isna(name):
            continue
        name = str(name).strip()

        entry = {
            'shifts_per_pay_period': None,
            'night_minimum': None,
            'weekend_minimum': None,
            'weekend_group': None,
            'email': None,
        }
        if len(cols) >= 2 and pd.notna(row.iloc[1]):
            entry['shifts_per_pay_period'] = int(float(row.iloc[1]))
        if len(cols) >= 3 and pd.notna(row.iloc[2]):
            entry['night_minimum'] = int(float(row.iloc[2]))
        if len(cols) >= 4 and pd.notna(row.iloc[3]):
            entry['weekend_minimum'] = int(float(row.iloc[3]))
        if len(cols) >= 5 and pd.notna(row.iloc[4]):
            wg = str(row.iloc[4]).strip().upper()
            if wg in ['A', 'B', 'C', 'D', 'E']:
                entry['weekend_group'] = wg
        if len(cols) >= 6 and pd.notna(row.iloc[5]):
            email = str(row.iloc[5]).strip()
            if email:
                entry['email'] = email

        result[name] = entry
    return result


def _bidding_role_bucket(role):
    """Collapse a raw role string to 'medic' or 'nurse' (nurse bucket includes dual) —
    mirrors the Nurse/Medic split used by the Manage Bid Access tables."""
    return 'medic' if str(role).strip().lower() == 'medic' else 'nurse'


# Probationary staff (under a year, not yet eligible to bid a track) who should be
# treated like management for bidding purposes: skipped in the bid order so they
# never get auto-notified/auto-opened when the person ahead of them bids. Unlike
# real management they still work clinical shifts, so — unlike management — they
# keep a normal SHIFTS PER PAY PERIOD in Requirements.xlsx (Summer Leave, PDF
# generation, and the track validators all depend on that field being accurate).
_BID_INELIGIBLE_STAFF = {"O'Flaherty", "Phelan", "VanderKooi"}


def _ordered_bidding_roster(staff_names, role_mapping, seniority_mapping, requirements_map, bucket):
    """
    Seniority-ascending (most senior first) list of staff in one role bucket ('nurse'
    or 'medic'), excluding anyone with no SHIFTS PER PAY PERIOD on file in
    Requirements.xlsx (management/non-bidding staff) or in _BID_INELIGIBLE_STAFF
    (probationary staff not yet eligible to bid) — both are skipped.
    """
    def _seniority_key(name):
        try:
            return (0, float(seniority_mapping.get(name)))
        except (TypeError, ValueError):
            return (1, 0)

    eligible = [
        name for name in staff_names
        if _bidding_role_bucket(role_mapping.get(name, '')) == bucket
        and requirements_map.get(name, {}).get('shifts_per_pay_period') is not None
        and name not in _BID_INELIGIBLE_STAFF
    ]
    return sorted(eligible, key=_seniority_key)


def _next_staff_in_rank(staff_name, staff_names, role_mapping, seniority_mapping, requirements_map):
    """Return the next staff member after staff_name in seniority rank order within
    staff_name's own role bucket (nurse incl. dual, or medic), or None if staff_name
    is last in that bucket, isn't in it (e.g. management), or the bucket is empty."""
    bucket = _bidding_role_bucket(role_mapping.get(staff_name, ''))
    roster = _ordered_bidding_roster(staff_names, role_mapping, seniority_mapping, requirements_map, bucket)
    if staff_name not in roster:
        return None
    idx = roster.index(staff_name)
    if idx + 1 < len(roster):
        return roster[idx + 1]
    return None


def _run_auto_bid_progression(staff_name, bid_track_name):
    """
    After staff_name's bid is saved, if automatic bid access & notification is turned
    on for bid_track_name: find the next staff member in seniority rank order (same
    role bucket), grant them bid access, and email them (+ admins) that their bid is
    open. If that staff member has no email on file, access is left untouched and the
    admins are emailed to enable/notify manually instead.

    If the next staff member already has bid access enabled for this track — e.g. an
    admin revised an earlier (more senior) staff member's bid after the cascade had
    already advanced past them — they're skipped entirely: no re-enabling access, and
    no duplicate "your bid is open" email. This only guards the next-staff cascade;
    the admin submission notice for the revision itself is sent unconditionally by
    the caller regardless of what happens here.

    Every outcome reached after the feature-enabled check is written to the
    bid_progression_log table (Manage Bid Access tab's notification log), whether or
    not an email actually went out.

    Returns:
        tuple (level, message) for display next to the existing bid-submission notice
        — level is one of 'success', 'warning', 'info' — or None if the feature is off.
    """
    cfg = get_track_config_by_name(bid_track_name)
    if not cfg or not cfg.get('auto_bid_progression'):
        return None

    def _log(next_staff, level, message, notified_email=None):
        log_bid_progression_event(bid_track_name, staff_name, next_staff, level, message, notified_email)
        return (level, message)

    try:
        ctx, roster_error = _load_bidding_data_files()
        if ctx is None:
            return _log(None, "warning", f"Automatic bid progression could not load roster data: {roster_error}")

        requirements_map = _load_requirements_map(ctx['requirements_df'])
        next_staff = _next_staff_in_rank(
            staff_name, ctx['staff_names'], ctx['role_mapping'], ctx['seniority_mapping'], requirements_map
        )
        if not next_staff:
            return _log(None, "info",
                        f"{staff_name} is last in seniority rank order — no next staff member to advance to.")

        if get_bid_access(next_staff, bid_track_name):
            return _log(next_staff, "info",
                        f"{next_staff} is next in rank, but already has bid access enabled for "
                        f"{bid_track_name} — skipped re-enabling access and did not send a duplicate notification.")

        next_requirements = requirements_map.get(next_staff, {})
        next_email = next_requirements.get('email')

        from modules.email_notifications import send_bid_access_opened_notification, send_missing_bidder_email_alert

        if not next_email:
            alert_ok, alert_msg = send_missing_bidder_email_alert(next_staff, bid_track_name)
            note = ("Admins have been emailed to enable access and notify them manually."
                    if alert_ok else f"The admin alert email also failed to send: {alert_msg}")
            return _log(next_staff, "warning",
                        f"{next_staff} is next in rank, but has no email on file in Requirements.xlsx — "
                        f"bid access was NOT automatically enabled. {note}")

        ok, _ = set_bid_access(next_staff, bid_track_name, True)
        if not ok:
            return _log(next_staff, "warning", f"{next_staff} is next in rank, but enabling their bid access failed.")

        sent_ok, sent_msg = send_bid_access_opened_notification(
            next_staff, next_email, bid_track_name, next_requirements)
        if sent_ok:
            return _log(next_staff, "success",
                        f"Bid access enabled for {next_staff} (next in rank) and notified at {next_email}.",
                        notified_email=next_email)
        else:
            return _log(next_staff, "warning",
                        f"Bid access enabled for {next_staff}, but the notification email failed: {sent_msg}")
    except Exception as e:
        return _log(None, "warning", f"Automatic bid progression failed unexpectedly: {e}")


def _send_manual_bid_notification(staff_name, manual_email, bid_track_name):
    """
    Admin-triggered "your bid is open" notification, for the fallback case where
    automatic progression couldn't send one itself (no email on file) or an admin
    otherwise wants to notify someone by hand.

    Unlike the automatic cascade, the recipient address is whatever the admin types
    in — never looked up from Requirements.xlsx — and bid access is left untouched
    (use Toggle Access below for that). Always logged, with trigger_type='manual' so
    it's distinguishable from automatic events in the notification log.

    Returns:
        tuple (level, message) — level is 'success' or 'warning'.
    """
    manual_email = (manual_email or "").strip()
    if not manual_email:
        return ("warning", "Enter an email address before sending.")

    ctx, _ = _load_bidding_data_files()
    requirements = {}
    if ctx is not None:
        requirements = _load_requirements_map(ctx['requirements_df']).get(staff_name, {})

    from modules.email_notifications import send_bid_access_opened_notification
    sent_ok, sent_msg = send_bid_access_opened_notification(staff_name, manual_email, bid_track_name, requirements)

    level = "success" if sent_ok else "warning"
    message = (f"Manually notified {staff_name} at {manual_email}."
               if sent_ok else f"Manual notification to {staff_name} at {manual_email} failed: {sent_msg}")
    log_bid_progression_event(
        bid_track_name, "Manual Send", staff_name, level, message,
        notified_email=manual_email if sent_ok else None, trigger_type='manual'
    )
    return (level, message)


# ──────────────────────────────────────────────
# Bid Analysis tab — per-day Nurse/Medic/Dual/Senior demand from submitted bids,
# mirroring the FY26 Track Analysis workbook's manual roll-up.
# ──────────────────────────────────────────────

# Medic reuses the blue nurse used to be shown in. Day/Night and D/N use a
# darker green/blue than the pastel Track Selection key for better contrast
# on the Where Staff Are Bidding and Maximum Achievable Crews charts.
_ROLE_COLORS = {'Nurse': '#f28b82', 'Medic': '#2a78d6', 'Dual': '#eda100'}
_PERIOD_COLORS = {'Day': '#66bb6a', 'Night': '#1976d2'}
_PERIOD_LINE_COLORS = {'Day': '#4c9950', 'Night': '#125aa3'}
_SHIFT_COLORS = {'D': '#66bb6a', 'N': '#1976d2', 'AT': '#898781', 'Off': '#f0efec'}


def _bid_role_and_senior(bid, role_mapping, no_matrix_mapping):
    """Resolve (effective role string, is_senior bool) for one submitted bid."""
    name = bid['staff_name']
    role = role_mapping.get(name)
    if not role or str(role).strip().lower() == 'unknown':
        role = (bid.get('metadata') or {}).get('original_role') or 'nurse'
    role = str(role).strip().lower()
    is_senior = bool(no_matrix_mapping.get(name))
    return role, is_senior


def _max_possible_shifts(nurse_n, medic_n, dual_n, senior_n):
    """
    Largest number of complete Nurse+Medic crews that day's bidders could staff.

    Dual-credentialed staff (already counted in nurse_n) can flex to the medic
    side; this tries every split and keeps the best pairing, then caps the
    result at how many no-matrix/senior staff bid that day. Direct translation
    of the LET() formula in rows 99/104 of the FY26 Track Analysis workbook.
    """
    best = max(min(nurse_n - x, medic_n + x) for x in range(dual_n + 1))
    return max(0, min(senior_n, best))


def _simulate_day_flex(day_nurse, day_medic, day_dual, day_senior,
                        night_nurse, night_medic, night_dual, night_senior,
                        min_night_crews):
    """
    Iteratively flex one night staffer's body (by plain nurse/medic headcount —
    not touching dual-credential or senior status) over to day, as long as it:
      - actually raises day's max achievable crews (_max_possible_shifts),
      - day hasn't already reached 7, and
      - the resulting night max achievable crews would stay >= min_night_crews.

    A flex is "free" (cost 0) when night has an unmatched/leftover body of that
    role on hand — pulling it doesn't change night's own max achievable crews
    at all, since it was never part of a crew pairing to begin with. Only when
    night has no leftover of that role does a flex cost a full night crew.

    Returns (simulated_day_max, simulated_night_max, night_crew_sacrificed,
    sim_day_nurse, sim_day_medic, sim_night_nurse, sim_night_medic) — the trailing
    four are the post-flex body counts (Staffing Rebalance needs them to compute
    what's still short; the chart ignores them).
    """
    day_max = _max_possible_shifts(day_nurse, day_medic, day_dual, day_senior)
    night_max = _max_possible_shifts(night_nurse, night_medic, night_dual, night_senior)
    if day_max >= 7:
        return day_max, night_max, False, day_nurse, day_medic, night_nurse, night_medic

    sim_day_nurse, sim_day_medic = day_nurse, day_medic
    sim_night_nurse, sim_night_medic = night_nurse, night_medic
    cur_day_max, cur_night_max = day_max, night_max
    sacrificed = False

    while cur_day_max < 7:
        best = None
        for role in ('medic', 'nurse'):
            night_count = sim_night_medic if role == 'medic' else sim_night_nurse
            if night_count <= 0:
                continue
            trial_day_medic = sim_day_medic + 1 if role == 'medic' else sim_day_medic
            trial_day_nurse = sim_day_nurse + 1 if role == 'nurse' else sim_day_nurse
            trial_day_max = _max_possible_shifts(trial_day_nurse, trial_day_medic, day_dual, day_senior)
            if trial_day_max <= cur_day_max:
                continue  # this flex wouldn't actually help day

            trial_night_medic = sim_night_medic - 1 if role == 'medic' else sim_night_medic
            trial_night_nurse = sim_night_nurse - 1 if role == 'nurse' else sim_night_nurse
            trial_night_max = _max_possible_shifts(trial_night_nurse, trial_night_medic, night_dual, night_senior)
            cost = cur_night_max - trial_night_max

            if cost > 0 and trial_night_max < min_night_crews:
                continue  # would drop night below the floor — not allowed

            # Prefer free flexes over costly ones; among equal cost, prefer the bigger day gain.
            key = (cost, -trial_day_max)
            payload = (trial_day_max, trial_night_max, cost,
                       trial_day_nurse, trial_day_medic, trial_night_nurse, trial_night_medic)
            if best is None or key < best[0]:
                best = (key, payload)

        if best is None:
            break  # no further eligible flex
        (_, (trial_day_max, trial_night_max, cost,
             sim_day_nurse, sim_day_medic, sim_night_nurse, sim_night_medic)) = best
        if cost > 0:
            sacrificed = True
        cur_day_max, cur_night_max = trial_day_max, trial_night_max

    return (cur_day_max, cur_night_max, sacrificed,
            sim_day_nurse, sim_day_medic, sim_night_nurse, sim_night_medic)


def _compute_bid_day_stats(days, bids, role_mapping, no_matrix_mapping):
    """One row per bid day with Nurse/Medic/Dual/Senior counts and Max Shifts, Day and Night."""
    resolved = [(_bid_role_and_senior(b, role_mapping, no_matrix_mapping), b) for b in bids]

    rows = []
    for i, day in enumerate(days, start=1):
        counts = {'day_label': day, 'day_index': i, 'weekday': day.split(' ')[0]}
        for period, code in (('day', 'D'), ('night', 'N')):
            nurse = medic = dual = senior = 0
            for (role, is_senior), b in resolved:
                if (b['track_data'] or {}).get(day) != code:
                    continue
                if role == 'medic':
                    medic += 1
                else:
                    nurse += 1
                    if role == 'dual':
                        dual += 1
                if is_senior:
                    senior += 1
            counts[f'{period}_nurse'] = nurse
            counts[f'{period}_medic'] = medic
            counts[f'{period}_dual'] = dual
            counts[f'{period}_senior'] = senior
            counts[f'{period}_max_shifts'] = _max_possible_shifts(nurse, medic, dual, senior)
        rows.append(counts)
    return pd.DataFrame(rows)


def _build_bid_heatmap(days, bids, role_mapping, no_matrix_mapping):
    """Staff x day grid (nurses A-Z, then medics A-Z) colored by submitted shift code."""
    role_of = {}
    for b in bids:
        role, _ = _bid_role_and_senior(b, role_mapping, no_matrix_mapping)
        role_of[b['staff_name']] = role
    staff_order = sorted(role_of.keys(), key=lambda n: (0 if role_of[n] != 'medic' else 1, n))

    rows = []
    for b in bids:
        name = b['staff_name']
        td = b['track_data'] or {}
        for day in days:
            code = td.get(day) or ''
            code = code if code in ('D', 'N', 'AT') else 'Off'
            rows.append({'staff': name, 'role': role_of[name], 'day_label': day, 'shift': code})
    df = pd.DataFrame(rows)

    color_scale = alt.Scale(domain=list(_SHIFT_COLORS.keys()), range=list(_SHIFT_COLORS.values()))
    return alt.Chart(df).mark_rect().encode(
        x=alt.X('day_label:N', sort=days, title=None,
                axis=alt.Axis(labelAngle=-90, labelFontSize=8, labelOverlap=False)),
        y=alt.Y('staff:N', sort=staff_order, title=None, axis=alt.Axis(labelFontSize=9)),
        color=alt.Color('shift:N', scale=color_scale, legend=alt.Legend(title='Shift')),
        tooltip=[alt.Tooltip('staff:N', title='Staff'), alt.Tooltip('role:N', title='Role'),
                 alt.Tooltip('day_label:N', title='Day'), alt.Tooltip('shift:N', title='Shift')],
    ).properties(height=max(300, 15 * len(staff_order)))


def _build_composition_chart(day_stats, period):
    """Stacked bar of Nurse(non-dual)/Dual/Medic bid counts across the 42 days, one shift period."""
    prefix = 'day_' if period == 'Day' else 'night_'
    df = day_stats[['day_label', f'{prefix}nurse', f'{prefix}dual', f'{prefix}medic']].copy()
    df['Nurse'] = df[f'{prefix}nurse'] - df[f'{prefix}dual']
    df['Dual'] = df[f'{prefix}dual']
    df['Medic'] = df[f'{prefix}medic']
    long_df = df.melt(id_vars=['day_label'], value_vars=['Nurse', 'Dual', 'Medic'],
                       var_name='Category', value_name='Count')

    color_scale = alt.Scale(domain=list(_ROLE_COLORS.keys()), range=list(_ROLE_COLORS.values()))
    order = day_stats['day_label'].tolist()
    return alt.Chart(long_df).mark_bar().encode(
        x=alt.X('day_label:N', sort=order, title=None,
                axis=alt.Axis(labelAngle=-90, labelFontSize=8, labelOverlap=False)),
        y=alt.Y('Count:Q', title='Staff bidding'),
        color=alt.Color('Category:N', scale=color_scale, legend=alt.Legend(title=None)),
        order=alt.Order('Category:N'),
        tooltip=['day_label:N', 'Category:N', 'Count:Q'],
    ).properties(title=f'{period} Shift — Bid Composition', height=220)


def _build_demand_vs_cap_chart(day_stats, period, role):
    """One role's bid count vs. its configured cap, across the 42 days."""
    prefix = 'day_' if period == 'Day' else 'night_'
    cap_prefix = 'day_cap_' if period == 'Day' else 'night_cap_'
    field, cap_field = f'{prefix}{role.lower()}', f'{cap_prefix}{role.lower()}'
    hue = _ROLE_COLORS[role]
    order = day_stats['day_label'].tolist()
    df = day_stats[['day_label', field, cap_field]].rename(columns={field: 'Bids', cap_field: 'Cap'})

    bars = alt.Chart(df).mark_bar(color=hue).encode(
        x=alt.X('day_label:N', sort=order, title=None,
                axis=alt.Axis(labelAngle=-90, labelFontSize=7, labelOverlap=True)),
        y=alt.Y('Bids:Q', title='Staff bidding'),
        tooltip=[alt.Tooltip('day_label:N', title='Day'), alt.Tooltip('Bids:Q', title='Bids'),
                 alt.Tooltip('Cap:Q', title='Cap')],
    )
    cap_line = alt.Chart(df).mark_line(strokeDash=[4, 3], strokeWidth=2, color=hue).encode(
        x=alt.X('day_label:N', sort=order), y='Cap:Q',
    )
    return (bars + cap_line).properties(title=f'{period} · {role} (dashed = cap)', height=180)


def _split_day_label(day_label):
    """"Sun A 1" -> ("Sun", "A1") — same letter+number join the block-table
    headers already use, kept apart from the weekday by one space."""
    parts = day_label.split()
    if len(parts) == 3:
        return parts[0], f"{parts[1]}{parts[2]}"
    return day_label, ""


# Vega-Lite has no declarative hatch-fill, so the flex-simulation's Night
# "capacity given up" gap (see _build_max_shifts_chart) references a literal
# url(#hatchNightBorrow) fill — this injects the matching SVG <pattern>
# definition into the page. SVG patterns resolve by ID lookup anywhere in the
# document, so this can live in its own tiny invisible <svg> rather than
# needing to reach into the chart's own generated markup.
_DIAGONAL_HATCH_PATTERN_HTML = """
<svg width="0" height="0" style="position:absolute">
  <defs>
    <pattern id="hatchNightBorrow" patternUnits="userSpaceOnUse" width="6" height="6" patternTransform="rotate(45)">
      <rect width="6" height="6" fill="white"></rect>
      <line x1="0" y1="0" x2="0" y2="6" stroke="#6fa8dc" stroke-width="2"></line>
    </pattern>
  </defs>
</svg>
"""


def _build_max_shifts_chart(day_stats, simulate=False, min_night_crews=4):
    """
    Max achievable Day/Night crews (see _max_possible_shifts) across the 42
    days, as a diverging bar chart — Day bars above the zero line, Night bars
    below. Solid lines mark the Day (7) and Night (min_night_crews) minimums;
    dashed lines mark the secondary Day references (5, 9). Day labels are a
    two-row weekday/tag strip (see _split_day_label) kept below the chart on
    a shared x-scale, since Vega-Lite axis labels can't reliably line-break.

    When simulate=True, days below the Day minimum get a solid blue extension
    (see _simulate_day_flex) — Night's own color — showing how far the day
    count could climb by flexing Night staff over. When a flex actually costs
    a full Night crew, the Night bar itself is drawn at its new, smaller
    height, and the gap it gave up (old max down to new max) is filled with a
    white/light-blue hatch rather than solid blue, so it reads as "no longer
    there" instead of double-counting.
    """
    df = day_stats.copy()
    split = df['day_label'].map(_split_day_label)
    df['weekday'] = split.map(lambda p: p[0])
    df['tag'] = split.map(lambda p: p[1])
    df['day_label_display'] = df['weekday'] + ' ' + df['tag']
    order = df['day_label_display'].tolist()

    if simulate:
        sim = df.apply(lambda r: _simulate_day_flex(
            r.day_nurse, r.day_medic, r.day_dual, r.day_senior,
            r.night_nurse, r.night_medic, r.night_dual, r.night_senior,
            min_night_crews), axis=1)
        df['sim_day_max'] = [s[0] for s in sim]
        df['sim_night_max'] = [s[1] for s in sim]
        df['sacrificed'] = [s[2] for s in sim]

    long_df = df.melt(id_vars=['day_label_display'], value_vars=['day_max_shifts', 'night_max_shifts'],
                       var_name='Period', value_name='Max Crews')
    long_df['Period'] = long_df['Period'].map({'day_max_shifts': 'Day', 'night_max_shifts': 'Night'})
    long_df['plot_value'] = long_df.apply(
        lambda r: r['Max Crews'] if r['Period'] == 'Day' else -r['Max Crews'], axis=1)

    if simulate:
        # A sacrificed flex lowers Night's own max, so the main Night bar is
        # drawn at that new, smaller height — the hatch gap below fills in
        # what it gave up, instead of the bar overstating what Night can
        # still staff.
        sac_by_day = df.set_index('day_label_display')['sacrificed']
        sim_night_by_day = df.set_index('day_label_display')['sim_night_max']
        sacrificed_rows = (long_df['Period'] == 'Night') & long_df['day_label_display'].map(sac_by_day)
        long_df.loc[sacrificed_rows, 'plot_value'] = \
            -long_df.loc[sacrificed_rows, 'day_label_display'].map(sim_night_by_day)

    color_scale = alt.Scale(domain=list(_PERIOD_COLORS.keys()), range=list(_PERIOD_COLORS.values()))
    shared_x = alt.X('day_label_display:N', sort=order, title=None, axis=None)

    bars = alt.Chart(long_df).mark_bar().encode(
        x=shared_x,
        y=alt.Y('plot_value:Q', title='Max Crews  (Day above · Night below)',
                axis=alt.Axis(labelExpr='abs(datum.value)', grid=False)),
        color=alt.Color('Period:N', scale=color_scale, legend=alt.Legend(title=None)),
        tooltip=[alt.Tooltip('day_label_display:N', title='Day'),
                 alt.Tooltip('Period:N', title='Period'),
                 alt.Tooltip('Max Crews:Q', title='Max Crews')],
    )

    zero_line = alt.Chart(pd.DataFrame({'y': [0]})).mark_rule(strokeWidth=1.5, color='#333').encode(y='y:Q')
    day_refs_soft = alt.Chart(pd.DataFrame({'y': [5, 9]})).mark_rule(
        strokeDash=[4, 3], strokeWidth=1, color=_PERIOD_COLORS['Day'], opacity=0.6).encode(y='y:Q')
    day_ref_min = alt.Chart(pd.DataFrame({'y': [7]})).mark_rule(
        strokeWidth=2, color=_PERIOD_COLORS['Day'], opacity=0.95).encode(y='y:Q')
    night_ref_min = alt.Chart(pd.DataFrame({'y': [-min_night_crews]})).mark_rule(
        strokeWidth=2, color=_PERIOD_COLORS['Night'], opacity=0.95).encode(y='y:Q')

    layers = [bars, zero_line, day_refs_soft, day_ref_min, night_ref_min]

    if simulate:
        ext_df = df[df['sim_day_max'] > df['day_max_shifts']].copy()
        if not ext_df.empty:
            ext_df['y0'] = ext_df['day_max_shifts']
            ext_df['y1'] = ext_df['sim_day_max']
            day_extension = alt.Chart(ext_df).mark_bar(
                fill=_PERIOD_COLORS['Night'], stroke=_PERIOD_LINE_COLORS['Night'], strokeWidth=1.2,
            ).encode(x=shared_x, y='y0:Q', y2='y1:Q',
                      tooltip=[alt.Tooltip('day_label_display:N', title='Day'),
                               alt.Tooltip('day_max_shifts:Q', title='Actual max'),
                               alt.Tooltip('sim_day_max:Q', title='Simulated max')])
            layers.append(day_extension)

        sac_df = df[df['sacrificed']].copy()
        if not sac_df.empty:
            sac_df['y0'] = -sac_df['night_max_shifts']
            sac_df['y1'] = -sac_df['sim_night_max']
            night_borrow_gap = alt.Chart(sac_df).mark_bar(
                fill='url(#hatchNightBorrow)', stroke=_PERIOD_LINE_COLORS['Night'], strokeWidth=1.2,
            ).encode(x=shared_x, y='y0:Q', y2='y1:Q',
                      tooltip=[alt.Tooltip('day_label_display:N', title='Day'),
                               alt.Tooltip('night_max_shifts:Q', title='Actual night max'),
                               alt.Tooltip('sim_night_max:Q', title='Simulated night max')])
            layers.append(night_borrow_gap)

    main = alt.layer(*layers).properties(height=340)

    weekday_row = alt.Chart(df).mark_text(fontSize=11, fontWeight='bold', dy=0).encode(x=shared_x, text='weekday:N')
    tag_row = alt.Chart(df).mark_text(fontSize=11, dy=16).encode(x=shared_x, text='tag:N')
    label_strip = (weekday_row + tag_row).properties(height=40)

    return alt.vconcat(main, label_strip, spacing=2).resolve_scale(x='shared')


def _build_bid_summary_table(day_stats):
    """Wide Max Shifts/Senior/Nurse/Dual/Medic x Day/Night table, days as columns (Excel-replica)."""
    idx = day_stats.set_index('day_label')
    row_order = [
        ('Max DAY Shifts', idx['day_max_shifts']),
        ('Day — Senior', idx['day_senior']),
        ('Day — Nurse', idx['day_nurse']),
        ('Day — Dual (counts as RN)', idx['day_dual']),
        ('Day — Medic', idx['day_medic']),
        ('Max NIGHT Shifts', idx['night_max_shifts']),
        ('Night — Senior', idx['night_senior']),
        ('Night — Nurse', idx['night_nurse']),
        ('Night — Dual (counts as RN)', idx['night_dual']),
        ('Night — Medic', idx['night_medic']),
    ]
    table = pd.DataFrame({label: series for label, series in row_order}).T
    table = table.reindex(columns=day_stats['day_label'].tolist())
    table.index.name = None
    return table


def _render_bid_analysis_tab(config_names, default_track_index):
    """Visual + tabular breakdown of submitted bids across the 42-day cycle, for tuning bid caps."""
    st.markdown("### Bid Analysis")

    if not config_names:
        st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        return

    analysis_track = st.selectbox(
        "Track Cycle:", config_names, index=default_track_index, key="analysis_track_select")

    ctx, roster_error = _load_bidding_data_files()
    if ctx is None:
        st.error(roster_error)
        return

    ok, bids_raw = get_all_bid_tracks(analysis_track)
    bids = bids_raw if ok else []
    if not bids:
        st.info(f"No bids submitted yet for {analysis_track}.")
        return

    days = ctx['days']
    role_mapping = ctx['role_mapping']
    no_matrix_mapping = ctx['no_matrix_mapping']
    staff_names = ctx['staff_names']

    submitted_names = {b['staff_name'] for b in bids}
    missing = sorted(n for n in staff_names if n not in submitted_names)
    m1, m2 = st.columns(2)
    m1.metric("Bids Submitted", f"{len(bids)} / {len(staff_names)}")
    m2.metric("Roster Staff Missing a Bid", len(missing))
    if missing:
        with st.expander(f"{len(missing)} staff without a submitted bid"):
            st.write(", ".join(missing))
    st.caption("Figures below reflect submitted bids only, and will shift as more staff submit.")

    day_stats = _compute_bid_day_stats(days, bids, role_mapping, no_matrix_mapping)
    weekday_caps = get_track_capacity_by_weekday(analysis_track)
    for period, cap_key_prefix in (('day', 'max_day_'), ('night', 'max_night_')):
        for role in ('nurse', 'medic'):
            # get_track_capacity_by_weekday() keys are plural: max_day_nurses/max_day_medics/...
            day_stats[f'{period}_cap_{role}'] = day_stats['weekday'].map(
                lambda w: weekday_caps.get(w, {}).get(f'{cap_key_prefix}{role}s', 0))

    st.markdown("#### Maximum Achievable Crews per Day")
    st.caption("The most complete Nurse+Medic crews that day's bidders could staff — letting dual-credentialed "
               "staff flex to whichever side is short, capped by how many no-matrix/senior staff bid that day. "
               "Solid lines mark the Day (7) and Night minimums; dashed lines are secondary Day references (5, 9).")
    simulate_flex = st.checkbox("Simulate N to D Flex?", value=False, key="bid_analysis_simulate_flex")
    min_night_crews = 4
    if simulate_flex:
        min_night_crews = st.selectbox(
            "Minimum Night Crews:", list(range(10)), index=4, key="bid_analysis_min_night_crews")
        st.caption("Solid blue extensions show how far a below-minimum Day count could climb by flexing "
                   "Night staff over, without dropping Night below this minimum. On any day where a flex "
                   "actually costs a full night crew, Night is redrawn at its new level and the white/"
                   "light-blue hatch marks the crew it gave up.")
        st.markdown(_DIAGONAL_HATCH_PATTERN_HTML, unsafe_allow_html=True)
    # Split into 14-day blocks like Bid Roster/Base Analysis — at the full 42-day
    # width the two-row weekday/tag labels don't have enough room per column and
    # start overlapping.
    for block_letter, block_start in (('A', 0), ('B', 14), ('C', 28)):
        st.markdown(f"##### Block {block_letter}")
        block_day_stats = day_stats.iloc[block_start:block_start + 14]
        st.altair_chart(_build_max_shifts_chart(block_day_stats, simulate_flex, min_night_crews),
                         use_container_width=True)

    st.markdown("#### Bid Composition by Day")
    st.altair_chart(_build_composition_chart(day_stats, 'Day'), use_container_width=True)
    st.altair_chart(_build_composition_chart(day_stats, 'Night'), use_container_width=True)

    st.markdown("#### Bid Demand vs. Configured Cap")
    st.caption("Solid bars are submitted bids; the dashed line is the current cap. "
               "A bar above the dashed line means more staff bid than there's room for.")
    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.altair_chart(_build_demand_vs_cap_chart(day_stats, 'Day', 'Nurse'), use_container_width=True)
        st.altair_chart(_build_demand_vs_cap_chart(day_stats, 'Night', 'Nurse'), use_container_width=True)
    with dcol2:
        st.altair_chart(_build_demand_vs_cap_chart(day_stats, 'Day', 'Medic'), use_container_width=True)
        st.altair_chart(_build_demand_vs_cap_chart(day_stats, 'Night', 'Medic'), use_container_width=True)

    with st.expander("Full Day/Night breakdown table (Max Shifts / Senior / Nurse / Dual / Medic)"):
        summary_table = _build_bid_summary_table(day_stats)
        st.dataframe(summary_table, use_container_width=True)

        from openpyxl.utils import get_column_letter
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            sheet_name = 'Day-Night Breakdown'
            summary_table.to_excel(writer, sheet_name=sheet_name, index=True)
            ws = writer.sheets[sheet_name]
            ws.column_dimensions['A'].width = 26
            for col_idx in range(2, len(summary_table.columns) + 2):
                ws.column_dimensions[get_column_letter(col_idx)].width = 10
        excel_buffer.seek(0)

        safe_track_name = ''.join(c if c.isalnum() else '_' for c in analysis_track)
        st.download_button(
            "📥 Download as Excel",
            data=excel_buffer,
            file_name=f"{safe_track_name}_bid_breakdown_{datetime.now(_eastern_tz).strftime('%Y%m%d%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_bid_breakdown_{analysis_track}",
        )

    with st.expander("Where Staff Are Bidding", expanded=False):
        st.caption("One row per staff member (nurses A–Z, then medics A–Z), one column per bid day.")
        st.altair_chart(_build_bid_heatmap(days, bids, role_mapping, no_matrix_mapping), use_container_width=True)


def _compute_bid_roster_table(analysis_track, ctx, bids):
    """
    One row per submitted bid: Staff/Role/Seniority, then one column per bid day
    holding "D (BASE)" / "N (BASE)" / "AT" / "" (empty for non-working days).

    Expected base reuses calculate_hypothetical_assignment — the same seniority
    competition simulation shown to a bidder for their own hypothetical schedule —
    scored against this cycle's submitted bids (bid_track_name=analysis_track), so
    the base shown here always matches what the bidder saw when they picked that day.
    """
    from modules.hypothetical_scheduler_new import calculate_hypothetical_assignment, _load_all_base_preferences
    from modules.db_utils import get_base_shift_counts

    days = ctx['days']
    all_base_prefs = _load_all_base_preferences()
    base_shift_counts = get_base_shift_counts(analysis_track)

    rows = []
    for b in bids:
        name = b['staff_name']
        track_data = b['track_data'] or {}
        raw_role, _ = _bid_role_and_senior(b, ctx['role_mapping'], ctx['no_matrix_mapping'])
        row = {
            'Staff': name,
            'Role': raw_role.title(),
            '_role_bucket': _bidding_role_bucket(raw_role),
            'Seniority': ctx['seniority_mapping'].get(name),
        }
        for day in days:
            code = track_data.get(day)
            if code in ('D', 'N'):
                result = calculate_hypothetical_assignment(
                    name, day, 'day' if code == 'D' else 'night',
                    ctx['preferences_df'], ctx['current_tracks_df'],
                    ctx['staff_col_prefs'], ctx['staff_col_tracks'], ctx['role_col'], ctx['seniority_col'],
                    use_database_logic=True, all_base_prefs=all_base_prefs,
                    bid_track_name=analysis_track, base_shift_counts=base_shift_counts,
                )
                base = result.get('assignment')
                row[day] = f"{code} ({base})" if base else code
            elif code == 'AT':
                row[day] = 'AT'
            else:
                row[day] = ''
        rows.append(row)
    return pd.DataFrame(rows)


def _seniority_sort_key(series):
    """Sort key for a 'Seniority' column: numeric ascending (most senior first), missing/non-numeric last."""
    def conv(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return float('inf')
    return series.map(conv)


def _render_bid_roster_block_table(df, block_days):
    """
    Fixed-width, non-draggable HTML table for one block of the Bid Roster: Staff/Role
    plus one column per day, every day column pinned to the same width regardless of
    content — st.dataframe's grid lets a user drag-reorder columns and auto-sizes them
    per content, which would let the day order drift and the widths go uneven.
    """
    # Fixed pixel widths, sized to the longest real value at this font (11px Source
    # Sans) plus padding/border — "Hanley-McCarthy" is the longest staff name on file,
    # "Medic"/"Nurse" the longest roles. Day <th>/<td> below deliberately carry no
    # width: table-layout:fixed auto-divides any remaining columns equally once the
    # sized ones are pinned, which is what actually gives the day columns the room
    # freed up by narrowing Staff/Role, at any table/container width.
    label_cols = [('Staff', 90), ('Role', 38)]

    weekday_header = ""
    tag_header = ""
    for day in block_days:
        parts = day.split()
        weekday = parts[0] if parts else day
        tag = f"{parts[1]}{parts[2]}" if len(parts) == 3 else ""
        weekday_header += (
            f'<th style="box-sizing:border-box; border:1px solid #ddd; '
            f'background-color:#f0f2f6; font-size:10px; font-weight:400; color:#666; '
            f'text-align:center; padding:2px 0; white-space:nowrap;">{weekday}</th>'
        )
        tag_header += (
            f'<th style="box-sizing:border-box; border:1px solid #ddd; '
            f'background-color:#f0f2f6; font-size:9px; font-weight:400; color:#666; '
            f'text-align:center; padding:2px 0; white-space:nowrap;">{tag}</th>'
        )

    label_th = "".join(
        f'<th rowspan="2" style="width:{w}px; box-sizing:border-box; '
        f'border:1px solid #ddd; background-color:#f0f2f6; font-size:10px; font-weight:500; '
        f'text-align:center; padding:2px;">{name}</th>'
        for name, w in label_cols
    )

    rows_html = ""
    for _, row in df.iterrows():
        label_tds = "".join(
            f'<td style="width:{w}px; box-sizing:border-box; border:1px solid #ddd; '
            f'font-size:11px; text-align:center; padding:3px 2px; white-space:nowrap; '
            f'overflow:hidden; text-overflow:ellipsis;">{row[name]}</td>'
            for name, w in label_cols
        )
        day_tds = ""
        for day in block_days:
            val = row.get(day, "")
            val = "" if pd.isna(val) else str(val)
            style = (
                "box-sizing:border-box; border:1px solid #ddd; font-size:10px; "
                "text-align:center; padding:3px 1px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;"
            )
            if val.startswith("D"):
                style += " background-color:#d4edda;"
            elif val.startswith("N"):
                style += " background-color:#cce5ff;"
            elif val == "AT":
                style += " background-color:#e2e3e5; font-weight:bold;"
            day_tds += f'<td style="{style}">{val}</td>'
        rows_html += f"<tr>{label_tds}{day_tds}</tr>"

    st.markdown(f"""
    <table style="width:100%; border-collapse:collapse; table-layout:fixed;">
        <thead>
            <tr>{label_th}{weekday_header}</tr>
            <tr>{tag_header}</tr>
        </thead>
        <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)


def _build_bid_roster_excel(table, days):
    """
    Bid Roster as a styled .xlsx: Staff/Role plus all 42 days side by side in
    one continuous sheet (rather than the on-screen separate Block A/B/C
    tables), with a merged header row naming each block's 14-day span so the
    blocks stay visually distinguishable in the combined view.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Roster"

    header_fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
    day_fills = {
        'D': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'N': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        'AT': PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
    }
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    blocks = [("Block A", days[0:14]), ("Block B", days[14:28]), ("Block C", days[28:42])]

    col = 3
    for block_name, block_days in blocks:
        start_col, end_col = col, col + len(block_days) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.border = border
        top_left = ws.cell(row=1, column=start_col, value=block_name)
        top_left.font = Font(bold=True)
        top_left.alignment = center
        col = end_col + 1

    for header_col, label in ((1, 'Staff'), (2, 'Role')):
        ws.merge_cells(start_row=1, start_column=header_col, end_row=3, end_column=header_col)
        cell = ws.cell(row=1, column=header_col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    col = 3
    for day in days:
        weekday, tag = _split_day_label(day)
        for row_idx, text in ((2, weekday), (3, tag)):
            cell = ws.cell(row=row_idx, column=col, value=text)
            cell.font = Font(size=9, color="666666")
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border
        col += 1

    row_idx = 4
    for _, row in table.iterrows():
        staff_cell = ws.cell(row=row_idx, column=1, value=row['Staff'])
        staff_cell.border = border
        role_cell = ws.cell(row=row_idx, column=2, value=row['Role'])
        role_cell.alignment = center
        role_cell.border = border
        col = 3
        for day in days:
            val = row.get(day, "")
            val = "" if pd.isna(val) else str(val)
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = border
            cell.font = Font(size=9, bold=(val == "AT"))
            if val.startswith("D"):
                cell.fill = day_fills['D']
            elif val.startswith("N"):
                cell.fill = day_fills['N']
            elif val == "AT":
                cell.fill = day_fills['AT']
            col += 1
        row_idx += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    for c in range(3, 3 + len(days)):
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.freeze_panes = "C4"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _render_bid_roster_tab(config_names, default_track_index):
    """Wide staff x day roster of submitted bids, with shift + expected base per day."""
    st.markdown("### Bid Roster")
    st.caption("One row per submitted bid, one column per bid day. D/N cells show the expected base "
               "computed the same way the bidder saw it; AT cells have no base; non-working days are blank.")

    if not config_names:
        st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        return

    roster_track = st.selectbox(
        "Track Cycle:", config_names, index=default_track_index, key="roster_track_select")

    ctx, roster_error = _load_bidding_data_files()
    if ctx is None:
        st.error(roster_error)
        return

    ok, bids_raw = get_all_bid_tracks(roster_track)
    bids = bids_raw if ok else []
    if not bids:
        st.info(f"No bids submitted yet for {roster_track}.")
        return

    view = st.radio(
        "View:", ["All Staff", "Split by Role (Nurse / Medic)"],
        horizontal=True, key="roster_view_mode")
    sort_by = st.selectbox("Sort by:", ["Seniority", "Staff Name"], key="roster_sort_by")

    def _sort_table(df):
        if sort_by == "Staff Name":
            return df.sort_values(by='Staff')
        return df.sort_values(by='Seniority', key=_seniority_sort_key)

    role_mapping = ctx['role_mapping']
    staff_options = sorted({b['staff_name'] for b in bids})
    nurse_names = sorted(n for n in staff_options if _bidding_role_bucket(role_mapping.get(n, '')) == 'nurse')
    medic_names = sorted(n for n in staff_options if _bidding_role_bucket(role_mapping.get(n, '')) == 'medic')

    if "roster_staff_filter" not in st.session_state:
        st.session_state["roster_staff_filter"] = staff_options

    selected_staff = st.session_state["roster_staff_filter"]
    with st.expander("Staff Filter", expanded=False):
        quick_col1, quick_col2, quick_col3 = st.columns(3)
        with quick_col1:
            if st.button("All Staff", key="roster_filter_all_btn", use_container_width=True):
                st.session_state["roster_staff_filter"] = staff_options
        with quick_col2:
            if st.button("All Nurses", key="roster_filter_nurses_btn", use_container_width=True):
                st.session_state["roster_staff_filter"] = nurse_names
        with quick_col3:
            if st.button("All Medics", key="roster_filter_medics_btn", use_container_width=True):
                st.session_state["roster_staff_filter"] = medic_names

        selected_staff = st.multiselect("Staff:", staff_options, key="roster_staff_filter")

    with st.spinner("Computing expected base assignments..."):
        table = _compute_bid_roster_table(roster_track, ctx, bids)

    table = table[table['Staff'].isin(selected_staff)]
    if table.empty:
        st.info("No staff match the current filter.")
        return

    days = ctx['days']
    blocks = [("A", days[0:14]), ("B", days[14:28]), ("C", days[28:42])]
    export_table = _sort_table(table)

    if view == "All Staff":
        for block_letter, block_days in blocks:
            st.markdown(f"#### Block {block_letter}")
            _render_bid_roster_block_table(export_table, block_days)
    else:
        for bucket_label, bucket in [("Nurses", "nurse"), ("Medics", "medic")]:
            st.markdown(f"### {bucket_label}")
            sub = _sort_table(table[table['_role_bucket'] == bucket])
            if sub.empty:
                st.caption("No bids yet.")
                continue
            for block_letter, block_days in blocks:
                st.markdown(f"#### Block {block_letter}")
                _render_bid_roster_block_table(sub, block_days)

    st.download_button(
        "⬇️ Download to Excel",
        data=_build_bid_roster_excel(export_table, days),
        file_name=f"Bid_Roster_{roster_track}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="bid_roster_download_excel",
    )


# ──────────────────────────────────────────────
# Base Analysis: per-base, per-slot fill status across the 42-day cycle
# ──────────────────────────────────────────────

_BASE_FULL_NAMES = {
    'KBED': 'Bedford',
    'KPYM': 'Plymouth',
    'KLWM': 'Lawrence',
    'KMHT': 'Manchester',
    '1B9': 'Mansfield',
}
_ALL_BASE_CODES = ['KMHT', 'KLWM', 'KBED', '1B9', 'KPYM']


def _compute_base_analysis_table(analysis_track, ctx, role_bucket):
    """
    One row per individual base/shift-type slot (e.g. "Bedford (KBED) Day #1"),
    one column per bid day, showing who currently wins that slot in the seniority
    competition — same simulate_full_shift_roster() used elsewhere, run once per
    (day, shift type) and reused across every slot at that shift type.
    """
    from modules.hypothetical_scheduler_new import (
        simulate_full_shift_roster, _load_all_base_preferences, _build_available_slots
    )
    from modules.db_utils import get_base_shift_counts

    days = ctx['days']
    all_base_prefs = _load_all_base_preferences()
    base_shift_counts = get_base_shift_counts(analysis_track)

    sim_cache = {}
    for shift_type in ('day', 'night'):
        for day in days:
            sim_cache[(day, shift_type)] = simulate_full_shift_roster(
                day, shift_type, role_bucket, ctx['preferences_df'], ctx['staff_col_prefs'],
                ctx['role_col'], ctx['seniority_col'], all_base_prefs=all_base_prefs,
                bid_track_name=analysis_track, base_shift_counts=base_shift_counts,
            )

    rows = []
    for shift_type, shift_label in (('day', 'Day'), ('night', 'Night')):
        available_shifts, shift_to_base = _build_available_slots(shift_type, base_shift_counts)
        for slot in available_shifts:
            base = shift_to_base[slot]
            slot_num = slot.split('#')[1]
            row = {
                'Base': f"{_BASE_FULL_NAMES.get(base, base)} ({base})",
                '_base_code': base,
                'Shift': shift_label,
                'Slot': f"#{slot_num}",
            }
            for day in days:
                winner = sim_cache[(day, shift_type)].get(slot)
                row[day] = winner['staff'] if winner else ''
            rows.append(row)
    return pd.DataFrame(rows)


def _compute_unassigned_rows(role_table, bids, role_mapping, no_matrix_mapping, role_bucket, role_label, days):
    """
    Two extra rows (Day/Night) for one role, not tied to any base: everyone
    who bid that shift type but the seniority competition never gave a slot —
    role_table already has who won each slot, so this just inverts that
    against who actually bid. '_base_code' is a sentinel that doesn't match
    any real base code, so _render_base_analysis_block_table's group-border
    logic naturally draws the same double-line separator ahead of these rows
    that it draws between any two different bases.
    """
    rows = []
    for shift_type, shift_label in (('day', 'Day'), ('night', 'Night')):
        code = 'D' if shift_type == 'day' else 'N'
        shift_rows = role_table[role_table['Shift'] == shift_label]
        row = {
            'Base': 'Unassigned', '_base_code': 'UNASSIGNED', 'Role': role_label,
            'Shift': shift_label, 'Slot': '—',
        }
        for day in days:
            winners = set(shift_rows[day]) - {''}
            bidders = set()
            for b in bids:
                if (b['track_data'] or {}).get(day) != code:
                    continue
                raw_role, _ = _bid_role_and_senior(b, role_mapping, no_matrix_mapping)
                if _bidding_role_bucket(raw_role) == role_bucket:
                    bidders.add(b['staff_name'])
            row[day] = ", ".join(sorted(bidders - winners))
        rows.append(row)
    return rows


def _render_base_analysis_block_table(df, block_days):
    """
    Fixed-width, non-draggable HTML table for one block of Base Analysis: Base/Role/
    Shift/Slot plus one column per day, green/red/blue fill status. df must already
    be sorted so every base's rows (and within a base, every shift's rows, and within
    a shift, every slot's rows) are contiguous — a three-tier top border marks each
    new group: a double line between bases, a thin black line between Day and Night
    within a base, and a thin gray line between numbered slots within the same base +
    shift. Night rows also get a faint blue tint on the label columns, and filled day
    cells are blue instead of green, matching the Day/Night colors used elsewhere
    (Bid Roster, track editors). Rows with _base_code == 'UNASSIGNED' (see
    _compute_unassigned_rows) render differently: cells hold a name list rather than
    a single winner, so they wrap instead of truncating, and amber/neutral fill
    replaces the green/blue/red fill status — a blank cell there is the good outcome.
    """
    # Fixed percentages (not weighted against day count) — Role/Shift/Slot only ever
    # hold a couple characters ("Medic", "Night", "#1") so they don't need much room,
    # which leaves most of the table for the day columns.
    label_cols = [('Base', 11), ('Role', 4), ('Shift', 4), ('Slot', 2.5)]
    label_pct_total = sum(w for _, w in label_cols)
    day_pct = (100 - label_pct_total) / len(block_days)

    weekday_header = ""
    tag_header = ""
    for day in block_days:
        parts = day.split()
        weekday = parts[0] if parts else day
        tag = f"{parts[1]}{parts[2]}" if len(parts) == 3 else ""
        weekday_header += (
            f'<th style="width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; '
            f'background-color:#f0f2f6; font-size:10px; font-weight:400; color:#666; '
            f'text-align:center; padding:2px 0; white-space:nowrap;">{weekday}</th>'
        )
        tag_header += (
            f'<th style="width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; '
            f'background-color:#f0f2f6; font-size:9px; font-weight:400; color:#666; '
            f'text-align:center; padding:2px 0; white-space:nowrap;">{tag}</th>'
        )

    label_th = "".join(
        f'<th rowspan="2" style="width:{w}%; box-sizing:border-box; '
        f'border:1px solid #ddd; background-color:#f0f2f6; font-size:10px; font-weight:500; '
        f'text-align:center; padding:2px;">{name}</th>'
        for name, w in label_cols
    )

    rows_html = ""
    prev_base = prev_shift = prev_slot = None
    for _, row in df.iterrows():
        is_new_base = prev_base is not None and row['_base_code'] != prev_base
        is_new_shift = not is_new_base and prev_shift is not None and row['Shift'] != prev_shift
        is_new_slot = not is_new_base and not is_new_shift and prev_slot is not None and row['Slot'] != prev_slot
        if is_new_base:
            top_border = "border-top:5px double #333;"
        elif is_new_shift:
            top_border = "border-top:1.5px solid #333;"
        elif is_new_slot:
            top_border = "border-top:1.5px solid #8a8a8a;"
        else:
            top_border = ""
        prev_base, prev_shift, prev_slot = row['_base_code'], row['Shift'], row['Slot']

        is_night = row['Shift'] == 'Night'
        is_unassigned = row['_base_code'] == 'UNASSIGNED'
        night_bg = "background-color:rgba(25, 118, 210, 0.10);" if is_night else ""

        label_tds = "".join(
            f'<td style="width:{w}%; box-sizing:border-box; border:1px solid #ddd; '
            f'{top_border} {night_bg} font-size:10px; text-align:center; padding:3px 2px; white-space:nowrap; '
            f'overflow:hidden; text-overflow:ellipsis;">{row[name]}</td>'
            for name, w in label_cols
        )
        day_tds = ""
        for day in block_days:
            val = row.get(day, "")
            val = "" if pd.isna(val) else str(val)
            if is_unassigned:
                # Names, not a single winner, so cells wrap instead of truncating —
                # and blank here is the good outcome (everyone who bid got a slot),
                # so it isn't colored red like a genuinely unfilled slot.
                fill = "#fff3cd" if val else "#f8f9fa"
                overflow_rule = "white-space:normal; word-break:break-word;"
            else:
                fill = ("#cce5ff" if is_night else "#d4edda") if val else "#f8d7da"
                overflow_rule = "white-space:nowrap; overflow:hidden; text-overflow:ellipsis;"
            style = (
                f"width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; {top_border} "
                f"background-color:{fill}; font-size:9.5px; text-align:center; padding:3px 1px; {overflow_rule}"
            )
            day_tds += f'<td style="{style}">{val}</td>'
        rows_html += f"<tr>{label_tds}{day_tds}</tr>"

    st.markdown(f"""
    <table style="width:100%; border-collapse:collapse; table-layout:fixed;">
        <thead>
            <tr>{label_th}{weekday_header}</tr>
            <tr>{tag_header}</tr>
        </thead>
        <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)


def _build_base_analysis_excel(filtered, days):
    """
    Base Analysis as a styled .xlsx: Base/Role/Shift/Slot plus all 42 days in
    one continuous sheet, with the same merged Block A/B/C header row as the
    Bid Roster export, and the same three-tier group separators (double line
    between bases, medium line between Day/Night, thin dark line between
    slots) rendered as Excel cell borders instead of CSS. filtered must
    already be sorted the way _render_base_analysis_tab sorts it, so each
    group's rows are contiguous.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Base Analysis"

    header_fill = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
    night_tint_fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")
    fills = {
        'day_filled': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'night_filled': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        'unfilled': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        'unassigned_has': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'unassigned_none': PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"),
    }
    thin_gray = Side(style='thin', color='DDDDDD')
    base_top = Side(style='double', color='333333')
    shift_top = Side(style='medium', color='333333')
    slot_top = Side(style='thin', color='8A8A8A')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    label_cols = [('Base', 1, 22), ('Role', 2, 8), ('Shift', 3, 8), ('Slot', 4, 6)]
    n_label_cols = len(label_cols)

    blocks = [("Block A", days[0:14]), ("Block B", days[14:28]), ("Block C", days[28:42])]
    col = n_label_cols + 1
    for block_name, block_days in blocks:
        start_col, end_col = col, col + len(block_days) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        top_left = ws.cell(row=1, column=start_col, value=block_name)
        top_left.font = Font(bold=True)
        top_left.alignment = center
        col = end_col + 1

    for label, header_col, width in label_cols:
        ws.merge_cells(start_row=1, start_column=header_col, end_row=3, end_column=header_col)
        cell = ws.cell(row=1, column=header_col, value=label)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.fill = header_fill
        cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        ws.column_dimensions[get_column_letter(header_col)].width = width

    col = n_label_cols + 1
    for day in days:
        weekday, tag = _split_day_label(day)
        for row_idx, text in ((2, weekday), (3, tag)):
            cell = ws.cell(row=row_idx, column=col, value=text)
            cell.font = Font(size=9, color="666666")
            cell.alignment = center
            cell.fill = header_fill
            cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        ws.column_dimensions[get_column_letter(col)].width = 7
        col += 1

    row_idx = 4
    prev_base = prev_shift = prev_slot = None
    for _, row in filtered.iterrows():
        is_new_base = prev_base is not None and row['_base_code'] != prev_base
        is_new_shift = not is_new_base and prev_shift is not None and row['Shift'] != prev_shift
        is_new_slot = not is_new_base and not is_new_shift and prev_slot is not None and row['Slot'] != prev_slot
        top_side = base_top if is_new_base else shift_top if is_new_shift else slot_top if is_new_slot else thin_gray
        prev_base, prev_shift, prev_slot = row['_base_code'], row['Shift'], row['Slot']

        is_night = row['Shift'] == 'Night'
        is_unassigned = row['_base_code'] == 'UNASSIGNED'
        row_border = Border(left=thin_gray, right=thin_gray, top=top_side, bottom=thin_gray)

        for label, col_idx, _ in label_cols:
            cell = ws.cell(row=row_idx, column=col_idx, value=row[label])
            cell.alignment = center
            cell.border = row_border
            if is_night:
                cell.fill = night_tint_fill

        col = n_label_cols + 1
        for day in days:
            val = row.get(day, "")
            val = "" if pd.isna(val) else str(val)
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = row_border
            cell.font = Font(size=9)
            if is_unassigned:
                cell.fill = fills['unassigned_has'] if val else fills['unassigned_none']
            elif val:
                cell.fill = fills['night_filled'] if is_night else fills['day_filled']
            else:
                cell.fill = fills['unfilled']
            col += 1
        row_idx += 1

    ws.freeze_panes = f"{get_column_letter(n_label_cols + 1)}4"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _render_base_analysis_tab(config_names, default_track_index):
    """Per-base, per-slot fill-status grid: green = filled (shows who), red = still open."""
    st.markdown("### Base Analysis")
    st.caption("Every base's individual shift slots, one row each, split into Blocks A/B/C "
               "(14 days each) so it's not one long horizontal scroll. Green cells show who "
               "currently wins that slot in the seniority competition; red cells are still "
               "unfilled. Filters below apply to all three blocks.")

    if not config_names:
        st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        return

    analysis_track = st.selectbox(
        "Track Cycle:", config_names, index=default_track_index, key="base_analysis_track_select")

    ctx, roster_error = _load_bidding_data_files()
    if ctx is None:
        st.error(roster_error)
        return

    ok, bids_raw = get_all_bid_tracks(analysis_track)
    bids = bids_raw if ok else []

    role_choices = st.multiselect(
        "Role:", ["Nurses", "Medics"], default=["Nurses", "Medics"], key="base_analysis_role")
    # Nurses and medics compete for their own independent copy of every slot (separate
    # crew roles, each with their own base-slot allotment), so showing both just runs
    # the same per-role simulation twice and stacks the results with a Role column.
    role_buckets = [
        (label, bucket) for label, bucket in (("Nurse", "nurse"), ("Medic", "medic"))
        if f"{label}s" in role_choices
    ]
    if not role_buckets:
        st.info("Select at least one role.")
        return

    base_options = [f"{_BASE_FULL_NAMES.get(b, b)} ({b})" for b in _ALL_BASE_CODES]
    base_code_by_option = dict(zip(base_options, _ALL_BASE_CODES))
    selected_base_options = st.multiselect(
        "Base:", base_options, default=base_options, key="base_analysis_base_filter")
    selected_bases = {base_code_by_option[o] for o in selected_base_options}

    selected_shifts = st.multiselect(
        "Shift Type:", ["Day", "Night"], default=["Day", "Night"], key="base_analysis_shift_filter")

    show_unassigned = st.checkbox(
        "Show Unassigned (bid but won no slot)", value=False, key="base_analysis_show_unassigned")
    if show_unassigned:
        st.caption("Adds an 'Unassigned' row per role/shift: everyone who bid that shift but every slot "
                   "of it was already won by someone more senior, so they aren't tied to a specific base.")

    days = ctx['days']

    with st.spinner("Computing base fill status..."):
        role_tables = []
        unassigned_rows = []
        for role_label, role_bucket in role_buckets:
            role_table = _compute_base_analysis_table(analysis_track, ctx, role_bucket)
            role_table.insert(1, 'Role', role_label)
            role_tables.append(role_table)
            if show_unassigned:
                unassigned_rows.extend(_compute_unassigned_rows(
                    role_table, bids, ctx['role_mapping'], ctx['no_matrix_mapping'],
                    role_bucket, role_label, days))
        table = pd.concat(role_tables, ignore_index=True)

    filtered = table[table['_base_code'].isin(selected_bases) & table['Shift'].isin(selected_shifts)]

    if unassigned_rows:
        unassigned_df = pd.DataFrame(unassigned_rows)
        unassigned_df = unassigned_df[unassigned_df['Shift'].isin(selected_shifts)]
        filtered = pd.concat([filtered, unassigned_df], ignore_index=True)

    if filtered.empty:
        st.info("No base/shift combinations match the current filters.")
        return

    # Group every base's rows together (across shift/slot/role) rather than grouping
    # by role first, so Nurse and Medic rows for the same base sit next to each other.
    # 'UNASSIGNED' is listed last so those rows sort as their own group at the end,
    # after every real base.
    filtered = filtered.copy()
    filtered['_base_order'] = pd.Categorical(
        filtered['_base_code'], categories=_ALL_BASE_CODES + ['UNASSIGNED'], ordered=True)
    filtered['_shift_order'] = pd.Categorical(filtered['Shift'], categories=['Day', 'Night'], ordered=True)
    filtered['_slot_order'] = pd.to_numeric(filtered['Slot'].str.lstrip('#'), errors='coerce').fillna(9999)
    filtered['_role_order'] = pd.Categorical(filtered['Role'], categories=['Nurse', 'Medic'], ordered=True)
    filtered = filtered.sort_values(['_base_order', '_shift_order', '_slot_order', '_role_order'])

    blocks = [("A", days[0:14]), ("B", days[14:28]), ("C", days[28:42])]
    for block_letter, block_days in blocks:
        st.markdown(f"#### Block {block_letter}")
        _render_base_analysis_block_table(filtered, block_days)

    st.download_button(
        "⬇️ Download to Excel",
        data=_build_base_analysis_excel(filtered, days),
        file_name=f"Base_Analysis_{analysis_track}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="base_analysis_download_excel",
    )


# ──────────────────────────────────────────────
# Admin mode toggle (small sidebar gate) + full-page admin dashboard
# ──────────────────────────────────────────────

def _render_admin_mode_toggle():
    """Render the sidebar password gate that flips the page into full admin mode."""
    if 'track_bidding_admin_mode' not in st.session_state:
        st.session_state.track_bidding_admin_mode = False

    with st.sidebar:
        st.markdown("## Track Bidding Admin")
        password = st.text_input("Enter admin password:", type="password", key="bid_admin_pw")

        # text_input already commits (and reruns) on Enter, so once the password
        # checks out there's nothing left to confirm — no separate button click needed.
        if check_admin_access(password) and not st.session_state.track_bidding_admin_mode:
            st.session_state.track_bidding_admin_mode = True
            st.rerun()

        if st.session_state.track_bidding_admin_mode:
            st.success("✅ Admin Mode Active")
            if st.button("👤 Switch to Staff View", key="bid_exit_admin_mode", use_container_width=True):
                st.session_state.track_bidding_admin_mode = False
                st.rerun()


def display_bidding_admin_interface():
    """Full-page Track Bidding admin dashboard (mirrors the Summer Leave admin page)."""
    st.header("🔧 Track Bidding Administration")
    st.markdown("---")

    all_configs = get_all_track_configs()
    bid_cfg = get_bidding_track_config()
    config_names = [c['track_name'] for c in all_configs]
    default_track_index = (
        config_names.index(bid_cfg['track_name'])
        if bid_cfg and bid_cfg['track_name'] in config_names else 0
    )

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs([
        "📊 Overview", "🛠️ Track Configs", "👥 Manage Bid Access", "➕ Add/Remove Selection", "📈 Bid Analysis",
        "📋 Bid Roster", "🏢 Base Analysis", "⚖️ Staffing Rebalance", "🔁 Needs Swap Requests"
    ])

    # ── Tab 1: Overview ──
    with tab1:
        st.markdown("### Bid Tracks Summary")

        if not all_configs:
            st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        else:
            stats_rows = []
            for cfg in all_configs:
                tn = cfg['track_name']
                ok, bids = get_all_bid_tracks(tn)
                bid_count = len(bids) if ok else 0
                access_count = sum(1 for v in get_all_bid_access_configs(tn).values() if v)
                status = 'Active' if cfg['is_active'] else ('Bidding Open' if cfg['is_bidding_open'] else 'Inactive')
                stats_rows.append({
                    'Track': tn,
                    'Status': status,
                    'Bids Submitted': bid_count,
                    'Staff w/ Access Enabled': access_count,
                    'Max Day Nurses': cfg['max_day_nurses'],
                    'Max Day Medics': cfg['max_day_medics'],
                    'Max Night Nurses': cfg['max_night_nurses'],
                    'Max Night Medics': cfg['max_night_medics'],
                })
            st.dataframe(pd.DataFrame(stats_rows), use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### Bids by Track")
            for cfg in all_configs:
                tn = cfg['track_name']
                ok, bids = get_all_bid_tracks(tn)
                bids = bids if ok else []
                if bids:
                    with st.expander(f"📅 {tn} ({len(bids)} bids submitted)"):
                        for b in bids:
                            role = b['metadata'].get('effective_role', '?')
                            st.markdown(f"- {b['staff_name']} (v{b['version']}, {role}, submitted {b['submission_date']})")

    # ── Tab 2: Track Configs ──
    with tab2:
        # ── Section 1: create a new bid track ──
        st.markdown("### Create New Bid Track")
        new_name = st.text_input("Track Name (e.g. FY27)", key="new_bid_name")

        st.markdown("**Bid Caps**")
        cap1, cap2 = st.columns(2)
        with cap1:
            dn = st.number_input("Max Day Nurses", 1, 50, 11, key="new_dn")
            nn = st.number_input("Max Night Nurses", 1, 50, 5, key="new_nn")
        with cap2:
            dm = st.number_input("Max Day Medics", 1, 50, 11, key="new_dm")
            nm = st.number_input("Max Night Medics", 1, 50, 5, key="new_nm")

        if st.button("Create Bid Track", key="create_bid_btn", use_container_width=True):
            if not new_name.strip():
                st.error("Please enter a track name.")
            else:
                ok, msg = create_track_config(new_name.strip(), dn, dm, nn, nm)
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

        # ── Section 2: manage existing track configs ──
        st.markdown("---")
        st.subheader("Manage Track Configs")

        if not all_configs:
            st.info("No track configs yet.")

        for cfg in all_configs:
            tn = cfg['track_name']
            just_saved = st.session_state.get(f'config_saved_{tn}', False)
            with st.expander(f"{'🟢' if cfg['is_active'] else '🔵' if cfg['is_bidding_open'] else '⚪'} {tn}", expanded=just_saved):
                status_label = 'Active' if cfg['is_active'] else ('Bidding Open' if cfg['is_bidding_open'] else 'Inactive')
                st.markdown(f"**Status:** {status_label}")

                if not cfg['is_active']:
                    new_bid_state = st.checkbox(
                        "Bidding Open", value=bool(cfg['is_bidding_open']),
                        key=f"toggle_bid_{tn}")
                    if new_bid_state != bool(cfg['is_bidding_open']):
                        toggle_bidding(tn, new_bid_state)
                        st.rerun()

                # Editable fields for ALL configs (active and non-active)
                st.markdown("**Bid Caps**")
                uc1, uc2 = st.columns(2)
                with uc1:
                    u_dn = st.number_input("Day Nurses", 1, 50, cfg['max_day_nurses'], key=f"u_dn_{tn}")
                    u_nn = st.number_input("Night Nurses", 1, 50, cfg['max_night_nurses'], key=f"u_nn_{tn}")
                with uc2:
                    u_dm = st.number_input("Day Medics", 1, 50, cfg['max_day_medics'], key=f"u_dm_{tn}")
                    u_nm = st.number_input("Night Medics", 1, 50, cfg['max_night_medics'], key=f"u_nm_{tn}")

                st.markdown("**Day-of-Week Limits** *(optional — further restricts the Bid Caps above on specific weekdays)*")
                use_wd_cap = st.checkbox(
                    "Use day-of-week specific limits", value=bool(cfg.get('use_weekday_capacity', False)),
                    key=f"use_wd_cap_{tn}")
                if use_wd_cap != bool(cfg.get('use_weekday_capacity', False)):
                    update_track_config(tn, use_weekday_capacity=1 if use_wd_cap else 0)
                    st.rerun()

                if use_wd_cap:
                    weekday_order = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
                    overrides = get_weekday_capacity_overrides(tn)
                    grid_rows = []
                    for wd in weekday_order:
                        ov = overrides.get(wd, {})
                        grid_rows.append({
                            "Day": wd,
                            "Max Day Nurses": ov.get('max_day_nurses') if ov.get('max_day_nurses') is not None else u_dn,
                            "Max Day Medics": ov.get('max_day_medics') if ov.get('max_day_medics') is not None else u_dm,
                            "Max Night Nurses": ov.get('max_night_nurses') if ov.get('max_night_nurses') is not None else u_nn,
                            "Max Night Medics": ov.get('max_night_medics') if ov.get('max_night_medics') is not None else u_nm,
                        })
                    edited_grid = st.data_editor(
                        pd.DataFrame(grid_rows),
                        hide_index=True, use_container_width=True, key=f"wd_grid_{tn}",
                        column_config={
                            "Day": st.column_config.TextColumn(disabled=True),
                            "Max Day Nurses": st.column_config.NumberColumn(min_value=0, max_value=50, step=1),
                            "Max Day Medics": st.column_config.NumberColumn(min_value=0, max_value=50, step=1),
                            "Max Night Nurses": st.column_config.NumberColumn(min_value=0, max_value=50, step=1),
                            "Max Night Medics": st.column_config.NumberColumn(min_value=0, max_value=50, step=1),
                        }
                    )
                    st.caption("Blank/unedited rows use the Bid Caps above. Saving writes all 7 days explicitly, so a later change to Bid Caps won't retroactively change days you've saved here.")
                    if st.button("Save Day-of-Week Limits", key=f"save_wd_cap_{tn}", use_container_width=True):
                        for _, row in edited_grid.iterrows():
                            set_weekday_capacity_override(
                                tn, row["Day"],
                                max_day_nurses=int(row["Max Day Nurses"]),
                                max_day_medics=int(row["Max Day Medics"]),
                                max_night_nurses=int(row["Max Night Nurses"]),
                                max_night_medics=int(row["Max Night Medics"]),
                            )
                        st.session_state[f'wd_cap_saved_{tn}'] = True
                        st.rerun()

                    if st.session_state.pop(f'wd_cap_saved_{tn}', False):
                        st.success(f"Day-of-week limits saved for {tn}")

                st.markdown("**Base Shift Counts** *(day/night shift slots per base, used for hypothetical bid assignments)*")
                bc_day, bc_night = st.columns(2)
                with bc_day:
                    st.caption("Day shifts per base")
                    u_day_kmht = st.number_input("KMHT", 0, 20, cfg.get('day_kmht', 1), key=f"u_day_kmht_{tn}")
                    u_day_klwm = st.number_input("KLWM", 0, 20, cfg.get('day_klwm', 2), key=f"u_day_klwm_{tn}")
                    u_day_kbed = st.number_input("KBED", 0, 20, cfg.get('day_kbed', 2), key=f"u_day_kbed_{tn}")
                    u_day_1b9 = st.number_input("1B9", 0, 20, cfg.get('day_1b9', 2), key=f"u_day_1b9_{tn}")
                    u_day_kpym = st.number_input("KPYM", 0, 20, cfg.get('day_kpym', 2), key=f"u_day_kpym_{tn}")
                with bc_night:
                    st.caption("Night shifts per base")
                    u_night_klwm = st.number_input("KLWM", 0, 20, cfg.get('night_klwm', 1), key=f"u_night_klwm_{tn}")
                    u_night_kbed = st.number_input("KBED", 0, 20, cfg.get('night_kbed', 2), key=f"u_night_kbed_{tn}")
                    u_night_kpym = st.number_input("KPYM", 0, 20, cfg.get('night_kpym', 2), key=f"u_night_kpym_{tn}")
                    st.caption("KMHT and 1B9 have no night shifts")

                if st.button("Save All Settings", key=f"save_cap_{tn}", use_container_width=True):
                    ok, msg = update_track_config(tn,
                                        max_day_nurses=u_dn, max_day_medics=u_dm,
                                        max_night_nurses=u_nn, max_night_medics=u_nm,
                                        day_kmht=u_day_kmht, day_klwm=u_day_klwm,
                                        day_kbed=u_day_kbed, day_1b9=u_day_1b9, day_kpym=u_day_kpym,
                                        night_klwm=u_night_klwm, night_kbed=u_night_kbed,
                                        night_kpym=u_night_kpym)
                    if ok:
                        st.session_state[f'config_saved_{tn}'] = True
                        st.rerun()
                    else:
                        st.error(f"Save failed: {msg}")

                if st.session_state.pop(f'config_saved_{tn}', False):
                    st.success(f"Settings saved for {tn}")

                # Bid count (individual bids are managed in the Add/Remove Selection tab)
                st.markdown("---")
                bid_tracks_result = get_all_bid_tracks(tn)
                bid_list = bid_tracks_result[1] if bid_tracks_result[0] else []
                bid_count = len(bid_list) if isinstance(bid_list, list) else 0
                st.markdown(f"**Bids submitted:** {bid_count}")
                st.caption("Manage individual bids in the Add/Remove Selection tab.")

                if bid_count > 0:
                    # Wipe all bids button
                    if st.button(f"Wipe All Bids for {tn}", key=f"wipe_bids_{tn}", use_container_width=True):
                        st.session_state[f'confirm_wipe_{tn}'] = True

                    if st.session_state.get(f'confirm_wipe_{tn}', False):
                        st.warning(f"This will delete **all {bid_count} bids** for {tn}. This cannot be undone.")
                        wc1, wc2 = st.columns(2)
                        with wc1:
                            if st.button("Yes, Wipe All", key=f"yes_wipe_{tn}"):
                                ok, msg = wipe_all_bids(tn)
                                st.session_state[f'confirm_wipe_{tn}'] = False
                                if ok:
                                    st.success(msg)
                                else:
                                    st.error(msg)
                                st.rerun()
                        with wc2:
                            if st.button("Cancel", key=f"no_wipe_{tn}"):
                                st.session_state[f'confirm_wipe_{tn}'] = False
                                st.rerun()

                if not cfg['is_active']:
                    # Promote to active
                    st.markdown("---")
                    if st.button(f"Promote {tn} to Active", key=f"promote_{tn}",
                                 type="primary", use_container_width=True):
                        st.session_state[f'confirm_promote_{tn}'] = True

                    if st.session_state.get(f'confirm_promote_{tn}', False):
                        st.warning(f"This will deactivate the current active track and make **{tn}** active. Are you sure?")
                        c1, c2 = st.columns(2)
                        with c1:
                            if st.button("Yes, Promote", key=f"confirm_yes_{tn}"):
                                ok, msg = promote_bid_to_active(tn)
                                if ok:
                                    st.success(msg)
                                    st.session_state[f'confirm_promote_{tn}'] = False
                                    st.rerun()
                                else:
                                    st.error(msg)
                        with c2:
                            if st.button("Cancel", key=f"confirm_no_{tn}"):
                                st.session_state[f'confirm_promote_{tn}'] = False
                                st.rerun()

                    # Delete track config
                    st.markdown("---")
                    if st.button(f"Delete {tn}", key=f"delete_cfg_{tn}", use_container_width=True):
                        st.session_state[f'confirm_delete_cfg_{tn}'] = True

                    if st.session_state.get(f'confirm_delete_cfg_{tn}', False):
                        st.error(f"Delete track config **{tn}** and all its bids? This cannot be undone.")
                        d1, d2 = st.columns(2)
                        with d1:
                            if st.button("Yes, Delete", key=f"yes_del_cfg_{tn}"):
                                ok, msg = delete_track_config(tn)
                                st.session_state[f'confirm_delete_cfg_{tn}'] = False
                                if ok:
                                    st.success(msg)
                                else:
                                    st.error(msg)
                                st.rerun()
                        with d2:
                            if st.button("Cancel", key=f"no_del_cfg_{tn}"):
                                st.session_state[f'confirm_delete_cfg_{tn}'] = False
                                st.rerun()

    # ── Tab 3: Manage Bid Access ──
    with tab3:
        st.markdown("### Manage Bid Access")

        if not config_names:
            st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        else:
            access_track = st.selectbox(
                "Track Cycle:", config_names, index=default_track_index, key="access_track_select")

            ctx, roster_error = _load_bidding_data_files()
            if ctx is None:
                st.error(roster_error)
            else:
                staff_names = ctx['staff_names']
                role_mapping = ctx['role_mapping']
                seniority_mapping = ctx['seniority_mapping']
                access_details = get_all_bid_access_details(access_track)
                ok, bids = get_all_bid_tracks(access_track)
                bids_lookup = {b['staff_name']: b for b in bids} if ok else {}

                staff_data = []
                for staff_name in staff_names:
                    role = role_mapping.get(staff_name, 'Unknown')
                    seniority = seniority_mapping.get(staff_name, '')
                    detail = access_details.get(staff_name, {})
                    access = detail.get('access', False)
                    bid = bids_lookup.get(staff_name)
                    staff_data.append({
                        'Staff Name': staff_name,
                        'Role': role,
                        'Seniority': seniority,
                        'Bid Access': '✅' if access else '❌',
                        'Notification Sent': detail.get('access_opened_date') or '',
                        'Has Bid': '✅' if bid else '❌',
                        'Version': bid['version'] if bid else '',
                        'Submitted': bid['submission_date'] if bid else '',
                    })

                def _seniority_key(row):
                    # Most-senior-first; blank/non-numeric seniority sorts last.
                    try:
                        return (0, float(row['Seniority']))
                    except (TypeError, ValueError):
                        return (1, 0)

                display_cols = ['Staff Name', 'Seniority', 'Bid Access', 'Notification Sent', 'Has Bid', 'Version', 'Submitted']
                nurse_data = sorted(
                    (d for d in staff_data if str(d['Role']).strip().lower() != 'medic'),
                    key=_seniority_key)
                medic_data = sorted(
                    (d for d in staff_data if str(d['Role']).strip().lower() == 'medic'),
                    key=_seniority_key)

                # Eligible-to-bid rosters (management staff — blank SHIFTS PER PAY PERIOD
                # in Requirements.xlsx — and _BID_INELIGIBLE_STAFF excluded), for the
                # submitted/remaining counters below. Same helpers the auto bid progression
                # cascade uses, so the counts
                # stay consistent with who actually gets skipped there.
                requirements_map = _load_requirements_map(ctx.get('requirements_df'))
                nurse_roster = _ordered_bidding_roster(
                    staff_names, role_mapping, seniority_mapping, requirements_map, 'nurse')
                medic_roster = _ordered_bidding_roster(
                    staff_names, role_mapping, seniority_mapping, requirements_map, 'medic')
                nurse_submitted_count = sum(1 for name in nurse_roster if name in bids_lookup)
                medic_submitted_count = sum(1 for name in medic_roster if name in bids_lookup)

                col_nurse, col_medic = st.columns(2)
                with col_nurse:
                    st.markdown(f"##### Nurses ({len(nurse_data)})")
                    st.dataframe(
                        pd.DataFrame(nurse_data, columns=display_cols),
                        use_container_width=True, hide_index=True)
                    st.caption(
                        f"**{nurse_submitted_count} submitted** · "
                        f"**{len(nurse_roster) - nurse_submitted_count} remaining** "
                        f"({len(nurse_roster)} eligible bidders; management/non-bidding staff excluded)")
                with col_medic:
                    st.markdown(f"##### Medics ({len(medic_data)})")
                    st.dataframe(
                        pd.DataFrame(medic_data, columns=display_cols),
                        use_container_width=True, hide_index=True)
                    st.caption(
                        f"**{medic_submitted_count} submitted** · "
                        f"**{len(medic_roster) - medic_submitted_count} remaining** "
                        f"({len(medic_roster)} eligible bidders; management/non-bidding staff excluded)")

                st.markdown("---")
                st.markdown("### Automatic Bid Access & Notification")
                st.caption(
                    "When enabled, submitting a bid automatically grants bid access to the next "
                    "staff member in seniority rank order (same role — Nurse/Dual or Medic) and "
                    "emails them that their bid is now open, along with the admins. Staff with a "
                    "blank **SHIFTS PER PAY PERIOD** in Requirements.xlsx are management, not "
                    "bidding on tracks, and are skipped. If that next staff member already has bid "
                    "access enabled (e.g. a revision to an earlier/more senior staff member's bid "
                    "after the cascade already passed them), they're skipped — no duplicate "
                    "notification is sent. The admin notice for the bid submission/revision itself "
                    "is always sent regardless."
                )
                track_cfg = get_track_config_by_name(access_track)
                auto_progression_on = bool(track_cfg.get('auto_bid_progression')) if track_cfg else False
                new_auto_progression_on = st.checkbox(
                    f"Enable automatic bid access & notification for {access_track}",
                    value=auto_progression_on, key=f"auto_progression_{access_track}")
                if new_auto_progression_on != auto_progression_on:
                    update_track_config(access_track, auto_bid_progression=1 if new_auto_progression_on else 0)
                    st.rerun()

                st.markdown("##### Manually Send Bid Notification")
                st.caption(
                    "Send the \"your bid is open\" notification to a specific staff member right "
                    "now, using an email address you enter below (not looked up from "
                    "Requirements.xlsx). This does not change their bid access — use "
                    "**Toggle Access** below for that."
                )
                manual_col1, manual_col2 = st.columns(2)
                with manual_col1:
                    manual_notify_staff = st.selectbox(
                        "Staff Member:", staff_names, key=f"manual_notify_staff_{access_track}")
                with manual_col2:
                    manual_notify_email = st.text_input(
                        "Send to email:", key=f"manual_notify_email_{access_track}",
                        placeholder="name@example.com")
                if st.button("Send Notification", key=f"manual_notify_send_{access_track}"):
                    level, message = _send_manual_bid_notification(
                        manual_notify_staff, manual_notify_email, access_track)
                    (st.success if level == "success" else st.warning)(message)

                st.markdown("##### Notification Log")
                progression_log = get_bid_progression_log(access_track, limit=100)
                if not progression_log:
                    st.caption(f"No bid-progression events yet for {access_track}.")
                else:
                    level_icon = {'success': '✅', 'warning': '⚠️', 'info': 'ℹ️'}
                    trigger_label = {'auto': 'Auto', 'manual': 'Manual'}
                    log_rows = [{
                        'Date/Time': entry['event_date'],
                        'Trigger': trigger_label.get(entry.get('trigger_type'), 'Auto'),
                        'Status': f"{level_icon.get(entry['level'], '')} {entry['level'].title()}".strip(),
                        'Submitted By': entry['submitted_by'],
                        'Next Staff': entry['next_staff'] or '—',
                        'Notified Email': entry['notified_email'] or '—',
                        'Details': entry['message'],
                    } for entry in progression_log]
                    st.dataframe(pd.DataFrame(log_rows), use_container_width=True, hide_index=True)
                    st.caption(f"Showing the {len(progression_log)} most recent event(s) for {access_track}.")

                st.markdown("---")
                st.markdown("### Toggle Access")

                selected_staff_access = st.selectbox(
                    "Select Staff Member:", staff_names, key="toggle_access_staff")

                if selected_staff_access:
                    current_status = access_details.get(selected_staff_access, {}).get('access', False)
                    staff_role = role_mapping.get(selected_staff_access, 'Unknown')

                    st.info(f"**{selected_staff_access}** ({staff_role}) - Bid Access: "
                            f"**{'Yes' if current_status else 'No'}**")

                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✅ Enable Bid Access", key="enable_access_btn", use_container_width=True):
                            ok, msg = set_bid_access(selected_staff_access, access_track, True)
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.error(msg)
                    with col2:
                        if st.button("❌ Disable Bid Access", key="disable_access_btn", use_container_width=True):
                            ok, msg = set_bid_access(selected_staff_access, access_track, False)
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.error(msg)

                # Bulk enable/disable
                st.markdown("---")
                st.markdown("### Bulk Operations")

                col3, col4 = st.columns(2)
                with col3:
                    if st.button("✅ Enable All Staff", key="enable_all_access_btn", use_container_width=True):
                        count = 0
                        for staff in staff_names:
                            ok, _ = set_bid_access(staff, access_track, True)
                            if ok:
                                count += 1
                        st.success(f"Enabled bid access for {count} staff members")
                        st.rerun()

                with col4:
                    if st.button("❌ Disable All Staff", key="disable_all_access_btn", use_container_width=True):
                        count = 0
                        for staff in staff_names:
                            ok, _ = set_bid_access(staff, access_track, False)
                            if ok:
                                count += 1
                        st.success(f"Disabled bid access for {count} staff members")
                        st.rerun()

    # ── Tab 4: Add/Remove Selection ──
    with tab4:
        st.markdown("### Add or Remove Selection")

        if not config_names:
            st.info("No track cycles exist yet. Create one in the Track Configs tab.")
        else:
            sel_track = st.selectbox(
                "Track Cycle:", config_names, index=default_track_index, key="admin_sel_track")

            ctx, roster_error = _load_bidding_data_files()
            if ctx is None:
                st.error(roster_error)
            else:
                staff_names = ctx['staff_names']
                role_mapping = ctx['role_mapping']

                admin_selected_staff = st.selectbox(
                    "Select Staff Member:", staff_names, key="admin_staff_select_bid")

                if admin_selected_staff:
                    staff_role = role_mapping.get(admin_selected_staff, 'Unknown')
                    st.info(f"**{admin_selected_staff}** ({staff_role})")

                    bid_result = get_bid_track_from_db(admin_selected_staff, sel_track)
                    has_bid = bid_result[0]

                    if has_bid:
                        b = bid_result[1]
                        st.success(f"Current bid: version {b['version']}, submitted {b['submission_date']}")

                        from modules.track_management.display import display_schedule_by_blocks
                        with st.expander("View submitted schedule"):
                            display_schedule_by_blocks(b['track_data'], ctx['days'], {})

                        if st.button("❌ Remove This Selection", key="remove_bid_btn"):
                            st.session_state['confirm_remove_bid'] = True

                        if st.session_state.get('confirm_remove_bid', False):
                            st.warning(f"Delete bid for **{admin_selected_staff}**? This cannot be undone.")
                            rc1, rc2 = st.columns(2)
                            with rc1:
                                if st.button("Yes, Delete", key="confirm_remove_bid_yes"):
                                    ok, msg = delete_bid(admin_selected_staff, sel_track)
                                    st.session_state['confirm_remove_bid'] = False
                                    if ok:
                                        st.success(msg)
                                    else:
                                        st.error(msg)
                                    st.rerun()
                            with rc2:
                                if st.button("Cancel", key="confirm_remove_bid_no"):
                                    st.session_state['confirm_remove_bid'] = False
                                    st.rerun()
                    else:
                        st.info(f"{admin_selected_staff} has not submitted a bid for {sel_track} yet.")

                    st.markdown("---")
                    st.markdown("### Add/Update Selection")
                    st.caption(
                        "Builds or edits a bid on this staff member's behalf using the same editor "
                        "staff use — Track Selection, Validation, and Submission included. Unlike the "
                        "staff-facing view, this never locks: saving here always creates a new version."
                    )

                    # The full editor re-runs schedule-competition calculations, which are expensive.
                    # Streamlit re-executes every tab's code on any admin-dashboard interaction, so the
                    # editor is gated behind an explicit open/close action — otherwise every click on
                    # Overview/Track Configs/Manage Bid Access would silently pay for recomputing it too.
                    editor_key = f'show_admin_bid_editor_{sel_track}_{admin_selected_staff}'
                    if not st.session_state.get(editor_key, False):
                        if st.button(f"📝 Open Bid Editor for {admin_selected_staff}",
                                     key="open_admin_editor_btn", type="primary", use_container_width=True):
                            st.session_state[editor_key] = True
                            st.rerun()
                    else:
                        if st.button("Close Editor", key="close_admin_editor_btn"):
                            st.session_state[editor_key] = False
                            st.rerun()

                        cap_for_track = get_track_capacity(sel_track)
                        st.session_state['preferences_df'] = ctx['preferences_df']
                        st.session_state['staff_col_prefs'] = ctx['staff_col_prefs']
                        st.session_state['role_col'] = ctx['role_col']

                        _display_bidding_staff_interface(
                            admin_selected_staff, ctx['preferences_df'], ctx['current_tracks_df'], ctx['requirements_df'],
                            ctx['days'], ctx['staff_col_prefs'], ctx['staff_col_tracks'], ctx['role_col'],
                            ctx['no_matrix_col'], ctx['reduced_rest_col'], ctx['seniority_col'],
                            ctx['preassignment_df'], sel_track, cap_for_track, is_admin=True
                        )

    # ── Tab 5: Bid Analysis ──
    with tab5:
        _render_bid_analysis_tab(config_names, default_track_index)

    # ── Tab 6: Bid Roster ──
    with tab6:
        _render_bid_roster_tab(config_names, default_track_index)

    # ── Tab 7: Base Analysis ──
    with tab7:
        _render_base_analysis_tab(config_names, default_track_index)

    # ── Tab 8: Staffing Rebalance ──
    with tab8:
        from modules.staffing_rebalance import _render_staffing_rebalance_tab
        _render_staffing_rebalance_tab(config_names, default_track_index)

    # ── Tab 9: Needs Swap Requests ──
    with tab9:
        from modules.track_needs_swap import _render_needs_swap_admin_tab
        _render_needs_swap_admin_tab(config_names, default_track_index)


# ──────────────────────────────────────────────
# Main bidding page (staff-facing)
# ──────────────────────────────────────────────

def _render_bidding_instructions():
    """
    Full step-by-step bidding walkthrough, shown at the top of the staff-facing
    Track Bidding page. This is the page the "bid access opened" notification
    email points to ("Detailed instructions are found linked at the top in the
    Track Bidding module.") — keep the tab names/emoji below in sync with the
    tab_labels in _display_bidding_staff_interface if those ever change.
    """
    with st.expander("📖 **Bidding Instructions — click to view the full step-by-step guide**"):
        st.markdown("""
Fiscal Year Info - First Sunday of Block A = Sept 27, 2026

1. **Select your name** from the dropdown to begin. The page will show the maximum staffing capacities (per day/night), your individual shift requirements based on years of service per the CBA, and a snapshot of your current active track for reference.

2. **Make sure your base preferences (rankings) are up to date before bidding.** No preferences entered means the system doesn't know where you want to work.
   - Check your **⚙️ Preferences** tab and, if needed, update them in **🛠️ Edit Preferences**.
   - Each base gets a 1-5 ranking for DAYS and a 1-3 ranking for NIGHTS (1 = first choice, most desired).
   - Confirm your current Zip Code (where you commute from), and choose whether to enroll in:
     - **Reduced Rest OK** — pre-approve getting scheduled for 10 hours between shifts when needed, to increase the likelihood of a shift at your preferred base.
     - **N to D Flex** — when drafting a schedule, do you want to be flexed to a DAY shift on the same date instead of a track NIGHT shift when staffing needs allow it? Includes a "Maybe" option where schedulers will ask first, but this might limit availability.

3. **Go to 🔄 Track Selection to begin bidding.**
   - The schedule tracks follow a 6-week rotation, with 3 pay periods (Blocks A, B, C) and corresponding week numbers (e.g. "Wed B4" = Block B, the second Wednesday of that pay period, which is the 4th week of the 6-week track).
   - If your role gets AT days, those are already pre-assigned to align with your position's scheduling needs and are built into your track as day shifts. Unlike a Night shift, which needs 2 days off before another day shift, only 1 day off is required after a Night → AT.

4. **Fill in Track Selection, one block/week at a time.**
   - Days highlighted in green are where your role is currently needed; a "Day Need" or "Night Need" count and a hypothetical base assignment appear under each day as a preview.
   - Select **"🔍 Validate and Save Block"** before moving on to the next block — this is important. This button sits between the block tabs and the weeks and stores your selections.
   - Not going to finish in one sitting? Click **"💾 Save Progress"** — it sits just above the block tabs — at any point. It saves everything you've entered so far across all three blocks and brings it back automatically the next time you sign in, even if you haven't submitted yet.

5. **Scheduler logic.** Under every day in Track Selection, and in full on the **🔮 Hypothetical Schedule** tab, CrewOps360 runs a live simulation of the upcoming competition: it sorts everyone bidding a given day/shift by seniority (nurses and medics separately — dual-trained staff compete as nurses), then hands out base seats one person at a time, most senior first, with each person getting their highest-ranked base that still has room.
   - **Choices disappear once a shift is full.** Each day/shift combination has a cap, shown in the header above. Once enough people have bid that shift to hit the cap, the D or N button for that day is dropped as an option.
   - **"Need exists but all named shifts are filled"** means all the named shifts are spoken for (ex. there are only 4 night shifts but the bid cap = 6). You can still take that D or N, but your location can't be accurately calculated — the beauty of the bid: do you value LOCATION or SCHEDULE?

6. **🔍 Validation tab.** Once all three blocks are filled, open Validation for a full pass against every rule, all in one place. Anything unmet is called out individually so you know exactly what to fix and where.

7. **📤 Submission.** Review the full six-week preview, then select **Submit Bid**. You'll have the option to download and/or email a PDF of your Bid Summary for your records — you can always come back and do this later.

8. **Communications.** We'll send a text message to notify you when it's your turn to bid, but we kindly ask that all schedule-related questions be sent via email to Matt, Aaron, and Jen. This helps consolidate responses, provide consistent information, and better manage communications when we're away from work.

Happy bidding!

~Charlie, Jen, Matt & Aaron
""")


def display_track_bidding():
    """Main entry point for the Track Bidding section."""
    st.markdown("")
    st.markdown("")

    if st.button("← Back to CrewOps360", key="back_from_bidding"):
        st.session_state.selected_module = None
        st.rerun()

    st.markdown("# Track Bidding")

    # Render admin sidebar (password gate -> full-page admin mode)
    _render_admin_mode_toggle()

    if st.session_state.get('track_bidding_admin_mode'):
        display_bidding_admin_interface()
        return

    _render_bidding_instructions()

    # Check if there is an open bidding track
    bid_cfg = get_bidding_track_config()
    active_cfg = get_active_track_config()

    # Track Needs Swap — opens after bidding closes and the shortfalls are known, so
    # it renders on its own regardless of whether a bidding cycle is currently open.
    from modules.track_needs_swap import display_staff_needs_swap
    swap_shown = display_staff_needs_swap()
    if swap_shown:
        st.markdown("---")

    if not bid_cfg:
        if not swap_shown:
            st.info("Bidding is currently closed. Check back later for the next bidding cycle.")
            if active_cfg:
                st.markdown(f"**Current active track:** {active_cfg['track_name']}")
        return

    bid_track_name = bid_cfg['track_name']
    st.markdown(f"### Bidding open for: **{bid_track_name}**")

    # Show capacity info
    cap = get_track_capacity(bid_track_name)
    st.markdown("**Staffing Capacity**")
    cap_cols = st.columns(4)
    cap_cols[0].metric("Max Day Nurses", cap['max_day_nurses'])
    cap_cols[1].metric("Max Day Medics", cap['max_day_medics'])
    cap_cols[2].metric("Max Night Nurses", cap['max_night_nurses'])
    cap_cols[3].metric("Max Night Medics", cap['max_night_medics'])

    if cap.get('use_weekday_capacity'):
        with st.expander("Day-of-week limits (in addition to the caps above)"):
            weekday_caps = get_track_capacity_by_weekday(bid_track_name)
            weekday_order = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
            table = [
                {
                    "Day": wd,
                    "Max Day Nurses": weekday_caps[wd]['max_day_nurses'],
                    "Max Day Medics": weekday_caps[wd]['max_day_medics'],
                    "Max Night Nurses": weekday_caps[wd]['max_night_nurses'],
                    "Max Night Medics": weekday_caps[wd]['max_night_medics'],
                }
                for wd in weekday_order
            ]
            st.dataframe(pd.DataFrame(table), use_container_width=True, hide_index=True)

    if active_cfg:
        st.markdown(f"*Prior active track: {active_cfg['track_name']}*")

    # ── Load data files (same as clinical track hub) ──
    _run_bidding_interface(bid_track_name, cap)


def _run_bidding_interface(bid_track_name, capacity):
    """Run the bidding interface — mirrors run_clinical_track_hub but for bids."""
    ctx, error = _load_bidding_data_files()
    if ctx is None:
        st.error(error)
        return

    st.markdown("---")
    selected_staff = st.selectbox("Select Your Name", [""] + ctx['staff_names'], key="bid_staff_select")

    if not selected_staff:
        st.info("Please select your name to begin.")
        return

    # Per-staff bidding access gate — mirrors Summer Leave's "LT Access" toggle.
    # Once an admin has locked this down, staff can't self-serve without being enabled.
    if not get_bid_access(selected_staff, bid_track_name):
        st.warning("⚠️ Bidding access is not available for you at this time.")
        st.info("Please contact your supervisor if you believe this is an error.")
        return

    # Store for submission access
    st.session_state['preferences_df'] = ctx['preferences_df']
    st.session_state['staff_col_prefs'] = ctx['staff_col_prefs']
    st.session_state['role_col'] = ctx['role_col']

    # Run the bidding staff interface
    _display_bidding_staff_interface(
        selected_staff, ctx['preferences_df'], ctx['current_tracks_df'], ctx['requirements_df'],
        ctx['days'], ctx['staff_col_prefs'], ctx['staff_col_tracks'], ctx['role_col'],
        ctx['no_matrix_col'], ctx['reduced_rest_col'], ctx['seniority_col'],
        ctx['preassignment_df'], bid_track_name, capacity, is_admin=False
    )


def _display_bidding_staff_interface(
    selected_staff, preferences_df, current_tracks_df, requirements_df,
    days, staff_col_prefs, staff_col_tracks, role_col,
    no_matrix_col, reduced_rest_col, seniority_col,
    preassignment_df, bid_track_name, capacity, is_admin=False
):
    """Render the tabbed bidding interface for a single staff member.

    When is_admin is True (called from the admin "Add/Update Selection" tab),
    the Submission tab never locks — an admin can always build or revise a bid
    on the selected staff member's behalf, regardless of what's already on file.
    """
    from modules.track_management.display import display_schedule_by_blocks
    from modules.track_management.preference_display import display_preferences
    from modules.preference_editor import display_location_preference_editor
    from modules.track_management.preassignment import get_staff_preassignments
    from modules.track_management.utils import reset_track_session_state
    from modules.enhanced_track_validator import validate_track_comprehensive
    from modules.enhanced_validation_display import display_comprehensive_validation
    from modules.track_modification_core import calculate_all_modification_options
    from modules.db_utils import get_location_preferences_from_db

    st.header(f"Track Bidding — {bid_track_name}")
    if is_admin:
        st.caption(f"🔧 Admin Mode — building/editing this bid on behalf of **{selected_staff}**.")

    # ── Extract requirements ──
    shifts_per_pay_period = 0
    night_minimum = 0
    weekend_minimum = 0
    weekend_group = None

    if requirements_df is not None and not requirements_df.empty:
        try:
            staff_req = None
            possible_cols = [requirements_df.columns[0], 'STAFF NAME', 'Staff Name', 'staff name', 'Name', 'NAME']
            for col_name in possible_cols:
                if col_name in requirements_df.columns:
                    staff_req = requirements_df[requirements_df[col_name] == selected_staff]
                    if staff_req.empty:
                        staff_req = requirements_df[requirements_df[col_name].str.lower() == selected_staff.lower()]
                    if not staff_req.empty:
                        break
            if staff_req is not None and not staff_req.empty:
                row = staff_req.iloc[0]
                if len(requirements_df.columns) >= 4:
                    if pd.notna(row.iloc[1]):
                        shifts_per_pay_period = int(float(row.iloc[1]))
                    if pd.notna(row.iloc[2]):
                        night_minimum = int(float(row.iloc[2]))
                    if pd.notna(row.iloc[3]):
                        weekend_minimum = int(float(row.iloc[3]))
                if len(requirements_df.columns) >= 5 and pd.notna(row.iloc[4]):
                    wg = str(row.iloc[4]).strip().upper()
                    if wg in ['A', 'B', 'C', 'D', 'E']:
                        weekend_group = wg
        except Exception as e:
            st.warning(f"Error loading requirements: {e}")

    # Requirements display
    st.markdown("### Staff Requirements")
    rc = st.columns(4)
    rc[0].metric("Shifts per Pay Period", shifts_per_pay_period)
    rc[1].metric("Night Minimum", night_minimum)
    rc[2].metric("Weekend Minimum", weekend_minimum)
    rc[3].metric("Weekend Group", weekend_group or "None")

    # Staff info
    staff_info = preferences_df[preferences_df[staff_col_prefs] == selected_staff].iloc[0]

    # Get preassignments
    staff_preassignments = {}
    if preassignment_df is not None:
        staff_preassignments = get_staff_preassignments(selected_staff, preassignment_df, days)

    # Check for existing bid
    bid_result = get_bid_track_from_db(selected_staff, bid_track_name)
    has_bid = bid_result[0]

    # Check for active track (reference)
    active_cfg = get_active_track_config()
    active_track_name = active_cfg['track_name'] if active_cfg else 'FY26'
    active_result = get_track_from_db(selected_staff, active_track_name)
    has_active = active_result[0]

    # Determine starting point for the bid editor: existing bid > saved draft > blank
    # (never the active track — a fresh bid should start empty, not a copy of last cycle's track)
    if has_bid:
        current_track_data = bid_result[1]['track_data']
    else:
        draft_result = get_bid_draft(selected_staff, bid_track_name)
        if draft_result[0]:
            current_track_data = draft_result[1]['track_data']
        else:
            current_track_data = {day: "" for day in days}

    # Apply preassignments
    if staff_preassignments:
        for day, pa in staff_preassignments.items():
            if day not in current_track_data or not current_track_data[day]:
                current_track_data[day] = pa

    # Store requirements
    st.session_state.shifts_per_pay_period = shifts_per_pay_period
    st.session_state.night_minimum = night_minimum
    st.session_state.weekend_minimum = weekend_minimum
    st.session_state.weekend_group = weekend_group

    # Session state keys for bidding (namespaced to avoid collision with clinical hub)
    bid_changes_key = f'bid_track_changes_{bid_track_name}'
    bid_modified_key = f'bid_modified_track_{bid_track_name}'

    # Clear button
    if st.button("Clear All Shifts", key=f"bid_clear_{selected_staff}_{bid_track_name}", use_container_width=True):
        blank = {day: "" for day in days}
        if staff_preassignments:
            for day, pa in staff_preassignments.items():
                blank[day] = pa
        st.session_state[bid_changes_key] = {selected_staff: blank}
        st.session_state[bid_modified_key] = {
            'staff': selected_staff, 'track': blank.copy(), 'valid': False, 'is_new': True
        }
        st.success("Cleared your in-progress selections below. If you already submitted a bid, it is unchanged until you submit again.")
        st.rerun()
    st.caption("Clears your working selections below — does not delete a bid you've already submitted.")

    # Initialize session state for bidding track changes
    if bid_changes_key not in st.session_state:
        st.session_state[bid_changes_key] = {}
    if selected_staff not in st.session_state[bid_changes_key]:
        st.session_state[bid_changes_key][selected_staff] = current_track_data.copy()
    if bid_modified_key not in st.session_state or st.session_state[bid_modified_key].get('staff') != selected_staff:
        st.session_state[bid_modified_key] = {
            'staff': selected_staff,
            'track': st.session_state[bid_changes_key][selected_staff].copy(),
            'valid': False,
            'is_new': not has_bid
        }

    # Alias into the main session keys so existing editor/validator modules work
    st.session_state.track_changes = st.session_state[bid_changes_key]
    st.session_state.modified_track = st.session_state[bid_modified_key]

    # ── Tabs ──
    tab_labels = [
        "📍 Current Track", "⚙️ Preferences", "🛠️ Edit Preferences",
        "🔄 Track Selection", "🔍 Validation", "📤 Submission",
        "🔮 Hypothetical Schedule"
    ]
    tabs = st.tabs(tab_labels)

    # Helper to build current track for validation
    def _build_track():
        vt = {day: "" for day in days}
        if selected_staff in st.session_state[bid_changes_key]:
            vt.update(st.session_state[bid_changes_key][selected_staff])
        if staff_preassignments:
            for day, pa in staff_preassignments.items():
                if pa == "AT":
                    vt[day] = "AT"
                elif pa in ["D", "N"]:
                    vt[day] = pa
                else:
                    vt[day] = "D"
        return vt

    # ── Tab 0: Current Track ──
    with tabs[0]:
        st.subheader("Current Track")
        if has_active:
            st.info(f"📊 Your active track: **{active_track_name}**")
            display_schedule_by_blocks(active_result[1]['track_data'], days, staff_preassignments)
        else:
            st.info("You do not have an active track on file yet.")

        st.markdown("---")

        st.subheader("Current Track Bid")
        if has_bid:
            st.info(f"📊 Your submitted bid for **{bid_track_name}** (version {bid_result[1]['version']}, submitted {bid_result[1]['submission_date']}).")
            display_schedule_by_blocks(bid_result[1]['track_data'], days, staff_preassignments)
        else:
            st.info(f"You have not submitted a bid for **{bid_track_name}** yet. Use the **Track Selection** tab to build your bid, then submit it from the **Submission** tab.")

    # ── Tab 1: Preferences ──
    with tabs[1]:
        display_preferences(selected_staff, staff_info, preferences_df)

    # ── Tab 2: Edit Preferences ──
    with tabs[2]:
        display_location_preference_editor(selected_staff)

    # ── Tab 3: Track Selection (the bidding editor) ──
    with tabs[3]:
        _display_track_selection_tab(
            selected_staff, preferences_df, current_tracks_df, days,
            staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
            reduced_rest_col, seniority_col,
            shifts_per_pay_period, night_minimum, weekend_minimum,
            staff_preassignments, weekend_group, requirements_df,
            capacity, bid_track_name, bid_changes_key, bid_modified_key
        )

    # ── Tab 4: Validation ──
    with tabs[4]:
        st.subheader(f"Bid Validation for {selected_staff}")
        current_track = _build_track()
        is_valid = display_comprehensive_validation(
            current_track, days, shifts_per_pay_period, night_minimum,
            weekend_minimum, staff_preassignments, weekend_group,
            requirements_df, selected_staff
        )
        st.session_state[bid_modified_key]['valid'] = is_valid
        st.session_state.modified_track = st.session_state[bid_modified_key]
        if is_valid:
            st.success("Your bid passes all validation requirements! Proceed to Submission.")
        else:
            st.warning("Your bid has validation issues. Review above and adjust in Track Selection.")

    # ── Tab 5: Submission ──
    with tabs[5]:
        _display_bid_submission(
            selected_staff, days, shifts_per_pay_period, night_minimum, weekend_minimum,
            staff_preassignments, bid_track_name, bid_changes_key, bid_modified_key,
            preferences_df, staff_col_prefs, role_col, is_admin=is_admin
        )

    # ── Tab 6: Hypothetical Schedule ──
    with tabs[6]:
        _display_bid_hypothetical(
            selected_staff, preferences_df, current_tracks_df, days,
            staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
            reduced_rest_col, seniority_col, capacity, bid_track_name,
            has_bid, bid_result, staff_preassignments, _build_track()
        )

    # Write back bidding-specific state
    st.session_state[bid_changes_key] = st.session_state.track_changes
    st.session_state[bid_modified_key] = st.session_state.modified_track


# ──────────────────────────────────────────────
# Tab implementations
# ──────────────────────────────────────────────

def _display_track_selection_tab(
    selected_staff, preferences_df, current_tracks_df, days,
    staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
    reduced_rest_col, seniority_col,
    shifts_per_pay_period, night_minimum, weekend_minimum,
    preassignments, weekend_group, requirements_df,
    capacity, bid_track_name, bid_changes_key, bid_modified_key
):
    """Track Selection tab — same as Track Modification but for bidding."""
    from modules.track_management.editor import display_track_modification_interface_enhanced
    from modules.track_modification_core import calculate_all_modification_options

    st.subheader(f"Track Selection for {selected_staff}")

    # Requirements
    st.markdown("### Requirements")
    rc = st.columns(4)
    rc[0].metric("Shifts/Pay Period", shifts_per_pay_period)
    rc[1].metric("Night Min", night_minimum)
    rc[2].metric("Weekend Min", weekend_minimum)
    rc[3].metric("Weekend Group", weekend_group or "None")

    st.info(f"Selecting shifts for **{bid_track_name}** bidding cycle.")

    staff_info = preferences_df[preferences_df[staff_col_prefs] == selected_staff].iloc[0]
    staff_role = staff_info[role_col]

    max_day_nurses = capacity['max_day_nurses']
    max_day_medics = capacity['max_day_medics']
    max_night_nurses = capacity['max_night_nurses']
    max_night_medics = capacity['max_night_medics']

    with st.spinner("Analyzing schedule needs and preferences..."):
        modification_results = calculate_all_modification_options(
            selected_staff, preferences_df, current_tracks_df, days,
            staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
            reduced_rest_col, seniority_col,
            max_day_nurses=max_day_nurses, max_day_medics=max_day_medics,
            max_night_nurses=max_night_nurses, max_night_medics=max_night_medics,
            bid_track_name=bid_track_name
        )
        options_by_day = modification_results["options_by_day"]
        day_assignments = modification_results["day_assignments"]
        night_assignments = modification_results["night_assignments"]
        assignment_details = modification_results["assignment_details"]

    # Reference track: always the active track — shown for comparison only, never edited
    bid_result = get_bid_track_from_db(selected_staff, bid_track_name)
    has_bid = bid_result[0]
    draft_result = get_bid_draft(selected_staff, bid_track_name)
    has_draft = draft_result[0]
    active_cfg = get_active_track_config()
    active_name = active_cfg['track_name'] if active_cfg else 'FY26'
    active_result = get_track_from_db(selected_staff, active_name)
    has_active = active_result[0]

    if has_active:
        reference_track = active_result[1]['track_data'].copy()
    elif current_tracks_df is not None and staff_col_tracks:
        st_df = current_tracks_df[current_tracks_df[staff_col_tracks] == selected_staff]
        if not st_df.empty:
            reference_track = {day: st_df.iloc[0][day] for day in days}
        else:
            reference_track = {day: "" for day in days}
    else:
        reference_track = {day: "" for day in days}

    # Initialize track changes: existing bid > saved draft > blank — never a copy of the reference track
    if selected_staff not in st.session_state.track_changes:
        if has_bid:
            track_data = bid_result[1]['track_data'].copy()
        elif has_draft:
            track_data = draft_result[1]['track_data'].copy()
        else:
            track_data = {day: "" for day in days}
        if preassignments:
            for day, pa in preassignments.items():
                if pa == "AT":
                    track_data[day] = "AT"
                else:
                    track_data[day] = "D"
        st.session_state.track_changes[selected_staff] = track_data

    if st.session_state.modified_track.get('staff') != selected_staff:
        st.session_state.modified_track = {
            'staff': selected_staff,
            'track': st.session_state.track_changes[selected_staff].copy(),
            'valid': False,
            'is_new': not has_bid
        }

    st.markdown("""
    ### How to Select Your Track

    1. Select days/nights where you want to work by selecting D, N, or Off to remove the selection
    2. Use **"Validate Block"** buttons to save individual 2-week blocks
    3. Pre-assignments (AT, if any) are shown as selected and locked
    4. Days where your role is needed are highlighted in green — darker green means it's the highest ranked hypothetical shift based on your preferences
    5. Not ready to finish in one sitting? Click **Save Progress** below at any time — your selections are stored under your name and reappear automatically the next time you come back, no submission required
    6. Go to the **Validation tab** to check your complete bid, then proceed to Submission
    7. **Note:** Hypothetical shifts are not guaranteed base assignments — your submitted track only designates a "D" or "N" for each day.
    """)

    if preassignments:
        from modules.track_management.preassignment import display_preassignments
        display_preassignments(selected_staff, preassignments)

    # Save progress: persists the whole in-progress bid (all three blocks) as a draft,
    # kept separate from the final submitted bid so staff can leave and resume freely
    # before Submission without triggering the "bid already submitted" lock.
    save_col1, save_col2, save_col3 = st.columns([1, 2, 1])
    with save_col2:
        if st.button("💾 Save Progress", key=f"save_bid_progress_{bid_track_name}_{selected_staff}",
                     use_container_width=True):
            ok, result = save_bid_draft(selected_staff, bid_track_name, st.session_state.track_changes[selected_staff])
            if ok:
                st.success(f"✅ Progress saved at {result}. Come back anytime before Submission to pick up where you left off.")
            else:
                st.error(f"Error saving progress: {result}")
        elif has_draft:
            st.caption(f"💾 Progress last saved {draft_result[1]['saved_date']}")
    st.markdown("---")

    # Use database logic
    use_database_logic = True
    has_db_track = has_bid

    display_track_modification_interface_enhanced(
        selected_staff, options_by_day, reference_track, days,
        preassignments, use_database_logic, has_db_track, staff_role, weekend_group,
        day_assignments, night_assignments, assignment_details
    )

    # Quick validation
    st.markdown("### Quick Validation Status")
    st.info("For comprehensive results, go to the **Validation tab**.")
    from modules.enhanced_track_validator import validate_track_comprehensive
    vt = {day: "" for day in days}
    if selected_staff in st.session_state.track_changes:
        vt.update(st.session_state.track_changes[selected_staff])
    if preassignments:
        for day, pa in preassignments.items():
            if pa == "AT":
                vt[day] = "AT"
            elif pa in ["D", "N"]:
                vt[day] = pa
            else:
                vt[day] = "D"

    val_result = validate_track_comprehensive(
        vt, shifts_per_pay_period, night_minimum,
        weekend_minimum, preassignments, days, weekend_group,
        requirements_df, selected_staff
    )
    is_valid = val_result['overall_valid']
    st.session_state.modified_track['valid'] = is_valid
    if is_valid:
        st.success("Your bid meets all requirements! Go to Submission to finalize.")
    else:
        total_issues = sum(len(r.get('issues', [])) for k, r in val_result.items()
                          if k != 'overall_valid' and isinstance(r, dict) and not r.get('status', True))
        st.warning(f"Your bid has {total_issues} validation issues. Check the Validation tab.")


def _display_bid_submission(
    selected_staff, days, shifts_per_pay_period, night_minimum, weekend_minimum,
    preassignments, bid_track_name, bid_changes_key, bid_modified_key,
    preferences_df, staff_col_prefs, role_col, is_admin=False
):
    """Handle bid submission.

    Once a bid exists in the database for a staff member, that staff member can no
    longer resubmit it themselves — the view becomes read-only (download/email the
    PDF only), matching Summer Leave's "contact your supervisor to make changes"
    lock. "Already submitted" is read straight from the database (not a per-session
    flag), so the lock holds even across a fresh browser session. Admins bypass the
    lock entirely (used by the Add/Update Selection admin tab) so they can build or
    revise a bid on a staff member's behalf.
    """
    from modules.enhanced_track_validator import validate_track_comprehensive
    from modules.pdf_generator import generate_bid_summary_pdf
    from modules.email_notifications import send_bid_submission_notification, send_bid_summary_email

    st.subheader(f"Submit Bid for {selected_staff}")

    existing = get_bid_track_from_db(selected_staff, bid_track_name)
    has_existing_bid = existing[0]

    admin_notice_key = f'bid_admin_notice_{bid_track_name}_{selected_staff}'
    progression_notice_key = f'bid_progression_notice_{bid_track_name}_{selected_staff}'
    email_result_key = f'bid_email_result_{bid_track_name}_{selected_staff}'
    staff_confirmation_key = f'bid_staff_confirmation_{bid_track_name}_{selected_staff}'
    just_submitted_key = f'bid_just_submitted_{bid_track_name}_{selected_staff}'

    # Shown regardless of admin/staff path or lock state, so both a staff member
    # submitting their own bid and an admin submitting on their behalf see the
    # outcome of the admin notification and the automatic bid-progression attempt.
    _notice_fn = {"success": st.success, "warning": st.warning, "info": st.info}
    if admin_notice_key in st.session_state:
        notice_type, notice_msg = st.session_state[admin_notice_key]
        _notice_fn.get(notice_type, st.warning)(notice_msg)
    if progression_notice_key in st.session_state:
        notice_type, notice_msg = st.session_state[progression_notice_key]
        _notice_fn.get(notice_type, st.warning)(notice_msg)

    if has_existing_bid and not is_admin:
        # Locked: staff can't resubmit once a bid is on file for this cycle.
        saved_bid = existing[1]

        # The balloons animation is fired by the front end on this render — calling
        # st.balloons() on the same run as the st.rerun() that got us here doesn't
        # give it time to actually play, so the submit handler below just sets this
        # flag and reruns; we pop it and celebrate here on the render right after.
        if st.session_state.pop(just_submitted_key, False):
            st.balloons()

        st.success(
            f"Your bid for **{bid_track_name}** has been submitted "
            f"(version {saved_bid['version']}, submitted {saved_bid['submission_date']})."
        )
        st.info("This bid is locked. Please contact your supervisor if you need to make changes.")

        if st.session_state.get(staff_confirmation_key, False):
            st.info(
                "📧 A confirmation email with your bid summary PDF was sent to the email "
                "address on file for you. If you don't see it in your inbox shortly, please "
                "check your Junk/Spam folder."
            )

        weekend_group = st.session_state.get('weekend_group')
        validation_result = validate_track_comprehensive(
            saved_bid['track_data'], shifts_per_pay_period, night_minimum,
            weekend_minimum, preassignments, days, weekend_group,
            staff_name=selected_staff
        )
        pdf_bytes, pdf_filename = generate_bid_summary_pdf(
            selected_staff, saved_bid['track_data'], days, bid_track_name,
            saved_bid['version'], saved_bid['submission_date'],
            shifts_per_pay_period, night_minimum, weekend_minimum,
            preassignments, validation_result, weekend_group
        )

        st.markdown("### Bid Summary PDF")
        dl_col, email_col = st.columns(2)

        with dl_col:
            st.download_button(
                "Download Bid Summary PDF", data=pdf_bytes, file_name=pdf_filename,
                mime="application/pdf", use_container_width=True,
                key=f"download_bid_pdf_{bid_track_name}_{selected_staff}"
            )

        with email_col:
            with st.form(key=f"bid_email_form_{bid_track_name}_{selected_staff}"):
                email_addr = st.text_input("Email this summary to:", placeholder="you@example.com")
                send_clicked = st.form_submit_button("Send PDF to Email", use_container_width=True)
            if send_clicked:
                send_ok, send_msg = send_bid_summary_email(
                    email_addr, selected_staff, bid_track_name, pdf_bytes, pdf_filename
                )
                st.session_state[email_result_key] = ("success", send_msg) if send_ok else ("error", send_msg)

            if email_result_key in st.session_state:
                result_type, result_msg = st.session_state[email_result_key]
                (st.success if result_type == "success" else st.error)(result_msg)
        return

    # ── Editable flow: no bid yet, or an admin building/revising one on staff's behalf ──
    if is_admin and has_existing_bid:
        saved_bid = existing[1]
        st.info(
            f"**{selected_staff}** already has a bid on file (version {saved_bid['version']}, "
            f"submitted {saved_bid['submission_date']}). Saving below will create version "
            f"{saved_bid['version'] + 1}."
        )
    elif is_admin:
        st.info(f"**{selected_staff}** has not submitted a bid yet. Building a new bid on their behalf.")
    else:
        st.info(f"Submitting bid for **{bid_track_name}**.")

    modified_track = st.session_state.track_changes.get(selected_staff, {})
    valid = st.session_state.modified_track.get('valid', False)

    if valid:
        st.success("This bid meets all requirements and is ready to submit.")
    else:
        st.error("This bid has validation issues. Please fix them in Track Selection before submitting.")

    # Schedule preview
    st.markdown("### Schedule Preview")
    blocks = ["A", "B", "C"]
    block_tabs = st.tabs([f"Block {b}" for b in blocks])
    for bi, bt in enumerate(block_tabs):
        with bt:
            start = bi * 14
            end = start + 14
            block_days = days[start:end]
            tdata = []
            for day in block_days:
                assignment = modified_track.get(day, "")
                if not assignment and preassignments and day in preassignments:
                    assignment = preassignments[day]
                tdata.append({"Day": day, "Assignment": assignment if assignment else ""})
            st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)

    if not valid:
        st.error("Cannot submit — fix validation issues first.")
        return

    button_label = "Update Bid" if (is_admin and has_existing_bid) else "Submit Bid"
    if st.button(button_label, use_container_width=True, type="primary",
                 key=f"submit_bid_{bid_track_name}_{selected_staff}"):
        with st.spinner("Saving bid..."):
            # Build track to save
            track_to_save = modified_track.copy()
            if preassignments:
                for day, pa in preassignments.items():
                    if day not in track_to_save or not track_to_save[day]:
                        track_to_save[day] = pa

            # Get role
            staff_role = 'nurse'
            effective_role = 'nurse'
            if preferences_df is not None and staff_col_prefs and role_col:
                si = preferences_df[preferences_df[staff_col_prefs] == selected_staff]
                if not si.empty:
                    staff_role = si.iloc[0][role_col]
                    effective_role = "nurse" if str(staff_role).lower().strip() in ["nurse", "dual"] else "medic"

            meta = {
                'original_role': staff_role,
                'effective_role': effective_role,
                'track_source': 'Bid',
                'has_preassignments': bool(preassignments),
                'preassignment_count': len(preassignments) if preassignments else 0,
            }

            ok, msg, tid = save_bid_track_to_db(selected_staff, track_to_save, bid_track_name, meta)
            if ok:
                # Bid is now officially submitted — clear any saved in-progress draft
                delete_bid_draft(selected_staff, bid_track_name)

                # Notify the admin recipients with bid summary statistics (sent from the admin
                # account), also including the submitting staff member - with their bid summary
                # PDF attached - when their email is on file in Requirements.xlsx.
                try:
                    bid_result = get_bid_track_from_db(selected_staff, bid_track_name)
                    if bid_result[0]:
                        saved_bid = bid_result[1]
                        weekend_group = st.session_state.get('weekend_group')
                        validation_result = validate_track_comprehensive(
                            saved_bid['track_data'], shifts_per_pay_period, night_minimum,
                            weekend_minimum, preassignments, days, weekend_group,
                            staff_name=selected_staff
                        )

                        staff_email = None
                        req_ctx, _ = _load_bidding_data_files()
                        if req_ctx is not None:
                            staff_email = _load_requirements_map(req_ctx['requirements_df']).get(
                                selected_staff, {}).get('email')

                        notice_pdf_bytes, notice_pdf_filename = generate_bid_summary_pdf(
                            selected_staff, saved_bid['track_data'], days, bid_track_name,
                            saved_bid['version'], saved_bid['submission_date'],
                            shifts_per_pay_period, night_minimum, weekend_minimum,
                            preassignments, validation_result, weekend_group
                        )

                        admin_ok, admin_msg = send_bid_submission_notification(
                            selected_staff, bid_track_name, saved_bid['track_data'],
                            saved_bid['version'], saved_bid['submission_date'], validation_result,
                            staff_email=staff_email, pdf_bytes=notice_pdf_bytes, pdf_filename=notice_pdf_filename
                        )
                        st.session_state[admin_notice_key] = ("success", admin_msg) if admin_ok else ("warning", admin_msg)
                        st.session_state[staff_confirmation_key] = bool(staff_email) and admin_ok
                except Exception as e:
                    st.session_state[admin_notice_key] = ("warning", f"Admin notification failed: {e}")

                # Automatic bid access & notification: hand bid access to the next
                # staff member in seniority rank order, if the feature is turned on.
                try:
                    progression_result = _run_auto_bid_progression(selected_staff, bid_track_name)
                    if progression_result:
                        st.session_state[progression_notice_key] = progression_result
                    else:
                        st.session_state.pop(progression_notice_key, None)
                except Exception as e:
                    st.session_state[progression_notice_key] = ("warning", f"Automatic bid progression failed: {e}")

                st.success(f"Bid saved successfully! {msg}")
                if not is_admin:
                    st.session_state[just_submitted_key] = True
                st.rerun()
            else:
                st.error(f"Error: {msg}")


def _hypothetical_date_block_options():
    """
    Successive 6-week (42-day) calendar ranges, Sunday-to-Saturday, starting at the
    first Block A/Week 1 date (9/27/2026) through the block that contains 10/1/2027 —
    the choices offered by the Hypothetical Schedule tab's optional date overlay.

    Returns:
        list[tuple[date, date]]: (block_start, block_end) pairs, in order.
    """
    start = date(2026, 9, 27)
    target = date(2027, 10, 1)
    ranges = []
    cur = start
    while True:
        end = cur + timedelta(days=41)
        ranges.append((cur, end))
        if cur <= target <= end:
            break
        cur = end + timedelta(days=1)
    return ranges


def _display_hypothetical_track_by_blocks(shift_track, base_track, days, calendar_dates=None):
    """
    Block-by-block schedule table matching Current Track's layout (all 3 blocks shown
    one after another, not tabs), rendered as a fixed-width HTML table so the day
    columns can't be dragged, reordered, or resized. Two rows per block: Assignment
    (bare "D"/"N"/"AT") and Possible Shift (the hypothetical base code), both shaded
    to match that day's Assignment value.

    calendar_dates, if given, is a flat list of 42 date objects (one per day across
    all 3 blocks, in order) — an extra row of real dates is drawn above the weekday
    header for that block. Omitting it (the default) leaves that row out entirely.
    """
    blocks = ["A", "B", "C"]
    weekday_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    label_pct = 10.0
    day_pct = (100 - label_pct) / 14

    for block_idx, block in enumerate(blocks):
        st.markdown(f"#### Block {block} (Pay Period {block_idx + 1})")

        start_idx = block_idx * 14
        end_idx = start_idx + 14
        block_days = days[start_idx:end_idx]
        block_dates = calendar_dates[start_idx:end_idx] if calendar_dates else None

        day_headers = []
        for i in range(14):
            day_num = i % 7
            week_num = (block_idx * 2) + (i // 7) + 1
            day_headers.append(f"{weekday_names[day_num]} {block} {week_num}")

        header_row_count = 2 if block_dates else 1
        corner_cell = (
            f'<th rowspan="{header_row_count}" style="width:{label_pct}%; box-sizing:border-box; '
            'border:1px solid #ddd; background-color:#f0f2f6;"></th>'
        )

        # One combined header cell per day (e.g. "Mon A2") instead of a separate
        # weekday row and A1/A2/etc. tag row.
        combined_header_cells = "".join(
            f'<th style="width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; '
            f'background-color:#f0f2f6; font-size:12px; font-weight:500; color:#333; '
            f'text-align:center; padding:4px 0; white-space:nowrap;">{dh.split()[0]} {dh.split()[1]}{dh.split()[2]}</th>'
            for dh in day_headers
        )

        if block_dates:
            date_cells = "".join(
                f'<th style="width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; '
                f'background-color:#fff; font-size:12px; font-weight:700; color:#000; '
                f'text-align:center; padding:3px 0; white-space:nowrap;">{d.strftime("%-m/%-d")}</th>'
                for d in block_dates
            )
            date_row = f"<tr>{corner_cell}{date_cells}</tr>"
            combined_header_row = f"<tr>{combined_header_cells}</tr>"
        else:
            date_row = ""
            combined_header_row = f"<tr>{corner_cell}{combined_header_cells}</tr>"

        row_defs = [("Assignment", shift_track), ("Possible Shift", base_track)]
        body_rows = ""
        for row_label, source in row_defs:
            cells = ""
            for day in block_days:
                val = source.get(day, "")
                shift_val = shift_track.get(day, "")
                style = (
                    f"width:{day_pct}%; box-sizing:border-box; border:1px solid #ddd; font-size:13px; "
                    "text-align:center; padding:5px 1px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;"
                )
                if shift_val == "D":
                    style += " background-color:#d4edda;"
                elif shift_val == "N":
                    style += " background-color:#cce5ff;"
                elif shift_val == "AT":
                    style += " background-color:#e2e3e5; font-weight:bold;"
                cells += f'<td style="{style}">{val}</td>'
            label_cell = (
                f'<td style="width:{label_pct}%; box-sizing:border-box; border:1px solid #ddd; '
                'background-color:#f0f2f6; font-size:13px; font-weight:500; text-align:center; '
                f'padding:5px 2px; white-space:nowrap;">{row_label}</td>'
            )
            body_rows += f"<tr>{label_cell}{cells}</tr>"

        # thead content is joined into one string (not spread across template lines)
        # because an empty date_row leaves a blank line, and CommonMark treats a
        # blank line inside an HTML block as ending it — everything after would
        # render as literal escaped text instead of an actual table.
        thead_content = date_row + combined_header_row
        st.markdown(
            f'<table style="width:100%; border-collapse:collapse; table-layout:fixed;">'
            f'<thead>{thead_content}</thead>'
            f'<tbody>{body_rows}</tbody>'
            f'</table>',
            unsafe_allow_html=True
        )
        st.write("")


_NOT_SUBMITTED_BANNER_HTML = (
    '<div style="background-color:#f8d7da; border:2px solid #dc3545; border-radius:5px; '
    'padding:10px; text-align:center; font-weight:700; color:#721c24; margin-bottom:10px;">'
    'PROPOSED TRACK NOT YET SUBMITTED</div>'
)


def _display_bid_hypothetical(
    selected_staff, preferences_df, current_tracks_df, days,
    staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
    reduced_rest_col, seniority_col, capacity, bid_track_name,
    has_bid, bid_result, staff_preassignments, current_working_track
):
    """
    Hypothetical schedule for bidding — shows expected assignments only for the
    days/nights actually present (as D or N) in the staff member's track, not every
    day the role has an opening. Works off the actual submitted bid once there is
    one; otherwise falls back to the staff member's saved-but-not-submitted working
    selections (current_working_track) so this tab doesn't require submission first,
    with a red banner top and bottom noting the track isn't submitted yet.
    """
    from modules.track_modification_core import calculate_all_modification_options
    from modules.db_utils import get_location_preferences_from_db

    st.subheader(f"Hypothetical Schedule for {selected_staff}")
    st.info(
        "**Note:** The 'Hypothetical Schedule' are not guaranteed base assignments — your "
        "proposed track only designates a \"D\" or \"N\" for each day. The expected base "
        "assignments shown here do not account for 10-hour rest or \"clock moving forward\" "
        "rules, and are for demonstration purposes only."
    )

    if not has_bid:
        st.markdown(_NOT_SUBMITTED_BANNER_HTML, unsafe_allow_html=True)

    if has_bid:
        submitted_track = bid_result[1]['track_data']
        bid_version = bid_result[1]['version']
        bid_submission_date = bid_result[1]['submission_date']
    else:
        submitted_track = current_working_track
        bid_version = "Draft"
        bid_submission_date = "not yet submitted"

    # Only the days the staff actually bid a working shift on — excludes Off days
    # and AT (preassigned days that aren't a D/N shift are covered on their own).
    bid_work_days = {day: shift for day, shift in submitted_track.items() if shift in ("D", "N")}

    if not bid_work_days:
        st.info(
            f"No working days selected yet for **{bid_track_name}**. Pick some D/N shifts "
            "in the **Track Selection** tab to see your hypothetical schedule here."
        )
        if not has_bid:
            st.markdown(_NOT_SUBMITTED_BANNER_HTML, unsafe_allow_html=True)
        return

    max_dn = capacity['max_day_nurses']
    max_dm = capacity['max_day_medics']
    max_nn = capacity['max_night_nurses']
    max_nm = capacity['max_night_medics']

    with st.spinner("Generating hypothetical schedule..."):
        results = calculate_all_modification_options(
            selected_staff, preferences_df, current_tracks_df, days,
            staff_col_prefs, staff_col_tracks, role_col, no_matrix_col,
            reduced_rest_col, seniority_col,
            max_day_nurses=max_dn, max_day_medics=max_dm,
            max_night_nurses=max_nn, max_night_medics=max_nm,
            bid_track_name=bid_track_name
        )

    day_assignments = results['day_assignments']
    night_assignments = results['night_assignments']

    staff_info = preferences_df[preferences_df[staff_col_prefs] == selected_staff].iloc[0]
    staff_role = staff_info[role_col]
    effective_role = "nurse" if str(staff_role).lower().strip() in ["nurse", "dual"] else "medic"

    has_base_prefs, _ = get_location_preferences_from_db(selected_staff)
    if has_base_prefs:
        st.success("Using your base preferences for this hypothetical schedule.")
    else:
        st.warning("No base preferences found. Set them in Edit Preferences for better results.")

    st.markdown("### Hypothetical Schedule")
    if has_bid:
        st.caption(
            f"Expected assignments for the days you bid a shift on your submitted "
            f"**{bid_track_name}** bid (version {bid_version}, submitted {bid_submission_date})."
        )
    else:
        st.caption(
            f"Expected assignments for the days you've selected a shift on so far for "
            f"**{bid_track_name}** — this bid has not been submitted yet."
        )

    total_day = sum(1 for day, shift in bid_work_days.items() if shift == "D" and day_assignments.get(day))
    total_night = sum(1 for day, shift in bid_work_days.items() if shift == "N" and night_assignments.get(day))
    sc = st.columns(3)
    sc[0].metric("Total Shifts", total_day + total_night)
    sc[1].metric("Day Shifts", total_day)
    sc[2].metric("Night Shifts", total_night)

    with st.expander("Simulate for specific dates?", expanded=False):
        dates_acknowledged = st.checkbox(
            "I acknowledge the hypothetical schedule displayed with dates will be for "
            "general reference and mapping out proposed track shifts only, and base "
            "assignments are not guaranteed as displayed. 9/27/26 will be the first "
            "date displayed as the first day of a 6-week block, but this does not mean "
            "new tracks will launch on that date. All tracks displayed here are "
            "proposed only until FY27 rebids are finalized.",
            key="hypo_dates_ack"
        )
        date_block_options = _hypothetical_date_block_options()
        date_option_labels = [
            f"{s.strftime('%-m/%-d/%y')} to {e.strftime('%-m/%-d/%y')}" for s, e in date_block_options
        ]
        selected_date_label = st.selectbox(
            "6-week block to map onto the schedule below:",
            date_option_labels, key="hypo_dates_range", disabled=not dates_acknowledged
        )

    calendar_dates = None
    if dates_acknowledged:
        block_start, _ = date_block_options[date_option_labels.index(selected_date_label)]
        calendar_dates = [block_start + timedelta(days=i) for i in range(42)]

    # Build track-shaped dicts like Current Track's: shift_track holds the bare
    # D/N/AT for the Assignment row, base_track holds the expected base (D/N days
    # with a resolved hypothetical assignment only) for the Possible Shift row.
    shift_track = {}
    base_track = {}
    for day in days:
        shift = submitted_track.get(day, "")
        if shift == "D":
            shift_track[day] = "D"
            assignment = day_assignments.get(day)
            if assignment:
                base_track[day] = assignment
        elif shift == "N":
            shift_track[day] = "N"
            assignment = night_assignments.get(day)
            if assignment:
                base_track[day] = assignment
        elif shift:
            shift_track[day] = shift  # e.g. "AT"

    _display_hypothetical_track_by_blocks(shift_track, base_track, days, calendar_dates)

    from modules.pdf_generator import generate_hypothetical_schedule_pdf
    pdf_bytes, pdf_filename = generate_hypothetical_schedule_pdf(
        selected_staff, shift_track, base_track, days,
        bid_track_name, bid_version, bid_submission_date
    )
    st.download_button(
        "📄 Download Hypothetical Schedule (PDF)",
        data=pdf_bytes, file_name=pdf_filename, mime="application/pdf"
    )

    if not has_bid:
        st.markdown(_NOT_SUBMITTED_BANNER_HTML, unsafe_allow_html=True)
