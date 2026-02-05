# fiscal_year_module.py - ENHANCED WITH STAFF MEMBER FILTERING
"""
Fiscal Year 2026 Track Display Module - ENHANCED WITH STAFF FILTERING
To be integrated with existing app.py
Place this file in your project root or modules/ directory
UPDATED: Added staff member filtering functionality alongside role filtering
UPDATED: Added Master Tracks tab and Version tab to Excel export
"""

import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import base64
import json
import pytz

_eastern_tz = pytz.timezone('America/New_York')

class FiscalYearDisplay:
    """Fiscal Year display component for integration with existing app"""
    
    def __init__(self, tracks_db_path='data/medflight_tracks.db'):
        self.db_path = tracks_db_path
        self.fiscal_year_start = datetime(2025, 9, 28)
        self.fiscal_year_end = datetime(2026, 9, 26)
        self.pattern_start = datetime(2025, 9, 14)  # Sun A 1
        self.pattern_length = 42  # 6 weeks
        
        # US Holidays
        self.holidays = {
            datetime(2025, 11, 27): "Thanksgiving",
            datetime(2025, 12, 24): "Christmas Eve",
            datetime(2025, 12, 25): "Christmas",
            datetime(2026, 1, 1): "New Year's Day",
            datetime(2026, 1, 19): "MLK Jr. Day",
            datetime(2026, 2, 16): "Presidents' Day",
            datetime(2026, 5, 25): "Memorial Day",
            datetime(2026, 6, 19): "Juneteenth",
            datetime(2026, 7, 4): "Independence Day",
            datetime(2026, 9, 7): "Labor Day"
        }
        
        self.shift_descriptions = {
            'D': 'Day Shift',
            'N': 'Night Shift',
            'AT': 'Admin Time',
            '': 'Off'
        }
        
        # Role mappings for filtering
        self.role_mappings = {
            'nurse': 'Nurse',
            'medic': 'Medic',
            'Nurse': 'Nurse',
            'Medic': 'Medic'
        }
    
    def load_tracks_from_db(self):
        """Load tracks from the existing medflight_tracks.db with proper role support"""
        tracks_data = {}
        staff_roles = {}
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get the latest active tracks for each staff member with role information
            cursor.execute("""
                SELECT staff_name, track_data, effective_role, submission_date
                FROM tracks
                WHERE is_active = 1
                ORDER BY staff_name, submission_date DESC
            """)
            
            results = cursor.fetchall()
            processed_staff = set()
            
            for staff_name, track_data_str, effective_role, submission_date in results:
                # Only take the latest entry for each staff member
                if staff_name not in processed_staff:
                    try:
                        # Parse the JSON track data
                        track_data = json.loads(track_data_str)
                        tracks_data[staff_name] = track_data
                        
                        # Use the effective_role from database, with fallback
                        role = self.normalize_role(effective_role)
                        staff_roles[staff_name] = role
                        
                        processed_staff.add(staff_name)
                        
                    except json.JSONDecodeError:
                        st.warning(f"Error parsing track data for {staff_name}")
                        continue
            
            conn.close()
            
            if not tracks_data:
                st.info("No active track data found in database.")
                
        except sqlite3.Error as e:
            st.error(f"Database error loading tracks: {str(e)}")
        except Exception as e:
            st.error(f"Error loading tracks from database: {str(e)}")
        
        return tracks_data, staff_roles
    
    def normalize_role(self, role):
        """Normalize role names for consistent display"""
        if not role:
            return "Unknown"
        
        role_lower = role.lower().strip()
        
        # Handle common variations
        if role_lower in ['nurse', 'rn']:
            return 'Nurse'
        elif role_lower in ['medic', 'paramedic', 'emt']:
            return 'Medic'
        else:
            # Capitalize first letter for display
            return role.strip().title()
    
    def get_pattern_day_name(self, date):
        """Get the pattern day name for a given date"""
        if isinstance(date, str):
            date = datetime.strptime(date, '%m/%d/%Y')
        
        days_since_start = (date - self.pattern_start).days
        pattern_day_index = days_since_start % self.pattern_length
        
        week_index = pattern_day_index // 7
        day_index = pattern_day_index % 7
        
        days_of_week = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        week_letters = ["A", "A", "B", "B", "C", "C"]
        week_numbers = [1, 2, 3, 4, 5, 6]
        
        if week_index < 6:
            day_name = days_of_week[day_index]
            week_letter = week_letters[week_index]
            week_number = week_numbers[week_index]
            return f"{day_name} {week_letter} {week_number}"
        
        return f"Day {pattern_day_index + 1}"
    
    def is_pay_period_end(self, date):
        """Check if date is end of pay period"""
        pattern_day = self.get_pattern_day_name(date)
        return 'Sat' in pattern_day and any(f' {n}' in pattern_day for n in ['2', '4', '6'])
    
    def get_fiscal_year_months(self):
        """Get list of months in fiscal year"""
        months = []
        current_date = self.fiscal_year_start
        
        while current_date <= self.fiscal_year_end:
            if current_date == self.fiscal_year_start:
                month_start = current_date
            else:
                month_start = current_date.replace(day=1)
            
            next_month = current_date.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
            
            if month_end > self.fiscal_year_end:
                month_end = self.fiscal_year_end
            
            months.append({
                'name': current_date.strftime('%B %Y'),
                'start': month_start,
                'end': month_end
            })
            
            current_date = month_end + timedelta(days=1)
        
        return months
    
    def get_role_statistics(self, staff_roles):
        """Get statistics about staff roles"""
        if not staff_roles:
            return {}
        
        role_counts = {}
        for role in staff_roles.values():
            role_counts[role] = role_counts.get(role, 0) + 1
        
        total_staff = len(staff_roles)
        role_stats = {}
        
        for role, count in role_counts.items():
            role_stats[role] = {
                'count': count,
                'percentage': round((count / total_staff) * 100, 1)
            }
        
        return role_stats
    
    def filter_staff_by_criteria(self, staff_list, staff_roles, role_filter, staff_filter):
        """Filter staff by both role and individual staff member criteria"""
        filtered_staff = list(staff_list)
        
        # Apply role filter first
        if role_filter != 'All Roles':
            filtered_staff = [s for s in filtered_staff if staff_roles.get(s) == role_filter]
        
        # Apply staff member filter
        if staff_filter != 'All Staff':
            if staff_filter == 'Selected Staff':
                # Get selected staff from session state
                selected_staff = st.session_state.get('fy_selected_staff', [])
                filtered_staff = [s for s in filtered_staff if s in selected_staff]
            else:
                # Single staff member selected
                filtered_staff = [s for s in filtered_staff if s == staff_filter]
        
        return filtered_staff
    
    def display_staff_selection_interface(self, available_staff, staff_roles):
        """Display multi-select interface for staff selection"""
        if 'fy_selected_staff' not in st.session_state:
            st.session_state.fy_selected_staff = []
        
        # Group staff by role for better organization
        staff_by_role = {}
        for staff in available_staff:
            role = staff_roles.get(staff, 'Unknown')
            if role not in staff_by_role:
                staff_by_role[role] = []
            staff_by_role[role].append(staff)
        
        # Sort staff within each role
        for role in staff_by_role:
            staff_by_role[role].sort()
        
        st.markdown("**Select specific staff members to view:**")
        
        # Quick selection buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("Select All", key="select_all_staff"):
                st.session_state.fy_selected_staff = available_staff.copy()
                st.rerun()
        with col2:
            if st.button("Clear All", key="clear_all_staff"):
                st.session_state.fy_selected_staff = []
                st.rerun()
        with col3:
            st.caption(f"Selected: {len(st.session_state.fy_selected_staff)}")
        
        # Multi-select organized by role
        for role in sorted(staff_by_role.keys()):
            with st.expander(f"👥 {role} ({len(staff_by_role[role])})"):
                role_staff = staff_by_role[role]
                
                # Role-specific select all/none
                role_col1, role_col2 = st.columns(2)
                with role_col1:
                    if st.button(f"Select All {role}", key=f"select_role_{role}"):
                        for staff in role_staff:
                            if staff not in st.session_state.fy_selected_staff:
                                st.session_state.fy_selected_staff.append(staff)
                        st.rerun()
                with role_col2:
                    if st.button(f"Clear {role}", key=f"clear_role_{role}"):
                        st.session_state.fy_selected_staff = [
                            s for s in st.session_state.fy_selected_staff if s not in role_staff
                        ]
                        st.rerun()
                
                # Individual checkboxes
                for staff in role_staff:
                    is_selected = staff in st.session_state.fy_selected_staff
                    if st.checkbox(
                        staff, 
                        value=is_selected, 
                        key=f"staff_check_{staff}_{role}"
                    ):
                        if staff not in st.session_state.fy_selected_staff:
                            st.session_state.fy_selected_staff.append(staff)
                    else:
                        if staff in st.session_state.fy_selected_staff:
                            st.session_state.fy_selected_staff.remove(staff)
    
    def display_fiscal_year_section(self):
        """Main display function to be called from app.py"""
        
        # Initialize session state for fullscreen if not exists
        if 'fy_show_fullscreen' not in st.session_state:
            st.session_state.fy_show_fullscreen = False
        
        # FIXED: Use st.empty() to replace content instead of conditional rendering
        content_container = st.empty()
        
        # Check if we're in fullscreen mode
        if st.session_state.fy_show_fullscreen:
            with content_container.container():
                self.show_fullscreen_view()
        else:
            with content_container.container():
                self.show_normal_view()
    
    def show_normal_view(self):
        """Show normal (non-fullscreen) view"""
        st.markdown('<h2 style="color: #1f2937; border-bottom: 3px solid #e5e7eb; padding-bottom: 0.5rem;">📊 Fiscal Year 2026 Track Display</h2>', unsafe_allow_html=True)
        st.caption("View tracks from September 28, 2025 through September 26, 2026")
        
        # Load tracks from database
        tracks_data, staff_roles = self.load_tracks_from_db()
        
        if not tracks_data:
            st.warning("No track data found in database. Please ensure tracks have been submitted and approved.")
            return
        
        # Initialize session state for filters
        if 'fy_role_filter' not in st.session_state:
            st.session_state.fy_role_filter = 'All Roles'
        if 'fy_staff_filter' not in st.session_state:
            st.session_state.fy_staff_filter = 'All Staff'
        
        # Available options for filters
        available_roles = ['All Roles'] + sorted(list(set(staff_roles.values())))
        available_staff = sorted(list(tracks_data.keys()))
        staff_filter_options = ['All Staff', 'Selected Staff'] + available_staff
        
        # Filter interface
        st.markdown("### 🔍 Filter Options")
        
        filter_col1, filter_col2 = st.columns(2)
        
        with filter_col1:
            st.markdown("**Filter by Role:**")
            role_filter = st.selectbox(
                "",
                options=available_roles,
                index=available_roles.index(st.session_state.fy_role_filter) if st.session_state.fy_role_filter in available_roles else 0,
                key="fy_role_select",
                label_visibility="collapsed"
            )
            st.session_state.fy_role_filter = role_filter
        
        with filter_col2:
            st.markdown("**Filter by Staff:**")
            staff_filter = st.selectbox(
                "",
                options=staff_filter_options,
                index=staff_filter_options.index(st.session_state.fy_staff_filter) if st.session_state.fy_staff_filter in staff_filter_options else 0,
                key="fy_staff_select",
                label_visibility="collapsed"
            )
            st.session_state.fy_staff_filter = staff_filter
        
        # Show staff selection interface if "Selected Staff" is chosen
        if staff_filter == 'Selected Staff':
            with st.expander("👥 Select Staff Members", expanded=True):
                self.display_staff_selection_interface(available_staff, staff_roles)
        
        # Filter summary
        filtered_staff = self.filter_staff_by_criteria(
            available_staff, staff_roles, role_filter, staff_filter
        )
        
        if filtered_staff:
            filter_summary = f"Showing {len(filtered_staff)} staff member(s)"
            if role_filter != 'All Roles':
                filter_summary += f" • Role: {role_filter}"
            if staff_filter not in ['All Staff']:
                if staff_filter == 'Selected Staff':
                    filter_summary += f" • Custom Selection"
                else:
                    filter_summary += f" • Staff: {staff_filter}"
            st.info(filter_summary)
        else:
            st.warning("No staff members match the selected filters.")
            return
        
        # Fullscreen button
        if st.button("⛶ Fullscreen", key="fy_fullscreen", type="primary", use_container_width=True):
            st.session_state.fy_show_fullscreen = True
            st.rerun()
        
        # Normal view - Get fiscal year months
        fiscal_months = self.get_fiscal_year_months()
        
        # Display months in expandable sections
        st.markdown("### 📅 Monthly Schedules")
        
        for month_info in fiscal_months:
            with st.expander(f"📆 {month_info['name']}"):
                self.display_month_schedule(
                    month_info, tracks_data, staff_roles, 
                    role_filter, staff_filter
                )
    
    def show_fullscreen_view(self):
        """Show fullscreen view - ENHANCED VERSION"""
        st.markdown("### 📊 Fiscal Year 2026 - Fullscreen View")
        
        # Back button
        back_col, spacer_col = st.columns([1, 4])
        with back_col:
            if st.button("← Back to Normal View", key="fy_back_to_normal", type="secondary", use_container_width=True):
                st.session_state.fy_show_fullscreen = False
                st.rerun()
        
        # Load tracks from database
        tracks_data, staff_roles = self.load_tracks_from_db()
        
        if not tracks_data:
            st.warning("No track data found in database.")
            return
        
        # Get current filters from session state
        current_role_filter = st.session_state.get('fy_role_filter', 'All Roles')
        current_staff_filter = st.session_state.get('fy_staff_filter', 'All Staff')
        
        # Available options
        available_roles = ['All Roles'] + sorted(list(set(staff_roles.values())))
        available_staff = sorted(list(tracks_data.keys()))
        staff_filter_options = ['All Staff', 'Selected Staff'] + available_staff
        
        # Filter interface in fullscreen
        with st.expander("🔍 Filter Options", expanded=False):
            filter_col1, filter_col2 = st.columns(2)
            
            with filter_col1:
                role_filter = st.selectbox(
                    "Filter by Role:",
                    options=available_roles,
                    index=available_roles.index(current_role_filter) if current_role_filter in available_roles else 0,
                    key="fy_fullscreen_role_select"
                )
                st.session_state.fy_role_filter = role_filter
            
            with filter_col2:
                staff_filter = st.selectbox(
                    "Filter by Staff:",
                    options=staff_filter_options,
                    index=staff_filter_options.index(current_staff_filter) if current_staff_filter in staff_filter_options else 0,
                    key="fy_fullscreen_staff_select"
                )
                st.session_state.fy_staff_filter = staff_filter
            
            # Show staff selection interface if needed
            if staff_filter == 'Selected Staff':
                self.display_staff_selection_interface(available_staff, staff_roles)
        
        # Display current filter status
        filtered_staff = self.filter_staff_by_criteria(
            available_staff, staff_roles, current_role_filter, current_staff_filter
        )
        
        if filtered_staff:
            filter_summary = f"📊 Currently showing {len(filtered_staff)} staff member(s)"
            if current_role_filter != 'All Roles':
                filter_summary += f" • Role: {current_role_filter}"
            if current_staff_filter not in ['All Staff']:
                if current_staff_filter == 'Selected Staff':
                    filter_summary += f" • Custom Selection ({len(st.session_state.get('fy_selected_staff', []))} selected)"
                else:
                    filter_summary += f" • Staff: {current_staff_filter}"
            st.info(filter_summary)
        else:
            st.warning("No staff members match the selected filters.")
            return
        
        fiscal_months = self.get_fiscal_year_months()
        month_names = [m['name'] for m in fiscal_months]
        
        # Create tabs for each month
        tabs = st.tabs(month_names)
        
        for idx, tab in enumerate(tabs):
            with tab:
                month_info = fiscal_months[idx]
                self.display_month_schedule(
                    month_info, 
                    tracks_data, 
                    staff_roles, 
                    current_role_filter,
                    current_staff_filter
                )

    def display_month_schedule(self, month_info, tracks_data, staff_roles, role_filter, staff_filter):
        """Display schedule for a single month with enhanced filtering"""
        # Filter staff by both role and staff criteria
        all_staff = list(tracks_data.keys())
        filtered_staff = self.filter_staff_by_criteria(
            all_staff, staff_roles, role_filter, staff_filter
        )
        
        if not filtered_staff:
            st.info(f"No staff found matching the selected filters.")
            return
        
        # Build date columns
        dates = []
        current_date = month_info['start']
        while current_date <= month_info['end']:
            dates.append({
                'date': current_date,
                'day': current_date.day,
                'weekday': current_date.strftime('%a'),
                'pattern_day': self.get_pattern_day_name(current_date),
                'is_holiday': current_date in self.holidays,
                'holiday_name': self.holidays.get(current_date, ''),
                'is_pay_period_end': self.is_pay_period_end(current_date)
            })
            current_date += timedelta(days=1)
        
        # Build DataFrame
        data = []
        columns = ['Staff', 'Role']
        
        for date_info in dates:
            col_name = f"{date_info['weekday']} {date_info['day']}"
            if date_info['is_holiday']:
                col_name += " 🎉"
            if date_info['is_pay_period_end']:
                col_name += " 💰"
            columns.append(col_name)
        
        # Sort staff for consistent display
        filtered_staff.sort()
        
        for staff_name in filtered_staff:
            row = [staff_name, staff_roles.get(staff_name, 'Unknown')]
            
            for date_info in dates:
                pattern_day = date_info['pattern_day']
                shift = tracks_data.get(staff_name, {}).get(pattern_day, '')
                row.append(shift)
            
            data.append(row)
        
        if data:
            df = pd.DataFrame(data, columns=columns)
            
            # Apply styling
            def style_shifts(val):
                if val == 'D':
                    return 'background-color: #dbeafe; color: #1e40af; font-weight: bold;'
                elif val == 'N':
                    return 'background-color: #ede9fe; color: #6b21a8; font-weight: bold;'
                elif val == 'AT':
                    return 'background-color: #d1fae5; color: #065f46; font-weight: bold;'
                elif val == '':
                    return 'background-color: #f9fafb; color: #6b7280;'
                return ''
            
            styled_df = df.style.applymap(style_shifts, subset=columns[2:])
            
            # Use larger height in fullscreen mode
            height = 600 if st.session_state.get('fy_show_fullscreen', False) else 400
            
            st.dataframe(styled_df, use_container_width=True, height=height, hide_index=True)
            
            # Show legend and statistics
            col1, col2 = st.columns(2)
            with col1:
                st.caption("🟦 **D** = Day Shift | 🟪 **N** = Night Shift | 🟢 **AT** = Admin Time")
            with col2:
                st.caption("🎉 = Holiday | 💰 = Pay Period End")
            
            # Show holidays and statistics
            holidays_in_month = [d for d in dates if d['is_holiday']]
            if holidays_in_month:
                holiday_text = ", ".join([f"{h['holiday_name']} ({h['day']})" for h in holidays_in_month])
                st.caption(f"🎉 **Holidays this month:** {holiday_text}")
            
            # Enhanced staff count display
            staff_count_text = f"👥 **Showing {len(filtered_staff)} staff member(s)**"
            if role_filter != 'All Roles':
                staff_count_text += f" ({role_filter})"
            if staff_filter not in ['All Staff']:
                if staff_filter == 'Selected Staff':
                    staff_count_text += f" - Custom Selection"
                else:
                    staff_count_text += f" - Individual: {staff_filter}"
            
            st.caption(staff_count_text)
            
            # Show breakdown by role for current filter
            if len(filtered_staff) > 1:
                role_breakdown = {}
                for staff in filtered_staff:
                    role = staff_roles.get(staff, 'Unknown')
                    role_breakdown[role] = role_breakdown.get(role, 0) + 1
                
                breakdown_text = " | ".join([f"{role}: {count}" for role, count in sorted(role_breakdown.items())])
                st.caption(f"📊 **Role breakdown:** {breakdown_text}")
                
        else:
            st.info("No staff data available for selected filters")
    
    def export_and_download(self, tracks_data, staff_roles, role_filter=None, staff_filter=None):
        """Generate and provide download for Excel export with filtering support"""
        file_path, message = self.export_to_excel(role_filter, staff_filter)
        
        if file_path and os.path.exists(file_path):
            try:
                with open(file_path, "rb") as f:
                    excel_data = f.read()
                
                b64 = base64.b64encode(excel_data).decode()
                href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="FY2026_Tracks_Filtered.xlsx">Click here to download Excel file</a>'
                
                st.success(f"✅ {message}")
                st.markdown(href, unsafe_allow_html=True)
                
                # Cleanup
                os.remove(file_path)
                
            except Exception as e:
                st.error(f"❌ Error preparing download: {str(e)}")
        else:
            st.error(f"❌ {message}")
    
    def export_to_excel(self, role_filter=None, staff_filter=None):
        """Export fiscal year to Excel with role and staff filtering, Master Tracks tab, and Version tab"""
        tracks_data, staff_roles = self.load_tracks_from_db()
        
        if not tracks_data:
            return None, "No track data available"
        
        # Apply filters if provided
        if role_filter or staff_filter:
            all_staff = list(tracks_data.keys())
            filtered_staff = self.filter_staff_by_criteria(
                all_staff, staff_roles, 
                role_filter or 'All Roles', 
                staff_filter or 'All Staff'
            )
            
            # Filter tracks_data and staff_roles
            tracks_data = {staff: tracks_data[staff] for staff in filtered_staff if staff in tracks_data}
            staff_roles = {staff: staff_roles[staff] for staff in filtered_staff if staff in staff_roles}
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Store creation timestamp in Eastern Time
            eastern = pytz.timezone('US/Eastern')
            now_eastern = datetime.now(eastern)
            # Format with timezone abbreviation (EST or EDT)
            creation_timestamp = now_eastern.strftime("%Y-%m-%d %H:%M:%S %Z")
            
            # Define styles
            header_font = Font(bold=True, size=12)
            role_font = Font(bold=True, size=10)
            holiday_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            pay_period_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
            
            shift_fills = {
                'D': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
                'N': PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid"),
                'AT': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            }
            
            role_fills = {
                'Nurse': PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
                'Medic': PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")
            }
            
            # ========================================
            # TAB 1: MASTER TRACKS (42-Day Pattern)
            # ========================================
            master_ws = wb.create_sheet(title="Master Tracks", index=0)
            
            # Generate 42 pattern days
            pattern_days = []
            days_of_week = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
            week_letters = ["A", "A", "B", "B", "C", "C"]
            week_numbers = [1, 2, 3, 4, 5, 6]
            
            for week_idx in range(6):
                for day_idx in range(7):
                    pattern_day = f"{days_of_week[day_idx]} {week_letters[week_idx]} {week_numbers[week_idx]}"
                    pattern_days.append(pattern_day)
            
            # Headers for Master Tracks
            master_ws['A1'] = 'Staff Name'
            master_ws['B1'] = 'Role'
            master_ws['A1'].font = header_font
            master_ws['B1'].font = header_font
            
            # Pattern day headers
            for col_idx, pattern_day in enumerate(pattern_days, start=3):
                cell = master_ws.cell(row=1, column=col_idx, value=pattern_day)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', text_rotation=90)
                master_ws.column_dimensions[get_column_letter(col_idx)].width = 4
            
            # Sort staff by role then name
            sorted_staff = sorted(tracks_data.keys(), key=lambda x: (staff_roles.get(x, 'Unknown'), x))
            
            # Populate Master Tracks data
            row_idx = 2
            for staff_name in sorted_staff:
                role = staff_roles.get(staff_name, 'Unknown')
                
                # Staff name and role
                name_cell = master_ws.cell(row=row_idx, column=1, value=staff_name)
                role_cell = master_ws.cell(row=row_idx, column=2, value=role)
                
                # Apply role-based background color
                if role in role_fills:
                    name_cell.fill = role_fills[role]
                    role_cell.fill = role_fills[role]
                
                # Fill in shifts for each pattern day
                for col_idx, pattern_day in enumerate(pattern_days, start=3):
                    shift = tracks_data.get(staff_name, {}).get(pattern_day, '')
                    cell = master_ws.cell(row=row_idx, column=col_idx, value=shift)
                    
                    # Apply shift-based styling
                    if shift in shift_fills:
                        cell.fill = shift_fills[shift]
                    
                    cell.alignment = Alignment(horizontal='center')
                
                row_idx += 1
            
            # Freeze panes for Master Tracks
            master_ws.freeze_panes = 'C2'
            
            # Set column widths
            master_ws.column_dimensions['A'].width = 18
            master_ws.column_dimensions['B'].width = 12
            
            fiscal_months = self.get_fiscal_year_months()
            
            # ========================================
            # MONTHLY SHEETS
            # ========================================
            # Create sheet for each month
            for month_info in fiscal_months:
                month_name = month_info['name'].replace(' ', '_')
                ws = wb.create_sheet(title=month_name[:31])
                
                # Headers
                ws['A1'] = 'Staff Name'
                ws['B1'] = 'Role'
                ws['A1'].font = header_font
                ws['B1'].font = header_font
                
                # Date columns
                col_idx = 3
                current_date = month_info['start']
                date_cols = []
                
                while current_date <= month_info['end']:
                    # Day of week
                    ws.cell(row=1, column=col_idx, value=current_date.strftime('%a')).font = header_font
                    # Day number
                    ws.cell(row=2, column=col_idx, value=current_date.day).font = header_font
                    
                    # Pattern day
                    pattern_day = self.get_pattern_day_name(current_date)
                    pattern_parts = pattern_day.split()
                    if len(pattern_parts) >= 3:
                        ws.cell(row=3, column=col_idx, value=f"{pattern_parts[1]} {pattern_parts[2]}")
                    
                    # Mark holidays
                    if current_date in self.holidays:
                        for row_num in range(1, 4):
                            ws.cell(row=row_num, column=col_idx).fill = holiday_fill
                            ws.cell(row=row_num, column=col_idx).comment = Comment(
                                self.holidays[current_date], "System"
                            )
                    
                    # Mark pay period ends
                    if self.is_pay_period_end(current_date):
                        for row_num in range(1, 4):
                            cell = ws.cell(row=row_num, column=col_idx)
                            if cell.fill.start_color.rgb != holiday_fill.start_color.rgb:
                                cell.fill = pay_period_fill
                    
                    date_cols.append((current_date, pattern_day))
                    col_idx += 1
                    current_date += timedelta(days=1)
                
                # Staff data - sort by role then name
                sorted_staff = sorted(tracks_data.keys(), key=lambda x: (staff_roles.get(x, ''), x))
                
                row_idx = 4
                for staff_name in sorted_staff:
                    role = staff_roles.get(staff_name, 'Unknown')
                    
                    # Staff name and role
                    name_cell = ws.cell(row=row_idx, column=1, value=staff_name)
                    role_cell = ws.cell(row=row_idx, column=2, value=role)
                    
                    # Apply role-based background color
                    if role in role_fills:
                        name_cell.fill = role_fills[role]
                        role_cell.fill = role_fills[role]
                    
                    # Shift assignments
                    for col_offset, (date, pattern_day) in enumerate(date_cols):
                        shift = tracks_data.get(staff_name, {}).get(pattern_day, '')
                        cell = ws.cell(row=row_idx, column=3 + col_offset, value=shift)
                        
                        # Apply shift-based styling
                        if shift in shift_fills:
                            cell.fill = shift_fills[shift]
                        
                        cell.alignment = Alignment(horizontal='center')
                    
                    row_idx += 1
                
                # Freeze panes
                ws.freeze_panes = 'C4'
                
                # Column widths
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 12
                for col in range(3, col_idx):
                    ws.column_dimensions[get_column_letter(col)].width = 6
            
            # ========================================
            # LAST TAB: VERSION
            # ========================================
            version_ws = wb.create_sheet(title="Version")
            version_ws['A1'] = f"Created: {creation_timestamp}"
            version_ws['A1'].font = Font(bold=True, size=11)
            version_ws.column_dimensions['A'].width = 30
            
            # Save to temp file
            export_dir = 'exports'
            os.makedirs(export_dir, exist_ok=True)
            
            # Include filter info in filename if applicable
            filename_suffix = ""
            if role_filter and role_filter != 'All Roles':
                filename_suffix += f"_{role_filter}"
            if staff_filter and staff_filter != 'All Staff':
                if staff_filter == 'Selected Staff':
                    filename_suffix += "_CustomSelection"
                else:
                    filename_suffix += f"_{staff_filter.replace(' ', '_')}"
            
            temp_file = os.path.join(
                export_dir, 
                f'FY2026_Tracks_Export{filename_suffix}_{datetime.now(_eastern_tz).strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            wb.save(temp_file)
            
            return temp_file, "Export successful"
            
        except Exception as e:
            return None, f"Export failed: {str(e)}"


# Integration function to add to your app.py - ENHANCED VERSION
def add_fiscal_year_display_to_app():
    """
    ENHANCED: Add this function call to your app.py where you want the fiscal year display.
    This version includes both role and staff member filtering.
    """
    
    # Initialize the fiscal year display
    fy_display = FiscalYearDisplay()
    
    # Display the fiscal year section
    fy_display.display_fiscal_year_section()
    
    return fy_display


# Enhanced admin integration function
def add_fiscal_year_export_to_admin(admin_authenticated=False):
    """
    ENHANCED: Add this to your admin section in app.py with filtering support
    """
    if admin_authenticated:
        st.markdown("### 📊 Fiscal Year 2026 Export")
        
        fy_display = FiscalYearDisplay()
        
        # Load current data for export options
        tracks_data, staff_roles = fy_display.load_tracks_from_db()
        
        if tracks_data:
            # Export options
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Export Options:**")
                
                # Role filter for export
                available_roles = ['All Roles'] + sorted(list(set(staff_roles.values())))
                export_role_filter = st.selectbox(
                    "Filter by Role:",
                    options=available_roles,
                    key="admin_export_role_filter"
                )
                
                # Staff filter for export
                available_staff = sorted(list(tracks_data.keys()))
                staff_filter_options = ['All Staff'] + available_staff
                export_staff_filter = st.selectbox(
                    "Filter by Staff:",
                    options=staff_filter_options,
                    key="admin_export_staff_filter"
                )
                
                # Export button
                if st.button("📥 Export Filtered Data to Excel", use_container_width=True):
                    fy_display.export_and_download(
                        tracks_data, 
                        staff_roles,
                        export_role_filter if export_role_filter != 'All Roles' else None,
                        export_staff_filter if export_staff_filter != 'All Staff' else None
                    )
            
            with col2:
                st.markdown("**Current Database Statistics:**")
                
                role_stats = fy_display.get_role_statistics(staff_roles)
                total_staff = sum(stats['count'] for stats in role_stats.values())
                
                st.info(f"📊 **Total Staff:** {total_staff} members with active tracks")
                
                for role, stats in role_stats.items():
                    st.caption(f"• {role}: {stats['count']} ({stats['percentage']}%)")
                
                # Show filtered count if applicable
                if export_role_filter != 'All Roles' or export_staff_filter != 'All Staff':
                    filtered_staff = fy_display.filter_staff_by_criteria(
                        list(tracks_data.keys()), 
                        staff_roles, 
                        export_role_filter, 
                        export_staff_filter
                    )
                    st.success(f"🎯 **Filtered Result:** {len(filtered_staff)} staff member(s) will be exported")
        else:
            st.warning("No track data available for export")


# Additional utility functions for enhanced functionality
def get_staff_workload_summary(tracks_data, staff_roles, fiscal_year_display):
    """
    Utility function to analyze staff workload across the fiscal year
    """
    if not tracks_data:
        return {}
    
    workload_summary = {}
    fiscal_months = fiscal_year_display.get_fiscal_year_months()
    
    for staff_name, track_data in tracks_data.items():
        role = staff_roles.get(staff_name, 'Unknown')
        
        # Count shifts by type
        shift_counts = {'D': 0, 'N': 0, 'AT': 0, 'Off': 0}
        
        for pattern_day, shift in track_data.items():
            if shift in shift_counts:
                shift_counts[shift] += 1
            else:
                shift_counts['Off'] += 1
        
        total_shifts = shift_counts['D'] + shift_counts['N'] + shift_counts['AT']
        
        workload_summary[staff_name] = {
            'role': role,
            'day_shifts': shift_counts['D'],
            'night_shifts': shift_counts['N'],
            'admin_time': shift_counts['AT'],
            'off_days': shift_counts['Off'],
            'total_working_days': total_shifts,
            'workload_percentage': round((total_shifts / len(track_data)) * 100, 1) if track_data else 0
        }
    
    return workload_summary


def display_staff_workload_analysis(fy_display):
    """
    Display detailed workload analysis for staff members
    """
    st.markdown("### 📈 Staff Workload Analysis")
    
    tracks_data, staff_roles = fy_display.load_tracks_from_db()
    
    if not tracks_data:
        st.warning("No track data available for analysis.")
        return
    
    workload_summary = get_staff_workload_summary(tracks_data, staff_roles, fy_display)
    
    # Create DataFrame for analysis
    analysis_data = []
    for staff_name, data in workload_summary.items():
        analysis_data.append([
            staff_name,
            data['role'],
            data['day_shifts'],
            data['night_shifts'],
            data['admin_time'],
            data['total_working_days'],
            data['off_days'],
            f"{data['workload_percentage']}%"
        ])
    
    df = pd.DataFrame(analysis_data, columns=[
        'Staff Name', 'Role', 'Day Shifts', 'Night Shifts', 'Admin Time', 
        'Total Working Days', 'Off Days', 'Workload %'
    ])
    
    # Sort by workload percentage
    df_sorted = df.sort_values('Total Working Days', ascending=False)
    
    st.dataframe(df_sorted, use_container_width=True, hide_index=True)
    
    # Summary statistics
    col1, col2, col3 = st.columns(3)
    
    with col1:
        avg_workload = sum(data['workload_percentage'] for data in workload_summary.values()) / len(workload_summary)
        st.metric("Average Workload", f"{avg_workload:.1f}%")
    
    with col2:
        total_day_shifts = sum(data['day_shifts'] for data in workload_summary.values())
        st.metric("Total Day Shifts", total_day_shifts)
    
    with col3:
        total_night_shifts = sum(data['night_shifts'] for data in workload_summary.values())
        st.metric("Total Night Shifts", total_night_shifts)


# Quick start guide for integration
def display_integration_guide():
    """
    Display guide for integrating the enhanced fiscal year module
    """
    st.markdown("""
    ## 🚀 Integration Guide
    
    ### Enhanced Features Added:
    
    1. **Staff Member Filtering**: 
       - Filter by individual staff members
       - Multi-select interface for custom staff selections
       - Combined role and staff filtering
    
    2. **Improved User Interface**:
       - Organized staff selection by role
       - Quick select/deselect all options
       - Enhanced filter summary displays
    
    3. **Enhanced Export Options**:
       - Export filtered data only
       - Filename includes filter information
       - Admin panel with filtering support
       - Master Tracks tab (42-day pattern view)
       - Version tab with EST timestamp
    
    ### Integration Steps:
    
    1. Replace your existing `fiscal_year.py` with this enhanced version
    2. Ensure `pytz` is installed: `pip install pytz`
    3. Update your `app.py` to use the new functions:
       ```python
       from fiscal_year import add_fiscal_year_display_to_app
       
       # In your main app
       add_fiscal_year_display_to_app()
       ```
    
    4. For admin functionality:
       ```python
       from fiscal_year import add_fiscal_year_export_to_admin
       
       # In your admin section
       add_fiscal_year_export_to_admin(admin_authenticated=True)
       ```
    
    ### New Session State Variables:
    - `fy_staff_filter`: Current staff filter selection
    - `fy_selected_staff`: List of individually selected staff members
    
    ### Excel Export Tab Order:
    1. Master Tracks (42-day pattern)
    2. Overview
    3. Monthly sheets (September 2025 - September 2026)
    4. Version (creation timestamp in EST)
    
    All existing functionality is preserved while adding these new filtering capabilities!
    """)


# Example usage and testing function
def test_enhanced_filtering():
    """
    Test function to demonstrate the enhanced filtering capabilities
    """
    st.markdown("### 🧪 Testing Enhanced Filtering")
    
    # This would be used for testing the new functionality
    fy_display = FiscalYearDisplay()
    
    # Test data loading
    tracks_data, staff_roles = fy_display.load_tracks_from_db()
    
    if tracks_data:
        st.success(f"✅ Loaded data for {len(tracks_data)} staff members")
        
        # Test filtering
        test_filters = [
            ('All Roles', 'All Staff'),
            ('Nurse', 'All Staff'),
            ('All Roles', 'Selected Staff'),
        ]
        
        for role_filter, staff_filter in test_filters:
            filtered_staff = fy_display.filter_staff_by_criteria(
                list(tracks_data.keys()), staff_roles, role_filter, staff_filter
            )
            st.info(f"Filter: {role_filter} + {staff_filter} → {len(filtered_staff)} staff")
    else:
        st.warning("No test data available")


# Main execution block for standalone testing
if __name__ == "__main__":
    st.set_page_config(
        page_title="Enhanced Fiscal Year Display",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 Enhanced Fiscal Year 2026 Track Display")
    st.caption("With Role and Staff Member Filtering")
    
    # Display integration guide
    with st.expander("📖 Integration Guide", expanded=False):
        display_integration_guide()
    
    # Main functionality
    add_fiscal_year_display_to_app()
    
    # Optional workload analysis
    with st.expander("📈 Workload Analysis", expanded=False):
        fy_display = FiscalYearDisplay()
        display_staff_workload_analysis(fy_display)