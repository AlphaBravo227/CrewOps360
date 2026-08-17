# modules/ccemt_schedule.py
"""
CCEMT schedules in the database.

The CCEMT group doesn't bid a 42-day track — they work a fixed 28-day (4-week) pattern
that repeats indefinitely from a start date. That pattern lived on the CCEMT tab of
Tracks.xlsx: one row per staff member, 28 columns of shift codes (GR, NG, PG, LG, NL, MG,
NP, SIM, …).

It now lives in `ccemt_schedules` — one row per staff member per day of the pattern — so
it can be edited in the app. `ccemt_schedule_config` holds the pattern's start date, which
is what anchors a day of the pattern to a calendar date.

Shift codes are stored as authored, and classified for scheduling the same way the
spreadsheet reader did: a code starting with N is a night shift, anything else is a day
shift.
"""

from datetime import datetime

import pytz

from .day_pattern import WEEKDAYS

_eastern_tz = pytz.timezone('America/New_York')

# The pattern is four weeks long and repeats.
PATTERN_WEEKS = 4
PATTERN_LENGTH = PATTERN_WEEKS * 7  # 28 days

# Anchor date the pattern starts on — the first Sunday of the schedule. Matches the value
# TrainingTrackManager used while the pattern was read from the spreadsheet.
DEFAULT_START_DATE = '2025-09-14'

_START_DATE_KEY = 'pattern_start_date'


def _get_conn():
    from .db_utils import get_db_connection
    return get_db_connection()


def _now():
    return datetime.now(_eastern_tz).strftime('%Y-%m-%d %H:%M:%S')


def initialize_ccemt_tables():
    """Create the CCEMT schedule tables if they don't exist. Safe to call repeatedly."""
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS ccemt_schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_name TEXT NOT NULL,
            day_index INTEGER NOT NULL,
            shift_code TEXT NOT NULL,
            created_date TEXT NOT NULL,
            modified_date TEXT NOT NULL,
            UNIQUE(staff_name, day_index)
        )
        ''')
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS ccemt_schedule_config (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            modified_date TEXT NOT NULL
        )
        ''')
        cursor.execute('''CREATE INDEX IF NOT EXISTS idx_ccemt_staff
                          ON ccemt_schedules(staff_name)''')
        conn.commit()
        return True
    except Exception as e:
        print(f"Error initializing CCEMT schedule tables: {e}")
        return False


def _table_exists():
    try:
        cursor = _get_conn().cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' "
                       "AND name='ccemt_schedules'")
        return cursor.fetchone() is not None
    except Exception:
        return False


def day_labels():
    """
    Column labels for the 28-day pattern: "Wk1 Sun" … "Wk4 Sat".

    The spreadsheet's header row was just Sun..Sat repeated four times, which is
    ambiguous once you're looking at a grid on screen, so the week is spelled out.
    """
    return [f"Wk{week + 1} {weekday}"
            for week in range(PATTERN_WEEKS)
            for weekday in WEEKDAYS]


def classify_shift_code(shift_code):
    """
    Scheduling class of a raw CCEMT code.

    Returns:
        str: 'N' for night (any code starting with N), 'D' otherwise, '' when blank.
    """
    code = str(shift_code or '').strip().upper()
    if not code:
        return ''
    return 'N' if code.startswith('N') else 'D'


def get_pattern_start_date():
    """
    The date pattern day 0 falls on.

    Returns:
        datetime: parsed start date, falling back to DEFAULT_START_DATE.
    """
    value = DEFAULT_START_DATE
    if _table_exists():
        try:
            cursor = _get_conn().cursor()
            cursor.execute("SELECT value FROM ccemt_schedule_config WHERE key = ?",
                           (_START_DATE_KEY,))
            row = cursor.fetchone()
            if row and row[0]:
                value = row[0]
        except Exception as e:
            print(f"Error reading the CCEMT pattern start date: {e}")
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d')
    except ValueError:
        return datetime.strptime(DEFAULT_START_DATE, '%Y-%m-%d')


def set_pattern_start_date(start_date, changed_by=None):
    """
    Set the date pattern day 0 falls on.

    Args:
        start_date: a date/datetime, or a YYYY-MM-DD string.

    Returns:
        tuple: (success, message)
    """
    if isinstance(start_date, str):
        try:
            parsed = datetime.strptime(start_date[:10], '%Y-%m-%d')
        except ValueError:
            return False, f"'{start_date}' is not a date in YYYY-MM-DD form."
    else:
        try:
            parsed = datetime(start_date.year, start_date.month, start_date.day)
        except AttributeError:
            return False, "A date is required."

    # The pattern's first column is a Sunday; anchoring it to another weekday would shift
    # every staff member's schedule by that many days.
    if parsed.weekday() != 6:
        return False, ("The pattern starts on a Sunday — "
                       f"{parsed.strftime('%Y-%m-%d')} is a {parsed.strftime('%A')}.")

    initialize_ccemt_tables()
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO ccemt_schedule_config (key, value, modified_date)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value,
                                           modified_date = excluded.modified_date
        """, (_START_DATE_KEY, parsed.strftime('%Y-%m-%d'), _now()))
        conn.commit()
        return True, f"CCEMT pattern now starts on {parsed.strftime('%Y-%m-%d')}."
    except Exception as e:
        return False, f"Error saving the CCEMT pattern start date: {e}"


def get_schedules():
    """
    Every CCEMT staff member's pattern.

    Returns:
        dict: {staff_name: {day_index: raw shift code}}
    """
    result = {}
    if not _table_exists():
        return result
    try:
        cursor = _get_conn().cursor()
        cursor.execute("SELECT staff_name, day_index, shift_code FROM ccemt_schedules")
        for staff_name, day_index, shift_code in cursor.fetchall():
            name = str(staff_name).strip()
            if not name:
                continue
            code = str(shift_code or '').strip().upper()
            if not code:
                continue
            result.setdefault(name, {})[int(day_index)] = code
    except Exception as e:
        print(f"Error reading CCEMT schedules: {e}")
    return result


def get_classified_schedules():
    """
    The same patterns reduced to 'D'/'N'.

    Returns:
        dict: {staff_name: {day_index: 'D' | 'N'}}
    """
    return {name: {index: classify_shift_code(code) for index, code in days.items()
                   if classify_shift_code(code)}
            for name, days in get_schedules().items()}


def get_staff_schedule(staff_name):
    """
    One staff member's pattern.

    Returns:
        dict: {day_index: raw shift code}
    """
    name = str(staff_name or '').strip()
    if not name:
        return {}
    schedules = get_schedules()
    if name in schedules:
        return schedules[name]
    lowered = name.lower()
    for key, value in schedules.items():
        if key.lower() == lowered:
            return value
    return {}


def get_ccemt_staff_names():
    """Names with a CCEMT pattern on file."""
    return sorted(get_schedules().keys())


def set_staff_schedule(staff_name, pattern, changed_by=None):
    """
    Replace one staff member's 28-day pattern.

    Args:
        pattern (dict): {day_index (0-27): shift code}. Blank codes are removed, so
            clearing a cell in the editor clears that day.

    Returns:
        tuple: (success, message)
    """
    name = str(staff_name or '').strip()
    if not name:
        return False, "Staff name is required."

    cleaned = {}
    for day_index, shift_code in (pattern or {}).items():
        try:
            index = int(day_index)
        except (TypeError, ValueError):
            return False, f"'{day_index}' is not a day of the pattern."
        if not 0 <= index < PATTERN_LENGTH:
            return False, f"Day {index} is outside the {PATTERN_LENGTH}-day pattern."
        code = str(shift_code or '').strip().upper()
        if code:
            cleaned[index] = code

    initialize_ccemt_tables()
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        now = _now()
        cursor.execute("DELETE FROM ccemt_schedules WHERE staff_name = ? COLLATE NOCASE",
                       (name,))
        for index, code in sorted(cleaned.items()):
            cursor.execute("""
                INSERT INTO ccemt_schedules
                (staff_name, day_index, shift_code, created_date, modified_date)
                VALUES (?, ?, ?, ?, ?)
            """, (name, index, code, now, now))
        conn.commit()
        if cleaned:
            return True, f"Saved {len(cleaned)} scheduled day(s) for {name}."
        return True, f"Cleared the CCEMT schedule for {name}."
    except Exception as e:
        return False, f"Error saving the CCEMT schedule for {name}: {e}"


def delete_staff_schedule(staff_name):
    """
    Remove a staff member's CCEMT pattern.

    Returns:
        tuple: (success, message)
    """
    name = str(staff_name or '').strip()
    if not name:
        return False, "Staff name is required."
    if not _table_exists():
        return False, "No CCEMT schedules on file."
    try:
        conn = _get_conn()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM ccemt_schedules WHERE staff_name = ? COLLATE NOCASE",
                       (name,))
        deleted = cursor.rowcount
        conn.commit()
        return True, f"Removed {name} from the CCEMT schedule ({deleted} day(s))."
    except Exception as e:
        return False, f"Error removing {name} from the CCEMT schedule: {e}"


def import_from_excel(path='upload files/Tracks.xlsx', sheet_name='CCEMT',
                      overwrite=False, dry_run=False):
    """
    One-time import of the CCEMT tab.

    Layout: row 1 is a header (column A blank, columns B..AC the weekday names), column A
    from row 2 down holds staff names, and columns B..AC hold the 28-day pattern.

    Returns:
        dict: report with keys staff_imported, days_imported, skipped, errors, dry_run.
    """
    import os

    report = {
        'staff_imported': 0,
        'days_imported': 0,
        'skipped': [],
        'errors': [],
        'dry_run': bool(dry_run),
    }

    if not os.path.exists(path):
        report['errors'].append(f"Tracks workbook not found: {path}")
        return report

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        report['errors'].append(f"Could not open {path}: {e}")
        return report

    if sheet_name not in workbook.sheetnames:
        report['errors'].append(f"No '{sheet_name}' tab in {path}.")
        return report

    sheet = workbook[sheet_name]
    rows = {}
    for row in sheet.iter_rows(min_row=2):
        staff_name = row[0].value if row else None
        if not staff_name:
            continue
        name = str(staff_name).strip()
        if not name:
            continue

        pattern = {}
        for day_index in range(PATTERN_LENGTH):
            cell_index = day_index + 1  # column B is index 1
            if cell_index >= len(row):
                break
            value = row[cell_index].value
            if value is None:
                continue
            code = str(value).strip().upper()
            if code:
                pattern[day_index] = code

        if pattern:
            rows[name] = pattern
        else:
            report['skipped'].append(name)

    report['staff_imported'] = len(rows)
    report['days_imported'] = sum(len(pattern) for pattern in rows.values())

    if dry_run:
        return report

    initialize_ccemt_tables()
    existing = get_schedules()
    if existing and not overwrite:
        report['errors'].append(
            f"CCEMT schedules already exist for {len(existing)} staff member(s). "
            "Choose to overwrite if you want to replace them.")
        report['staff_imported'] = 0
        report['days_imported'] = 0
        return report

    try:
        conn = _get_conn()
        cursor = conn.cursor()
        now = _now()
        if existing:
            cursor.execute("DELETE FROM ccemt_schedules")
        for name, pattern in rows.items():
            for day_index, code in sorted(pattern.items()):
                cursor.execute("""
                    INSERT INTO ccemt_schedules
                    (staff_name, day_index, shift_code, created_date, modified_date)
                    VALUES (?, ?, ?, ?, ?)
                """, (name, day_index, code, now, now))
        conn.commit()
    except Exception as e:
        report['errors'].append(f"Error saving imported CCEMT schedules: {e}")

    return report
