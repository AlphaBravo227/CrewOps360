# modules/track_display.py - Enhanced with Comprehensive Staff Member filtering
"""
Enhanced track display module with fullscreen capability and comprehensive filtering
Uses database effective_role information instead of Excel files

ENHANCED CHANGES:
1. Role selection with three buttons (maintained)
2. Full schedule display (all 42 days) in proper order (maintained)
3. ADDED: Multi-select staff interface similar to fiscal year module
4. ADDED: Staff selection organized by role
5. ADDED: Quick select/deselect all functionality
6. ADDED: Enhanced filter summary and statistics
7. ADDED: Persistent staff selections across views
8. IMPROVED: Better UI consistency with fiscal year module
"""

import streamlit as st
import pandas as pd
from .db_utils import get_all_active_tracks

def display_track_viewer():
    """
    Enhanced track viewer with fullscreen option for better viewing of active tracks
    """
    # Check if we should show fullscreen mode
    if st.session_state.get('track_viewer_fullscreen', False):
        display_fullscreen_track_viewer()
        return
    
    # Original compact view in right column
    display_compact_track_viewer()

def display_compact_track_viewer():
    """
    Compact track viewer for the split-screen layout
    ENHANCED: Now includes comprehensive staff filtering with multi-select capability
    """
    if st.session_state.get('current_tracks_df') is None:
        st.warning("Track data not available")
        return
    
    # Get all tracks for filter options
    try:
        success, all_tracks = get_all_active_tracks()
        if not success or not all_tracks:
            st.warning("No track data available")
            return
    except Exception as e:
        st.error(f"Error loading track data: {str(e)}")
        return
    
    # Initialize session state for filters if not exists
    if 'compact_role_select' not in st.session_state:
        st.session_state.compact_role_select = 'All'
    if 'compact_staff_filter' not in st.session_state:
        st.session_state.compact_staff_filter = 'All Staff'
    if 'compact_selected_staff' not in st.session_state:
        st.session_state.compact_selected_staff = []
    
    # Filter Options Header (matching fiscal year module)
    st.markdown("### 🔍 Filter Options")
    
    # Available options for filters
    available_roles = ['All Roles', 'Nurse', 'Medic']
    current_role = st.session_state.get('compact_role_select', 'All')
    
    # Map display names to internal values
    role_mapping = {'All Roles': 'All', 'Nurse': 'nurse', 'Medic': 'medic'}
    reverse_role_mapping = {'All': 'All Roles', 'nurse': 'Nurse', 'medic': 'Medic'}
    
    # Filter interface matching fiscal year layout
    filter_col1, filter_col2 = st.columns(2)
    
    with filter_col1:
        st.markdown("**Filter by Role:**")
        role_display = reverse_role_mapping.get(current_role, 'All Roles')
        
        role_filter = st.selectbox(
            "",
            options=available_roles,
            index=available_roles.index(role_display) if role_display in available_roles else 0,
            key="compact_role_select_box",
            label_visibility="collapsed"
        )
        
        # Update session state and reset staff filter when role changes
        new_role_value = role_mapping[role_filter]
        if new_role_value != st.session_state.get('compact_role_select'):
            st.session_state.compact_role_select = new_role_value
            st.session_state.compact_staff_filter = 'All Staff'
            st.session_state.compact_selected_staff = []
    
    with filter_col2:
        st.markdown("**Filter by Staff:**")
        # Get staff options based on current role filter (using mapped value)
        available_staff = get_available_staff_by_role(all_tracks, new_role_value)
        staff_filter_options = ['All Staff', 'Selected Staff'] + available_staff
        
        current_staff_filter = st.session_state.get('compact_staff_filter', 'All Staff')
        
        # Ensure current selection is still valid
        if current_staff_filter not in staff_filter_options:
            current_staff_filter = 'All Staff'
            st.session_state.compact_staff_filter = 'All Staff'
        
        # Staff filter selectbox
        selected_staff_filter = st.selectbox(
            "",
            options=staff_filter_options,
            index=staff_filter_options.index(current_staff_filter) if current_staff_filter in staff_filter_options else 0,
            key="compact_staff_filter_select",
            label_visibility="collapsed"
        )
        
        # Update session state if selection changed
        if selected_staff_filter != st.session_state.get('compact_staff_filter'):
            st.session_state.compact_staff_filter = selected_staff_filter
    
    # Show staff selection interface if "Selected Staff" is chosen
    if selected_staff_filter == 'Selected Staff':
        with st.expander("👥 Select Staff Members", expanded=True):
            display_compact_staff_selection_interface(available_staff, all_tracks)
    
    # Filter summary
    filtered_staff_count = get_filtered_staff_count(all_tracks, new_role_value, selected_staff_filter)
    if filtered_staff_count > 0:
        filter_summary = f"Showing {filtered_staff_count} staff member(s)"
        if new_role_value != 'All':
            filter_summary += f" • Role: {role_filter}"
        if selected_staff_filter not in ['All Staff']:
            if selected_staff_filter == 'Selected Staff':
                selected_count = len(st.session_state.get('compact_selected_staff', []))
                filter_summary += f" • Custom Selection ({selected_count})"
            else:
                filter_summary += f" • {selected_staff_filter}"
        st.info(filter_summary)
    else:
        st.warning("No staff members match the selected filters.")
        return
    
    # Fullscreen toggle button
    if st.button("⛶ Fullscreen", key="track_fullscreen", type="primary", use_container_width=True):
        st.session_state.track_viewer_fullscreen = True
        st.session_state.selected_role_fullscreen = new_role_value
        st.session_state.selected_staff_filter_fullscreen = selected_staff_filter
        st.rerun()
    
    # Display compact track table with enhanced filtering
    display_role_tracks_compact(new_role_value, selected_staff_filter)

def display_fullscreen_track_viewer():
    """
    Fullscreen track viewer with enhanced layout and comprehensive filtering
    ENHANCED: Includes full multi-select staff interface in fullscreen mode
    """
    # Header with back button
    col1, col2 = st.columns([1, 6])
    
    with col1:
        if st.button("← Back", use_container_width=True):
            st.session_state.track_viewer_fullscreen = False
            st.session_state.pop('selected_role_fullscreen', None)
            st.session_state.pop('selected_staff_filter_fullscreen', None)
            st.rerun()
    
    with col2:
        st.markdown("## 📊 Active Tracks - Fullscreen View")
    
    # Get all tracks for filter options
    try:
        success, all_tracks = get_all_active_tracks()
        if not success or not all_tracks:
            st.warning("No track data available")
            return
    except Exception as e:
        st.error(f"Error loading track data: {str(e)}")
        return
    
    # Get the previously selected filters or defaults
    selected_role = st.session_state.get('selected_role_fullscreen', st.session_state.get('compact_role_select', 'All'))
    selected_staff_filter = st.session_state.get('selected_staff_filter_fullscreen', st.session_state.get('compact_staff_filter', 'All Staff'))
    
    # Available options for filters (matching compact view)
    available_roles = ['All Roles', 'Nurse', 'Medic']
    role_mapping = {'All Roles': 'All', 'Nurse': 'nurse', 'Medic': 'medic'}
    reverse_role_mapping = {'All': 'All Roles', 'nurse': 'Nurse', 'medic': 'Medic'}
    
    # Filter interface matching fiscal year layout
    with st.expander("🔍 Filter Options", expanded=False):
        filter_col1, filter_col2 = st.columns(2)
        
        with filter_col1:
            role_display = reverse_role_mapping.get(selected_role, 'All Roles')
            role_filter = st.selectbox(
                "Filter by Role:",
                options=available_roles,
                index=available_roles.index(role_display) if role_display in available_roles else 0,
                key="fs_role_select_dropdown"
            )
            new_role_value = role_mapping[role_filter]
            st.session_state.selected_role_fullscreen = new_role_value
        
        with filter_col2:
            # Get available staff based on selected role
            available_staff = get_available_staff_by_role(all_tracks, new_role_value)
            staff_filter_options = ['All Staff', 'Selected Staff'] + available_staff
            
            # Ensure current selection is still valid
            if selected_staff_filter not in staff_filter_options:
                selected_staff_filter = 'All Staff'
                st.session_state.selected_staff_filter_fullscreen = 'All Staff'
            
            staff_filter = st.selectbox(
                "Filter by Staff:",
                options=staff_filter_options,
                index=staff_filter_options.index(selected_staff_filter) if selected_staff_filter in staff_filter_options else 0,
                key="fs_staff_filter_select_dropdown"
            )
            st.session_state.selected_staff_filter_fullscreen = staff_filter
        
        # Show staff selection interface if needed
        if staff_filter == 'Selected Staff':
            display_fullscreen_staff_selection_interface(available_staff, all_tracks, new_role_value)
    
    # Display current filter status
    display_filter_status(new_role_value, staff_filter, all_tracks)
    
    # Display fullscreen track table
    display_role_tracks_fullscreen(new_role_value, staff_filter)

def display_compact_staff_selection_interface(available_staff, all_tracks):
    """
    Compact staff selection interface for the right column
    """
    if not available_staff:
        st.info("No staff available for selection")
        return
    
    # Group staff by role for organized display
    staff_by_role = get_staff_grouped_by_role(available_staff, all_tracks)
    
    # Quick selection buttons
    sel_col1, sel_col2, sel_col3 = st.columns(3)
    with sel_col1:
        if st.button("Select All", key="compact_select_all_staff", use_container_width=True):
            st.session_state.compact_selected_staff = available_staff.copy()
            st.rerun()
    with sel_col2:
        if st.button("Clear All", key="compact_clear_all_staff", use_container_width=True):
            st.session_state.compact_selected_staff = []
            st.rerun()
    with sel_col3:
        selected_count = len(st.session_state.get('compact_selected_staff', []))
        st.caption(f"Selected: {selected_count}")
    
    # Compact checkboxes (2 columns)
    for role, staff_list in sorted(staff_by_role.items()):
        st.markdown(f"**{role.title()}s ({len(staff_list)}):**")
        
        # Display in 2 columns for compact view
        for i in range(0, len(staff_list), 2):
            check_col1, check_col2 = st.columns(2)
            
            # First column
            if i < len(staff_list):
                staff = staff_list[i]
                with check_col1:
                    is_selected = staff in st.session_state.get('compact_selected_staff', [])
                    if st.checkbox(staff, value=is_selected, key=f"compact_staff_check_{staff}_{i}"):
                        if staff not in st.session_state.compact_selected_staff:
                            st.session_state.compact_selected_staff.append(staff)
                    else:
                        if staff in st.session_state.compact_selected_staff:
                            st.session_state.compact_selected_staff.remove(staff)
            
            # Second column
            if i + 1 < len(staff_list):
                staff = staff_list[i + 1]
                with check_col2:
                    is_selected = staff in st.session_state.get('compact_selected_staff', [])
                    if st.checkbox(staff, value=is_selected, key=f"compact_staff_check_{staff}_{i+1}"):
                        if staff not in st.session_state.compact_selected_staff:
                            st.session_state.compact_selected_staff.append(staff)
                    else:
                        if staff in st.session_state.compact_selected_staff:
                            st.session_state.compact_selected_staff.remove(staff)

def display_fullscreen_staff_selection_interface(available_staff, all_tracks, selected_role):
    """
    Full staff selection interface for fullscreen mode - similar to fiscal year module
    """
    if not available_staff:
        st.info("No staff available for selection")
        return
    
    # Group staff by role for better organization
    staff_by_role = get_staff_grouped_by_role(available_staff, all_tracks)
    
    # Quick selection buttons
    quick_col1, quick_col2, quick_col3, quick_col4 = st.columns(4)
    with quick_col1:
        if st.button("Select All", key="fs_select_all_staff", use_container_width=True):
            st.session_state.compact_selected_staff = available_staff.copy()
            st.rerun()
    with quick_col2:
        if st.button("Clear All", key="fs_clear_all_staff", use_container_width=True):
            st.session_state.compact_selected_staff = []
            st.rerun()
    with quick_col3:
        selected_count = len(st.session_state.get('compact_selected_staff', []))
        total_available = len(available_staff)
        st.metric("Selected", f"{selected_count}/{total_available}")
    with quick_col4:
        if selected_count > 0:
            percentage = round((selected_count / total_available) * 100, 1)
            st.metric("Percentage", f"{percentage}%")
    
    # Role-organized selection
    for role in sorted(staff_by_role.keys()):
        role_staff = staff_by_role[role]
        with st.expander(f"👥 {role.title()}s ({len(role_staff)})", expanded=True):
            
            # Role-specific quick actions
            role_col1, role_col2 = st.columns(2)
            with role_col1:
                if st.button(f"Select All {role.title()}s", key=f"fs_select_role_{role}"):
                    for staff in role_staff:
                        if staff not in st.session_state.compact_selected_staff:
                            st.session_state.compact_selected_staff.append(staff)
                    st.rerun()
            with role_col2:
                if st.button(f"Clear {role.title()}s", key=f"fs_clear_role_{role}"):
                    st.session_state.compact_selected_staff = [
                        s for s in st.session_state.compact_selected_staff if s not in role_staff
                    ]
                    st.rerun()
            
            # Individual checkboxes in 3 columns for fullscreen
            for i in range(0, len(role_staff), 3):
                check_cols = st.columns(3)
                for j, col in enumerate(check_cols):
                    if i + j < len(role_staff):
                        staff = role_staff[i + j]
                        with col:
                            is_selected = staff in st.session_state.get('compact_selected_staff', [])
                            if st.checkbox(
                                staff, 
                                value=is_selected, 
                                key=f"fs_staff_check_{staff}_{role}_{i+j}"
                            ):
                                if staff not in st.session_state.compact_selected_staff:
                                    st.session_state.compact_selected_staff.append(staff)
                            else:
                                if staff in st.session_state.compact_selected_staff:
                                    st.session_state.compact_selected_staff.remove(staff)

def get_available_staff_by_role(all_tracks, role_filter):
    """
    Get list of available staff members based on role filter
    
    Args:
        all_tracks (list): All track data
        role_filter (str): Current role filter ('All', 'nurse', 'medic')
        
    Returns:
        list: List of staff names (without 'All' option)
    """
    staff_names = set()
    
    for track in all_tracks:
        metadata = track.get('metadata', {})
        effective_role = metadata.get('effective_role', 'nurse')
        
        # Filter by role if not 'All'
        if role_filter == 'All' or effective_role == role_filter:
            staff_names.add(track['staff_name'])
    
    return sorted(list(staff_names))

def get_staff_grouped_by_role(staff_list, all_tracks):
    """
    Group staff members by their roles
    
    Args:
        staff_list (list): List of staff names
        all_tracks (list): All track data
        
    Returns:
        dict: Staff grouped by role
    """
    staff_by_role = {}
    
    # Create lookup for staff roles
    staff_roles = {}
    for track in all_tracks:
        metadata = track.get('metadata', {})
        effective_role = metadata.get('effective_role', 'nurse')
        staff_roles[track['staff_name']] = effective_role
    
    # Group staff by role
    for staff in staff_list:
        role = staff_roles.get(staff, 'unknown')
        if role not in staff_by_role:
            staff_by_role[role] = []
        staff_by_role[role].append(staff)
    
    # Sort staff within each role
    for role in staff_by_role:
        staff_by_role[role].sort()
    
    return staff_by_role

def get_role_statistics(all_tracks):
    """
    Get statistics about staff roles
    
    Returns:
        dict: Role counts
    """
    role_counts = {}
    
    for track in all_tracks:
        metadata = track.get('metadata', {})
        effective_role = metadata.get('effective_role', 'nurse')
        role_counts[effective_role] = role_counts.get(effective_role, 0) + 1
    
    return role_counts

def get_filtered_staff_count(all_tracks, role_filter, staff_filter):
    """
    Get count of staff members matching current filters
    """
    tracks = get_tracks_from_database_by_filters(role_filter, staff_filter)
    return len(tracks)

def display_filter_status(selected_role, staff_filter, all_tracks):
    """
    Display current filter status with statistics
    """
    # Get filtered count
    filtered_count = get_filtered_staff_count(all_tracks, selected_role, staff_filter)
    total_count = len(all_tracks)
    
    # Create status message
    status_parts = []
    if selected_role != 'All':
        status_parts.append(f"Role: {selected_role.title()}")
    
    if staff_filter not in ['All Staff']:
        if staff_filter == 'Selected Staff':
            selected_count = len(st.session_state.get('compact_selected_staff', []))
            status_parts.append(f"Custom Selection ({selected_count} staff)")
        else:
            status_parts.append(f"Staff: {staff_filter}")
    
    if status_parts:
        status_text = " • ".join(status_parts)
        st.info(f"📊 **Current Filters:** {status_text} • **Results:** {filtered_count}/{total_count} staff members")
    else:
        st.info(f"📊 **Showing all staff:** {filtered_count} total members")

def get_ordered_day_columns():
    """
    Return day columns in the proper order as specified in data
    Order: Sun A 1, Mon A 1, Tue A 1, ..., Sat A 1, then Sun A 2, Mon A 2, etc.
    """
    day_order = [
        "Sun A 1", "Mon A 1", "Tue A 1", "Wed A 1", "Thu A 1", "Fri A 1", "Sat A 1",
        "Sun A 2", "Mon A 2", "Tue A 2", "Wed A 2", "Thu A 2", "Fri A 2", "Sat A 2",
        "Sun B 3", "Mon B 3", "Tue B 3", "Wed B 3", "Thu B 3", "Fri B 3", "Sat B 3",
        "Sun B 4", "Mon B 4", "Tue B 4", "Wed B 4", "Thu B 4", "Fri B 4", "Sat B 4",
        "Sun C 5", "Mon C 5", "Tue C 5", "Wed C 5", "Thu C 5", "Fri C 5", "Sat C 5",
        "Sun C 6", "Mon C 6", "Tue C 6", "Wed C 6", "Thu C 6", "Fri C 6", "Sat C 6"
    ]
    return day_order

def display_role_tracks_compact(selected_role, selected_staff_filter='All Staff'):
    """
    Display tracks in compact mode with enhanced filtering
    """
    try:
        # Get tracks from database with enhanced filtering
        tracks = get_tracks_from_database_by_filters(selected_role, selected_staff_filter)
        
        if not tracks:
            st.info("No tracks match the selected filters.")
            return
        
        # Get ordered day columns
        ordered_days = get_ordered_day_columns()
        
        # Convert to DataFrame for display
        display_data = []
        for track in tracks:
            staff_name = track['staff_name']
            track_data = track['track_data']
            metadata = track.get('metadata', {})
            effective_role = metadata.get('effective_role', 'nurse')
            
            # Create row with all days in proper order
            row = {'Staff Name': staff_name, 'Role': effective_role.title()}
            
            # Add all 42 days in the specified order
            for day in ordered_days:
                row[day] = track_data.get(day, '')
            
            display_data.append(row)
        
        if display_data:
            df = pd.DataFrame(display_data)
            
            # Apply styling for better visualization
            def style_shifts(val):
                if val == 'D':
                    return 'background-color: #dbeafe; color: #1e40af; font-weight: bold;'
                elif val == 'N':
                    return 'background-color: #ede9fe; color: #6b21a8; font-weight: bold;'
                elif val == 'AT':
                    return 'background-color: #d1fae5; color: #065f46; font-weight: bold;'
                elif val == '':
                    return 'background-color: #f9fafb; color: #6b7280;'
                return ''
            
            # Apply styling to shift columns only
            shift_columns = [col for col in df.columns if col not in ['Staff Name', 'Role']]
            styled_df = df.style.applymap(style_shifts, subset=shift_columns)
            
            st.dataframe(styled_df, use_container_width=True, hide_index=True, height=400)
            
            # Show legend
            st.caption("🟦 **D** = Day Shift | 🟪 **N** = Night Shift | 🟢 **AT** = Admin Time | ⚪ = Off")
        
        st.caption("Showing full 6-week schedule (42 days). Use fullscreen for better viewing of all columns.")
        
    except Exception as e:
        st.error(f"Error displaying tracks: {str(e)}")

def display_role_tracks_fullscreen(selected_role, selected_staff_filter='All Staff'):
    """
    Display tracks in fullscreen mode with enhanced filtering
    """
    try:
        # Get tracks from database with enhanced filtering
        tracks = get_tracks_from_database_by_filters(selected_role, selected_staff_filter)
        
        if not tracks:
            st.info("No tracks match the selected filters.")
            return
        
        # Get ordered day columns
        ordered_days = get_ordered_day_columns()
        
        # Convert to DataFrame for display
        display_data = []
        for track in tracks:
            staff_name = track['staff_name']
            track_data = track['track_data']
            metadata = track.get('metadata', {})
            effective_role = metadata.get('effective_role', 'nurse')
            
            # Create row with all days in proper order
            row = {'Staff Name': staff_name, 'Role': effective_role.title()}
            
            # Add all 42 days in the specified order
            for day in ordered_days:
                row[day] = track_data.get(day, '')
            
            display_data.append(row)
        
        if display_data:
            df = pd.DataFrame(display_data)
            
            # Apply styling for better visualization
            def style_shifts(val):
                if val == 'D':
                    return 'background-color: #dbeafe; color: #1e40af; font-weight: bold;'
                elif val == 'N':
                    return 'background-color: #ede9fe; color: #6b21a8; font-weight: bold;'
                elif val == 'AT':
                    return 'background-color: #d1fae5; color: #065f46; font-weight: bold;'
                elif val == '':
                    return 'background-color: #f9fafb; color: #6b7280;'
                return ''
            
            # Apply styling to shift columns only
            shift_columns = [col for col in df.columns if col not in ['Staff Name', 'Role']]
            styled_df = df.style.applymap(style_shifts, subset=shift_columns)
            
            st.dataframe(
                styled_df, 
                use_container_width=True,
                hide_index=True,
                height=600
            )
            
            # Enhanced legend and statistics
            legend_col1, legend_col2 = st.columns(2)
            with legend_col1:
                st.caption("🟦 **D** = Day Shift | 🟪 **N** = Night Shift | 🟢 **AT** = Admin Time | ⚪ = Off")
            with legend_col2:
                # Calculate shift statistics
                shift_stats = calculate_shift_statistics(tracks)
                if shift_stats:
                    stats_text = " | ".join([f"{shift}: {count}" for shift, count in shift_stats.items()])
                    st.caption(f"📊 **Shift Distribution:** {stats_text}")
        
        st.caption("Full 6-week schedule displayed in chronological order: Sun A 1 → Mon A 1 → ... → Sat C 6")
        
    except Exception as e:
        st.error(f"Error displaying tracks: {str(e)}")

def calculate_shift_statistics(tracks):
    """
    Calculate statistics about shift distribution
    """
    shift_counts = {}
    
    for track in tracks:
        track_data = track.get('track_data', {})
        for day, shift in track_data.items():
            if shift:
                shift_counts[shift] = shift_counts.get(shift, 0) + 1
    
    return shift_counts

def get_tracks_from_database_by_filters(selected_role, selected_staff_filter='All Staff'):
    """
    Get tracks from database filtered by role and comprehensive staff filtering
    
    Args:
        selected_role (str): Role to filter by ('All', 'nurse', 'medic')
        selected_staff_filter (str): Staff filter type ('All Staff', 'Selected Staff', or individual name)
        
    Returns:
        list: List of track data with comprehensive filtering applied
    """
    try:
        success, tracks_data = get_all_active_tracks()
        if not success:
            return []
        
        filtered_tracks = []
        
        for track in tracks_data:
            metadata = track.get('metadata', {})
            effective_role = metadata.get('effective_role', 'nurse')
            staff_name = track.get('staff_name', '')
            
            # Filter by role
            role_match = (selected_role == 'All' or effective_role == selected_role)
            
            # Filter by staff member using comprehensive filtering
            staff_match = False
            
            if selected_staff_filter == 'All Staff':
                staff_match = True
            elif selected_staff_filter == 'Selected Staff':
                # Check if staff member is in the selected list
                selected_staff_list = st.session_state.get('compact_selected_staff', [])
                staff_match = staff_name in selected_staff_list
            else:
                # Individual staff member selected
                staff_match = (staff_name == selected_staff_filter)
            
            if role_match and staff_match:
                filtered_tracks.append(track)
        
        # Sort tracks by role then staff name for consistent display
        filtered_tracks.sort(key=lambda x: (
            x.get('metadata', {}).get('effective_role', 'nurse'),
            x.get('staff_name', '')
        ))
        
        return filtered_tracks
    
    except Exception as e:
        st.error(f"Error retrieving tracks from database: {str(e)}")
        return []

# Legacy function for backward compatibility
def get_tracks_from_database_by_role(selected_role):
    """
    Legacy function - now calls the new comprehensive filtering function
    """
    return get_tracks_from_database_by_filters(selected_role, 'All Staff')

def get_staff_options(all_tracks, role_filter):
    """
    Legacy function for backward compatibility - now uses enhanced staff retrieval
    """
    staff_list = get_available_staff_by_role(all_tracks, role_filter)
    return ['All'] + staff_list

# Additional utility functions for enhanced functionality

def export_filtered_tracks_to_excel(selected_role='All', selected_staff_filter='All Staff'):
    """
    Export filtered tracks to Excel format
    Enhanced version with comprehensive filtering support
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import tempfile
        import os
        
        # Get filtered tracks
        tracks = get_tracks_from_database_by_filters(selected_role, selected_staff_filter)
        
        if not tracks:
            return None, "No tracks match the selected filters for export."
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Active_Tracks"
        
        # Headers
        ws['A1'] = 'Staff Name'
        ws['B1'] = 'Role'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        # Day columns
        ordered_days = get_ordered_day_columns()
        for idx, day in enumerate(ordered_days, start=3):
            cell = ws.cell(row=1, column=idx, value=day)
            cell.font = Font(bold=True)
        
        # Define shift colors
        shift_fills = {
            'D': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            'N': PatternFill(start_color="E6CCFF", end_color="E6CCFF", fill_type="solid"),
            'AT': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        }
        
        # Add track data
        for row_idx, track in enumerate(tracks, start=2):
            staff_name = track['staff_name']
            metadata = track.get('metadata', {})
            effective_role = metadata.get('effective_role', 'nurse')
            track_data = track.get('track_data', {})
            
            # Staff name and role
            ws.cell(row=row_idx, column=1, value=staff_name)
            ws.cell(row=row_idx, column=2, value=effective_role.title())
            
            # Shift data
            for col_idx, day in enumerate(ordered_days, start=3):
                shift = track_data.get(day, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=shift)
                
                # Apply styling
                if shift in shift_fills:
                    cell.fill = shift_fills[shift]
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        for col_idx in range(3, len(ordered_days) + 3):
            ws.column_dimensions[chr(64 + col_idx)].width = 8
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name, f"Successfully exported {len(tracks)} tracks to Excel."
    
    except Exception as e:
        return None, f"Export failed: {str(e)}"

def display_track_summary_statistics():
    """
    Display summary statistics for the current track data
    """
    try:
        success, all_tracks = get_all_active_tracks()
        if not success or not all_tracks:
            st.warning("No track data available for statistics.")
            return
        
        st.markdown("### 📊 Track Summary Statistics")
        
        # Overall statistics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_tracks = len(all_tracks)
            st.metric("Total Active Tracks", total_tracks)
        
        with col2:
            role_stats = get_role_statistics(all_tracks)
            nurse_count = role_stats.get('nurse', 0)
            st.metric("Nurses", nurse_count)
        
        with col3:
            medic_count = role_stats.get('medic', 0)
            st.metric("Medics", medic_count)
        
        with col4:
            if total_tracks > 0:
                nurse_percentage = round((nurse_count / total_tracks) * 100, 1)
                st.metric("Nurse %", f"{nurse_percentage}%")
        
        # Shift distribution analysis
        st.markdown("#### Shift Distribution Analysis")
        
        shift_analysis = analyze_shift_distribution(all_tracks)
        if shift_analysis:
            analysis_col1, analysis_col2 = st.columns(2)
            
            with analysis_col1:
                st.markdown("**Total Shifts by Type:**")
                for shift_type, count in shift_analysis['total_shifts'].items():
                    st.caption(f"• {shift_type}: {count} shifts")
            
            with analysis_col2:
                st.markdown("**Average Shifts per Staff:**")
                for shift_type, avg in shift_analysis['avg_per_staff'].items():
                    st.caption(f"• {shift_type}: {avg:.1f} shifts/person")
        
    except Exception as e:
        st.error(f"Error displaying statistics: {str(e)}")

def analyze_shift_distribution(tracks):
    """
    Analyze shift distribution across all tracks
    """
    if not tracks:
        return None
    
    shift_totals = {'D': 0, 'N': 0, 'AT': 0}
    staff_count = len(tracks)
    
    for track in tracks:
        track_data = track.get('track_data', {})
        for day, shift in track_data.items():
            if shift in shift_totals:
                shift_totals[shift] += 1
    
    # Calculate averages
    shift_averages = {}
    for shift_type, total in shift_totals.items():
        shift_averages[shift_type] = total / staff_count if staff_count > 0 else 0
    
    return {
        'total_shifts': shift_totals,
        'avg_per_staff': shift_averages,
        'staff_count': staff_count
    }

def display_workload_analysis():
    """
    Display detailed workload analysis for current tracks
    """
    try:
        success, all_tracks = get_all_active_tracks()
        if not success or not all_tracks:
            st.warning("No track data available for analysis.")
            return
        
        st.markdown("### 📈 Staff Workload Analysis")
        
        # Calculate workload for each staff member
        workload_data = []
        
        for track in all_tracks:
            staff_name = track['staff_name']
            metadata = track.get('metadata', {})
            effective_role = metadata.get('effective_role', 'nurse')
            track_data = track.get('track_data', {})
            
            # Count shifts
            shift_counts = {'D': 0, 'N': 0, 'AT': 0, 'Off': 0}
            
            for day, shift in track_data.items():
                if shift in shift_counts:
                    shift_counts[shift] += 1
                else:
                    shift_counts['Off'] += 1
            
            total_working = shift_counts['D'] + shift_counts['N'] + shift_counts['AT']
            total_days = len(track_data)
            workload_percentage = (total_working / total_days * 100) if total_days > 0 else 0
            
            workload_data.append({
                'Staff Name': staff_name,
                'Role': effective_role.title(),
                'Day Shifts': shift_counts['D'],
                'Night Shifts': shift_counts['N'],
                'Admin Time': shift_counts['AT'],
                'Off Days': shift_counts['Off'],
                'Total Working Days': total_working,
                'Workload %': f"{workload_percentage:.1f}%"
            })
        
        # Create DataFrame and display
        if workload_data:
            df = pd.DataFrame(workload_data)
            df_sorted = df.sort_values('Total Working Days', ascending=False)
            
            st.dataframe(df_sorted, use_container_width=True, hide_index=True)
            
            # Summary statistics
            summary_col1, summary_col2, summary_col3 = st.columns(3)
            
            with summary_col1:
                avg_workload = sum(float(row['Workload %'].rstrip('%')) for row in workload_data) / len(workload_data)
                st.metric("Average Workload", f"{avg_workload:.1f}%")
            
            with summary_col2:
                total_day_shifts = sum(row['Day Shifts'] for row in workload_data)
                st.metric("Total Day Shifts", total_day_shifts)
            
            with summary_col3:
                total_night_shifts = sum(row['Night Shifts'] for row in workload_data)
                st.metric("Total Night Shifts", total_night_shifts)
    
    except Exception as e:
        st.error(f"Error displaying workload analysis: {str(e)}")

# Enhanced integration functions

def add_track_display_with_export_to_admin(admin_authenticated=False):
    """
    Add track display with export functionality to admin section
    """
    if admin_authenticated:
        st.markdown("### 📊 Track Display & Export")
        
        try:
            success, all_tracks = get_all_active_tracks()
            if not success or not all_tracks:
                st.warning("No track data available.")
                return
            
            # Export options
            export_col1, export_col2 = st.columns(2)
            
            with export_col1:
                st.markdown("**Export Filtered Tracks:**")
                
                # Role filter for export
                export_role = st.selectbox(
                    "Export Role Filter:",
                    options=['All', 'nurse', 'medic'],
                    key="admin_export_role"
                )
                
                # Staff filter for export
                available_staff = get_available_staff_by_role(all_tracks, export_role)
                staff_filter_options = ['All Staff'] + available_staff
                export_staff_filter = st.selectbox(
                    "Export Staff Filter:",
                    options=staff_filter_options,
                    key="admin_export_staff"
                )
                
                if st.button("📥 Export to Excel", use_container_width=True):
                    file_path, message = export_filtered_tracks_to_excel(export_role, export_staff_filter)
                    
                    if file_path:
                        import base64
                        with open(file_path, "rb") as f:
                            excel_data = f.read()
                        
                        b64 = base64.b64encode(excel_data).decode()
                        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="Active_Tracks_Export.xlsx">Click here to download Excel file</a>'
                        
                        st.success(f"✅ {message}")
                        st.markdown(href, unsafe_allow_html=True)
                        
                        # Cleanup
                        import os
                        os.remove(file_path)
                    else:
                        st.error(f"❌ {message}")
            
            with export_col2:
                st.markdown("**Current Database Statistics:**")
                
                role_stats = get_role_statistics(all_tracks)
                total_tracks = len(all_tracks)
                
                st.info(f"📊 **Total Active Tracks:** {total_tracks}")
                
                for role, count in role_stats.items():
                    percentage = round((count / total_tracks) * 100, 1)
                    st.caption(f"• {role.title()}: {count} ({percentage}%)")
                
                # Show filtered count
                if export_role != 'All' or export_staff_filter != 'All Staff':
                    filtered_tracks = get_tracks_from_database_by_filters(export_role, export_staff_filter)
                    st.success(f"🎯 **Filtered Result:** {len(filtered_tracks)} tracks will be exported")
        
        except Exception as e:
            st.error(f"Error in admin track display: {str(e)}")

# Main execution for testing
if __name__ == "__main__":
    st.set_page_config(
        page_title="Enhanced Track Display",
        page_icon="📊",
        layout="wide"
    )
    
    st.title("📊 Enhanced Track Display Module")
    st.caption("With Comprehensive Role and Staff Filtering")
    
    # Test the enhanced functionality
    display_track_viewer()
    
    # Additional analysis sections
    with st.expander("📊 Summary Statistics", expanded=False):
        display_track_summary_statistics()
    
    with st.expander("📈 Workload Analysis", expanded=False):
        display_workload_analysis()