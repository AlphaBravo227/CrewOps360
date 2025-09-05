# training_modules/track_manager.py - Updated to include CCEMT schedule integration

from datetime import datetime, timedelta
import sqlite3
import json
import pandas as pd
from openpyxl import load_workbook

class TrainingTrackManager:
    """Enhanced Track Manager that includes CCEMT schedule integration"""
    
    def __init__(self, tracks_db_path=None):
        self.tracks_db_path = tracks_db_path
        self.tracks_cache = {}
        self.ccemt_schedule_cache = {}
        self.excel_handler = None  # Will be set externally
        
        # ADD THESE MISSING PROPERTIES:
        self.pattern_start = datetime(2025, 9, 14)  # Sun A 1 - pattern start date
        self.pattern_length = 42  # 6 weeks = 42 days
        
        # Shift descriptions for display
        self.shift_descriptions = {
            'D': 'Day Shift',
            'N': 'Night Shift',
            'AT': 'AT Shift',
            'LT': 'Day Shift (LT)'  # LT treated as Day shift
        }
        
        # Load tracks if database exists
        if self.tracks_db_path:
            self.reload_tracks()

    # ADD THIS MISSING METHOD:
    def get_pattern_day_name(self, date):
        """
        Get the pattern day name (e.g., "Sun A 1") for a given date.
        
        Args:
            date: datetime object or string date
            
        Returns:
            str: Pattern day name
        """
        if isinstance(date, str):
            try:
                date = datetime.strptime(date, '%m/%d/%Y')
            except:
                return ""
        
        # Calculate days since pattern start
        days_since_start = (date - self.pattern_start).days
        
        # Get position in current pattern cycle
        pattern_day_index = days_since_start % self.pattern_length
        
        # Calculate week and day
        week_index = pattern_day_index // 7
        day_index = pattern_day_index % 7
        
        # Define components
        days_of_week = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        week_letters = ["A", "A", "B", "B", "C", "C"]
        week_numbers = [1, 2, 3, 4, 5, 6]
        
        if week_index < 6:
            day_name = days_of_week[day_index]
            week_letter = week_letters[week_index]
            week_number = week_numbers[week_index]
            return f"{day_name} {week_letter} {week_number}"
        
        return f"Day {pattern_day_index + 1}"
    
    def set_excel_handler(self, excel_handler):
        """Set the Excel handler for CCEMT schedule access"""
        self.excel_handler = excel_handler
        self.load_ccemt_schedules()
    
    def load_ccemt_schedules(self):
        """Load CCEMT schedules from the CCEMT CRM Sched tab"""
        if not self.excel_handler or not self.excel_handler.workbook:
            return
        
        try:
            # Access the CCEMT CRM Sched worksheet
            if 'CCEMT CRM Sched' not in self.excel_handler.workbook.sheetnames:
                print("CCEMT CRM Sched tab not found in workbook")
                return
            
            ccemt_sheet = self.excel_handler.workbook['CCEMT CRM Sched']
            
            # Read the header row to get dates (row 1, starting from column B)
            dates = []
            date_columns = {}
            
            for col_idx, col in enumerate(ccemt_sheet.iter_cols(min_col=2, max_col=20, min_row=1, max_row=1), start=2):
                cell_value = col[0].value
                if cell_value:
                    # Convert date to string format that matches class dates
                    if isinstance(cell_value, datetime):
                        date_str = cell_value.strftime('%m/%d/%Y')
                    else:
                        # Try to parse string date
                        try:
                            date_obj = datetime.strptime(str(cell_value), '%m/%d/%Y')
                            date_str = date_obj.strftime('%m/%d/%Y')
                        except:
                            continue
                    
                    dates.append(date_str)
                    date_columns[date_str] = col_idx
                        
            # Read staff schedules (starting from row 2, column A for names)
            for row in ccemt_sheet.iter_rows(min_row=2, max_col=1):
                staff_name = row[0].value
                if not staff_name:
                    continue
                
                staff_name = str(staff_name).strip()
                if not staff_name:
                    continue
                
                # Initialize schedule for this staff member
                self.ccemt_schedule_cache[staff_name] = {}
                
                # Read schedule data for each date
                for date_str, col_idx in date_columns.items():
                    schedule_cell = ccemt_sheet.cell(row=row[0].row, column=col_idx)
                    schedule_value = schedule_cell.value
                    
                    if schedule_value:
                        schedule_str = str(schedule_value).strip().upper()
                        
                        # Normalize schedule values
                        if schedule_str in ['D', 'N', 'LT']:
                            # Treat LT as D for conflict calculations
                            normalized_value = 'D' if schedule_str == 'LT' else schedule_str
                            self.ccemt_schedule_cache[staff_name][date_str] = normalized_value
                        else:
                            # Store original value for any other notation
                            self.ccemt_schedule_cache[staff_name][date_str] = schedule_str                
        
        except Exception as e:
            print(f"Error loading CCEMT schedules: {e}")
            import traceback
            traceback.print_exc()
    
    def reload_tracks(self):
        """Reload track data from database"""
        if not self.tracks_db_path:
            return
        
        try:
            conn = sqlite3.connect(self.tracks_db_path)
            cursor = conn.cursor()
            
            # Get active tracks
            cursor.execute("""
                SELECT staff_name, track_data 
                FROM tracks 
                WHERE is_active = 1
            """)
            
            results = cursor.fetchall()
            self.tracks_cache = {}
            
            for staff_name, track_data_json in results:
                if track_data_json:
                    try:
                        track_data = json.loads(track_data_json)
                        self.tracks_cache[staff_name] = track_data
                    except json.JSONDecodeError:
                        continue
            
            conn.close()
            print(f"Loaded {len(self.tracks_cache)} tracks from database")
            
        except Exception as e:
            print(f"Error loading tracks: {e}")
    
    def get_staff_role(self, staff_name):
        """Get the role of a staff member from Excel handler"""
        if not self.excel_handler:
            return None
        
        try:
            # Access the enrollment sheet to get role information
            enrollment_sheet = self.excel_handler.enrollment_sheet
            if not enrollment_sheet:
                return None
            
            # Find staff member's row and get their role
            for row in enrollment_sheet.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == staff_name:
                    # Role is typically in column B (index 1)
                    role_cell = row[1].value if len(row) > 1 else None
                    return str(role_cell).strip() if role_cell else None
            
            return None
            
        except Exception as e:
            print(f"Error getting staff role for {staff_name}: {e}")
            return None
    
    def has_track_data(self, staff_name):
        """Check if staff member has track data (regular track OR CCEMT schedule)"""
        # Check regular track data first
        if staff_name in self.tracks_cache:
            return True
        
        # Check if staff member is CCEMT with schedule data
        staff_role = self.get_staff_role(staff_name)
        if staff_role == 'CCEMT' and hasattr(self, 'ccemt_schedule_cache') and staff_name in self.ccemt_schedule_cache:
            return True
        
        return False
    
    def get_staff_shift(self, staff_name, date):
            """
            Get the shift assignment for a staff member on a specific date.
            Handles both regular track staff and CCEMT staff.
            
            Args:
                staff_name: Name of the staff member
                date: Date to check (datetime or string)
                
            Returns:
                str: Shift code (D, N, AT, or empty string)
            """
            # Convert date to string format if needed
            if isinstance(date, datetime):
                date_str = date.strftime('%m/%d/%Y')
            else:
                date_str = str(date)
            
            # Check if this is a CCEMT staff member
            staff_role = self.get_staff_role(staff_name)
            
            if staff_role == 'CCEMT':
                # Use CCEMT schedule data
                if hasattr(self, 'ccemt_schedule_cache') and staff_name in self.ccemt_schedule_cache:
                    return self.ccemt_schedule_cache[staff_name].get(date_str, "")
                else:
                    return ""  # No CCEMT schedule data available
            
            # Fall back to existing logic for regular track staff
            if staff_name not in self.tracks_cache:
                return ""  # No track data available (return empty string, not None)
            
            # Use pattern day logic for regular tracks
            pattern_day = self.get_pattern_day_name(date)
            if not pattern_day:
                return ""
            
            track_data = self.tracks_cache.get(staff_name, {})
            return track_data.get(pattern_day, "")    
    
    def check_class_conflict(self, staff_name, class_date, is_two_day=False, can_work_n_prior=False):
        """
        Check if a staff member has a conflict with a class date.
        Works for both regular track staff and CCEMT staff.
        
        Args:
            staff_name: Name of the staff member
            class_date: Date of the class (datetime or string)
            is_two_day: Whether this is a two-day class
            can_work_n_prior: Whether night shift prior is allowed for this class
            
        Returns:
            tuple: (has_conflict, conflict_details)
        """
        if isinstance(class_date, str):
            try:
                class_date = datetime.strptime(class_date, '%m/%d/%Y')
            except:
                return (False, "Invalid date format")
        
        # Check if we have any data for this staff member
        if not self.has_track_data(staff_name):
            return (False, "No schedule data available")
        
        conflicts = []
        
        # Check day before class
        day_before = class_date - timedelta(days=1)
        shift_before = self.get_staff_shift(staff_name, day_before)
        
        if shift_before == 'N' and not can_work_n_prior:
            conflicts.append(f"Night shift on {day_before.strftime('%m/%d/%Y')}")
        
        # Check class day
        shift_class_day = self.get_staff_shift(staff_name, class_date)
        if shift_class_day in ['D', 'AT', 'N']:
            shift_desc = self.shift_descriptions.get(shift_class_day, shift_class_day)
            conflicts.append(f"{shift_desc} on {class_date.strftime('%m/%d/%Y')}")
        
        # Check second day if two-day class
        if is_two_day:
            day_two = class_date + timedelta(days=1)
            shift_day_two = self.get_staff_shift(staff_name, day_two)
            
            if shift_day_two in ['D', 'AT', 'N']:
                shift_desc = self.shift_descriptions.get(shift_day_two, shift_day_two)
                conflicts.append(f"{shift_desc} on {day_two.strftime('%m/%d/%Y')}")
        
        has_conflict = len(conflicts) > 0
        conflict_details = "; ".join(conflicts) if conflicts else "No conflicts"
        
        return (has_conflict, conflict_details)
    
    def get_class_date_conflicts(self, staff_name, class_dates, is_two_day=False, can_work_n_prior_list=None):
        """
        Check conflicts for multiple class dates.
        
        Args:
            staff_name: Name of the staff member
            class_dates: List of class dates
            is_two_day: Whether this is a two-day class
            can_work_n_prior_list: List of booleans for each date indicating if N prior is allowed
            
        Returns:
            dict: Dictionary mapping dates to conflict information
        """
        if can_work_n_prior_list is None:
            can_work_n_prior_list = [False] * len(class_dates)
        
        conflicts = {}
        
        for i, date in enumerate(class_dates):
            if date:  # Skip None/empty dates
                can_work_n = can_work_n_prior_list[i] if i < len(can_work_n_prior_list) else False
                has_conflict, details = self.check_class_conflict(
                    staff_name, date, is_two_day, can_work_n
                )
                
                # Get the shift for display
                shift = self.get_staff_shift(staff_name, date)
                
                conflicts[date] = {
                    'has_conflict': has_conflict,
                    'details': details,
                    'shift': shift,
                    'shift_desc': self.shift_descriptions.get(shift, 'Off')
                }
        
        return conflicts
    
    def get_conflict_summary(self, conflicts_dict):
        """
        Generate a summary of conflicts for a class.
        
        Args:
            conflicts_dict: Dictionary of date conflicts
            
        Returns:
            str: Summary text
        """
        if not conflicts_dict:
            return "No schedule data available"
        
        # Handle case where conflicts_dict is not in expected format
        if not isinstance(conflicts_dict, dict):
            return "No schedule data available"
        
        # Check if the conflicts_dict has the expected structure
        if not conflicts_dict or all(not isinstance(v, dict) for v in conflicts_dict.values()):
            return "No schedule data available"
        
        total_dates = len(conflicts_dict)
        
        # Safely count conflicts by checking if 'has_conflict' key exists
        conflict_count = 0
        for conflict_info in conflicts_dict.values():
            if isinstance(conflict_info, dict) and conflict_info.get('has_conflict', False):
                conflict_count += 1
        
        if conflict_count == 0:
            return f"✅ All {total_dates} dates available"
        elif conflict_count == total_dates:
            return f"⚠️ Conflicts on all {total_dates} dates"
        else:
            return f"⚠️ {conflict_count} of {total_dates} dates have conflicts"
    
    def get_all_staff_with_tracks(self):
        """Get list of all staff with track data (regular tracks OR CCEMT schedules)"""
        all_staff = set()
        
        # Add staff with regular tracks
        all_staff.update(self.tracks_cache.keys())
        
        # Add CCEMT staff with schedules
        if hasattr(self, 'ccemt_schedule_cache'):
            all_staff.update(self.ccemt_schedule_cache.keys())
        
        return list(all_staff)

# Integration points for existing codebase:

def integrate_ccemt_schedules(track_manager, excel_handler):
    """
    Helper function to integrate CCEMT schedules into an existing track manager.
    This should be called after both components are initialized.
    """
    track_manager.set_excel_handler(excel_handler)

# Usage example for app.py integration:
"""
# In app.py, after initializing both components:

if 'training_track_manager' not in st.session_state:
    st.session_state.training_track_manager = TrainingTrackManager('data/medflight_tracks.db')

if 'training_excel_handler' not in st.session_state:
    st.session_state.training_excel_handler = ExcelHandler(excel_path)

# Integrate CCEMT schedules
integrate_ccemt_schedules(
    st.session_state.training_track_manager, 
    st.session_state.training_excel_handler
)
"""