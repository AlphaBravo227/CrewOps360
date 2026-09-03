# training_modules/class_catalog.py
"""
The class catalog: classes, their dates and their assigned staff, held in the database.

Until FY27 this lived in the roster workbook — one column per class on the
`Class_Enrollment` sheet, one detail sheet per class with its settings at fixed cell
addresses. That layout is what capped a class at 14 dates and at one location per date,
because rows 1-14 of column B were the dates and column E was the single location beside
each. Both limits were the spreadsheet's, not the program's.

Classes now live in four tables and the workbook is an import source only (see
`import_workbook`, wired to the "Import roster workbook" button in Training Admin):

    training_classes              one row per class per training year, holding the
                                  settings that used to sit at fixed cells
    training_class_dates          one row per date, any number of them
    training_class_options        the bookable options on a date — a location, its own
                                  times and its own seat count. One option per date is
                                  the ordinary case and matches how the workbook read;
                                  several options is a day taught at several sites.
    training_class_assignments    who is assigned to the class, one row per person

Everything is scoped by `training_year`, so FY26 and FY27 hold separate classes under
the same name and a closed year keeps reporting its own schedule.

`ClassCatalog` below re-implements the reader the rest of the training module already
calls — the same method names, the same flat `date_1`, `date_2`, … detail dictionary —
so enrollment, educator signups, conflict checking and the admin reports did not have to
change to stop reading Excel. The one addition is `date_count`: callers that used to
loop `range(1, 15)` now loop to the number of dates a class actually has.
"""

import json
import os
import sqlite3
from datetime import datetime, time, timedelta

try:
    from modules import staff_database as staffdb
except Exception as e:  # pragma: no cover - the staff database is optional at import
    staffdb = None
    print(f"Staff database unavailable to the class catalog: {e}")


DEFAULT_DB_PATH = os.path.join('data', 'medflight_tracks.db')

# Columns on the workbook's Class_Enrollment sheet that name something other than a
# class. Only needed while importing; nothing reads the sheet otherwise.
NON_CLASS_COLUMNS = [
    'STAFF NAME',
    'Role',
    'MGMT',
    'DUAL',
    'Educator AT',
]

# What a class looks like before anyone configures it. Also what a lookup for a class
# that isn't in the catalog returns, so a caller gets a usable shape rather than None.
DEFAULT_CLASS_DETAILS = {
    'students_per_class': 21,
    'nurses_medic_separate': 'No',
    'classes_per_day': 1,
    'is_two_day_class': 'No',
    'time_1_start': '08:00',
    'time_1_end': '16:00',
    'instructors_per_day': 0,
    'is_multi_session': 'No',
    'session_length': None,
    'is_count_exempt': False,
    'has_ccemt': 'No',
    'calendar_display': '',
    'is_staff_meeting': False,
    'date_count': 0,
}

# The settings that used to live at fixed cells on a class's detail sheet, paired with
# the cell each one came from. Kept here rather than in the importer so the mapping from
# the old layout to the new columns is readable in one place.
CLASS_SETTING_COLUMNS = (
    'has_ccemt', 'is_multi_session', 'session_length', 'is_count_exempt',
    'students_per_class', 'nurses_medic_separate', 'classes_per_day',
    'is_two_day_class', 'time_1_start', 'time_1_end', 'time_2_start', 'time_2_end',
    'time_3_start', 'time_3_end', 'time_4_start', 'time_4_end', 'instructors_per_day',
    'calendar_display',
)


# ---------------------------------------------------------------------------
# Value parsing
#
# Imported values arrive as whatever openpyxl made of a spreadsheet cell, and
# admin-entered values arrive from Streamlit widgets. Both go through here so a class
# built by hand and a class read out of a workbook store identically.
# ---------------------------------------------------------------------------

def parse_checkbox(value):
    """A checkbox cell or form value as a bool."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        cleaned = value.strip().upper()
        if cleaned in ('YES', 'Y', 'TRUE', 'T', 'X', '✓', '1'):
            return True
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def parse_time(value):
    """A time cell or form value as 'HH:MM', or None when there isn't one."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        if ':' in value:
            return value
        if value.isdigit():
            if len(value) == 3:
                return f"{value[0]}:{value[1:3]}"
            if len(value) == 4:
                return f"{value[0:2]}:{value[2:4]}"
        return value
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    if isinstance(value, time):
        return value.strftime('%H:%M')
    if isinstance(value, (int, float)):
        # Excel keeps a time as a fraction of a day.
        hours = int(value * 24)
        minutes = int(round((value * 24 - hours) * 60))
        return f"{hours:02d}:{minutes:02d}"
    return None


def parse_date(value):
    """A date cell or form value as 'MM/DD/YYYY', or None when there isn't one."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%m/%d/%Y')
    if hasattr(value, 'strftime'):  # datetime.date
        return value.strftime('%m/%d/%Y')
    text = str(value).strip()
    if not text:
        return None
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%m-%d-%Y'):
        try:
            return datetime.strptime(text, fmt).strftime('%m/%d/%Y')
        except ValueError:
            continue
    return text


def parse_int(value, default=None):
    """An integer cell or form value, or `default` when it isn't one."""
    if value is None or value == '':
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def date_indices(class_details):
    """
    The 1-based date positions a class's details actually carry.

    Call sites used to walk `range(1, 15)` because the workbook gave a class fourteen
    date rows and no way to say how many were filled. A class now has as many dates as
    it has, and this is how a caller asks. `date_count` is always present on details
    that came from the catalog; counting the keys covers a dictionary built by hand.
    """
    details = class_details or {}
    count = details.get('date_count')
    if count is None:
        count = 0
        while details.get(f'date_{count + 1}'):
            count += 1
    return range(1, int(count) + 1)


def sort_key(date_str):
    """Sort dates chronologically, keeping unparseable ones last in entry order."""
    try:
        return (0, datetime.strptime(date_str, '%m/%d/%Y'))
    except (TypeError, ValueError):
        return (1, datetime.max)


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

def initialize_catalog_tables(db_path=DEFAULT_DB_PATH):
    """Create the catalog tables. Safe to call on every start."""
    os.makedirs(os.path.dirname(db_path) or '.', exist_ok=True)
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_classes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            training_year TEXT NOT NULL,
            class_name TEXT NOT NULL,
            display_order INTEGER DEFAULT 0,
            has_ccemt INTEGER DEFAULT 0,
            is_multi_session INTEGER DEFAULT 0,
            session_length INTEGER,
            is_count_exempt INTEGER DEFAULT 0,
            students_per_class INTEGER DEFAULT 21,
            nurses_medic_separate INTEGER DEFAULT 0,
            classes_per_day INTEGER DEFAULT 1,
            is_two_day_class INTEGER DEFAULT 0,
            time_1_start TEXT,
            time_1_end TEXT,
            time_2_start TEXT,
            time_2_end TEXT,
            time_3_start TEXT,
            time_3_end TEXT,
            time_4_start TEXT,
            time_4_end TEXT,
            instructors_per_day INTEGER DEFAULT 0,
            is_staff_meeting INTEGER,
            calendar_display TEXT,
            assignment_source TEXT,
            source TEXT DEFAULT 'app',
            notes TEXT,
            created_date TEXT,
            modified_date TEXT,
            UNIQUE(training_year, class_name)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_class_dates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            date_index INTEGER NOT NULL,
            class_date TEXT NOT NULL,
            has_live INTEGER DEFAULT 0,
            can_work_n_prior INTEGER DEFAULT 0,
            UNIQUE(class_id, class_date),
            FOREIGN KEY (class_id) REFERENCES training_classes(id) ON DELETE CASCADE
        )
    ''')

    # A date's bookable options. `location` is what a staff member picks between when a
    # day runs at more than one site; `start_time`/`end_time`/`capacity` fall back to the
    # class-level settings when left empty, which is what an imported date does.
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_class_options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_id INTEGER NOT NULL,
            option_index INTEGER NOT NULL,
            location TEXT,
            start_time TEXT,
            end_time TEXT,
            capacity INTEGER,
            UNIQUE(date_id, option_index),
            FOREIGN KEY (date_id) REFERENCES training_class_dates(id) ON DELETE CASCADE
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS training_class_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER NOT NULL,
            staff_name TEXT NOT NULL,
            UNIQUE(class_id, staff_name),
            FOREIGN KEY (class_id) REFERENCES training_classes(id) ON DELETE CASCADE
        )
    ''')

    # Added after the catalog was already in use, so a database built before it needs
    # the column adding rather than the table recreating.
    existing_columns = {row[1] for row in
                        cursor.execute("PRAGMA table_info(training_classes)")}
    if 'calendar_display' not in existing_columns:
        cursor.execute('ALTER TABLE training_classes ADD COLUMN calendar_display TEXT')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_class_year '
                   'ON training_classes(training_year)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_class_dates_class '
                   'ON training_class_dates(class_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_class_options_date '
                   'ON training_class_options(date_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_class_assignments_staff '
                   'ON training_class_assignments(staff_name)')

    conn.commit()
    conn.close()


def _connect(db_path):
    conn = sqlite3.connect(db_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return conn


def _now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def save_class(training_year, class_name, settings=None, dates=None, assigned_staff=None,
               class_id=None, source='app', assignment_source=None,
               db_path=DEFAULT_DB_PATH):
    """
    Create or update one class, with its dates and its assigned staff.

    `dates` is a list of dicts, each one date:

        {'class_date': '11/12/2026',
         'has_live': False,
         'can_work_n_prior': True,
         'options': [{'location': 'Bedford',   'start_time': '08:00',
                      'end_time': '16:00', 'capacity': 21},
                     {'location': 'Worcester', ...}]}

    A date with no options gets one unnamed option, which is how a class taught at a
    single site behaves. Dates and options are rewritten wholesale on each save: they
    are the class's schedule, and reconciling them row by row would leave an option
    behind whenever an admin removed one.

    Returns the class's id.
    """
    settings = dict(settings or {})
    dates = list(dates or [])
    assigned_staff = list(assigned_staff or [])

    class_name = str(class_name).strip()
    if not class_name:
        raise ValueError("A class needs a name")

    initialize_catalog_tables(db_path)
    conn = _connect(db_path)
    cursor = conn.cursor()
    try:
        columns = {
            'has_ccemt': int(parse_checkbox(settings.get('has_ccemt'))),
            'is_multi_session': int(parse_checkbox(settings.get('is_multi_session'))),
            'session_length': parse_int(settings.get('session_length')),
            'is_count_exempt': int(parse_checkbox(settings.get('is_count_exempt'))),
            'students_per_class': parse_int(settings.get('students_per_class'), 21),
            'nurses_medic_separate': int(parse_checkbox(
                settings.get('nurses_medic_separate'))),
            'classes_per_day': parse_int(settings.get('classes_per_day'), 1),
            'is_two_day_class': int(parse_checkbox(settings.get('is_two_day_class'))),
            'instructors_per_day': parse_int(settings.get('instructors_per_day'), 0),
            'notes': settings.get('notes') or None,
            # The short label the comprehensive schedule report prints in a day's
            # cell. NULL when the admin left it blank, which is what tells the report
            # to fall back to the full class name.
            'calendar_display': (str(settings.get('calendar_display')).strip() or None
                                 if settings.get('calendar_display') else None),
            'assignment_source': (json.dumps(assignment_source)
                                  if assignment_source is not None else None),
        }
        for slot in ('time_1_start', 'time_1_end', 'time_2_start', 'time_2_end',
                     'time_3_start', 'time_3_end', 'time_4_start', 'time_4_end'):
            columns[slot] = parse_time(settings.get(slot))
        columns['time_1_start'] = columns['time_1_start'] or '08:00'
        columns['time_1_end'] = columns['time_1_end'] or '16:00'

        # `is_staff_meeting` stays NULL unless an admin sets it, in which case it
        # overrides the "SM appears in the name" rule the workbook relied on. A class
        # called "SMART course" is not a staff meeting; one called "Q1 Meeting" may be.
        meeting_flag = settings.get('is_staff_meeting')
        columns['is_staff_meeting'] = (None if meeting_flag is None
                                       else int(parse_checkbox(meeting_flag)))

        if class_id:
            columns['class_name'] = class_name
            columns['modified_date'] = _now()
            assignments = ', '.join(f"{k} = ?" for k in columns)
            cursor.execute(
                f"UPDATE training_classes SET {assignments} "
                f"WHERE id = ? AND training_year = ?",
                list(columns.values()) + [class_id, training_year])
            if cursor.rowcount == 0:
                raise ValueError(f"No class {class_id} in {training_year} to update")
        else:
            existing = cursor.execute(
                "SELECT id FROM training_classes "
                "WHERE training_year = ? AND class_name = ?",
                (training_year, class_name)).fetchone()
            if existing:
                conn.close()
                return save_class(training_year, class_name, settings, dates,
                                  assigned_staff, class_id=existing['id'],
                                  source=source, assignment_source=assignment_source,
                                  db_path=db_path)
            columns['training_year'] = training_year
            columns['class_name'] = class_name
            columns['source'] = source
            columns['created_date'] = _now()
            columns['modified_date'] = _now()
            placeholders = ', '.join('?' for _ in columns)
            cursor.execute(
                f"INSERT INTO training_classes ({', '.join(columns)}) "
                f"VALUES ({placeholders})", list(columns.values()))
            class_id = cursor.lastrowid

        _write_dates(cursor, class_id, dates)
        _write_assignments(cursor, class_id, assigned_staff)
        conn.commit()
        return class_id
    finally:
        conn.close()


def _write_dates(cursor, class_id, dates):
    """Replace a class's dates and their options."""
    cursor.execute("DELETE FROM training_class_dates WHERE class_id = ?", (class_id,))

    prepared = []
    for entry in dates:
        class_date = parse_date(entry.get('class_date'))
        if not class_date:
            continue
        prepared.append((class_date, entry))
    # Chronological, so a date added later lands in the right place rather than at the
    # end. The date_index is only ever an ordering, never an identity.
    prepared.sort(key=lambda pair: sort_key(pair[0]))

    seen = set()
    index = 0
    for class_date, entry in prepared:
        if class_date in seen:
            continue
        seen.add(class_date)
        index += 1
        cursor.execute(
            "INSERT INTO training_class_dates "
            "(class_id, date_index, class_date, has_live, can_work_n_prior) "
            "VALUES (?, ?, ?, ?, ?)",
            (class_id, index, class_date,
             int(parse_checkbox(entry.get('has_live'))),
             int(parse_checkbox(entry.get('can_work_n_prior')))))
        date_id = cursor.lastrowid

        options = [opt for opt in (entry.get('options') or [])
                   if any(str(opt.get(field) or '').strip()
                          for field in ('location', 'start_time', 'end_time', 'capacity'))]
        if not options:
            # No options entered means one option with nothing said about it: the class
            # times and the class capacity apply, and no location is named.
            options = [{}]
        for option_index, option in enumerate(options, start=1):
            cursor.execute(
                "INSERT INTO training_class_options "
                "(date_id, option_index, location, start_time, end_time, capacity) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (date_id, option_index,
                 (str(option.get('location')).strip()
                  if option.get('location') else None),
                 parse_time(option.get('start_time')),
                 parse_time(option.get('end_time')),
                 parse_int(option.get('capacity'))))


def _write_assignments(cursor, class_id, assigned_staff):
    """Replace who is assigned to a class."""
    cursor.execute("DELETE FROM training_class_assignments WHERE class_id = ?",
                   (class_id,))
    for name in dict.fromkeys(str(n).strip() for n in assigned_staff if str(n).strip()):
        cursor.execute(
            "INSERT OR IGNORE INTO training_class_assignments (class_id, staff_name) "
            "VALUES (?, ?)", (class_id, name))


def delete_class(training_year, class_name, db_path=DEFAULT_DB_PATH):
    """
    Remove a class from the catalog.

    Enrollments are not touched: they are records of who attended what and are keyed by
    class name, not by the catalog row. Deleting a class that people have enrolled in
    leaves those enrollments without a schedule to display, so the admin page checks the
    enrollment count and says so before offering this.
    """
    conn = _connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM training_classes "
                       "WHERE training_year = ? AND class_name = ?",
                       (training_year, class_name))
        removed = cursor.rowcount
        conn.commit()
        return removed > 0
    finally:
        conn.close()


def rename_class(training_year, old_name, new_name, db_path=DEFAULT_DB_PATH):
    """
    Rename a class, carrying its enrollments and educator signups with it.

    Both of those tables key on the class name, so renaming the catalog row alone would
    orphan every record made under the old name.
    """
    old_name = str(old_name).strip()
    new_name = str(new_name).strip()
    if not new_name:
        raise ValueError("A class needs a name")
    if old_name == new_name:
        return True

    conn = _connect(db_path)
    try:
        cursor = conn.cursor()
        clash = cursor.execute(
            "SELECT id FROM training_classes WHERE training_year = ? AND class_name = ?",
            (training_year, new_name)).fetchone()
        if clash:
            raise ValueError(f"{training_year} already has a class called '{new_name}'")

        cursor.execute("UPDATE training_classes SET class_name = ?, modified_date = ? "
                       "WHERE training_year = ? AND class_name = ?",
                       (new_name, _now(), training_year, old_name))
        if cursor.rowcount == 0:
            raise ValueError(f"No class '{old_name}' in {training_year}")

        for table in ('training_enrollments', 'training_educator_signups',
                      'training_enrollment_audit', 'training_educator_audit'):
            try:
                cursor.execute(f"UPDATE {table} SET class_name = ? "
                               f"WHERE class_name = ? AND training_year = ?",
                               (new_name, old_name, training_year))
            except sqlite3.OperationalError:
                # The table isn't there yet on a database that has never run the
                # training module. Nothing to carry over.
                pass
        conn.commit()
        return True
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def get_class_row(training_year, class_name, db_path=DEFAULT_DB_PATH):
    """The raw catalog row for one class, or None."""
    conn = _connect(db_path)
    try:
        row = conn.execute(
            "SELECT * FROM training_classes WHERE training_year = ? AND class_name = ? "
            "COLLATE NOCASE", (training_year, class_name)).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


def get_class_names(training_year, db_path=DEFAULT_DB_PATH):
    """Every class in a training year, in display order then name."""
    conn = _connect(db_path)
    try:
        rows = conn.execute(
            "SELECT class_name FROM training_classes WHERE training_year = ? "
            "ORDER BY display_order, class_name", (training_year,)).fetchall()
        return [row['class_name'] for row in rows]
    finally:
        conn.close()


def get_dates_with_options(class_id, db_path=DEFAULT_DB_PATH, conn=None):
    """A class's dates, each with its bookable options, in chronological order."""
    own_connection = conn is None
    conn = conn or _connect(db_path)
    try:
        date_rows = conn.execute(
            "SELECT * FROM training_class_dates WHERE class_id = ? "
            "ORDER BY date_index", (class_id,)).fetchall()
        dates = []
        for date_row in date_rows:
            option_rows = conn.execute(
                "SELECT * FROM training_class_options WHERE date_id = ? "
                "ORDER BY option_index", (date_row['id'],)).fetchall()
            dates.append({
                'id': date_row['id'],
                'date_index': date_row['date_index'],
                'class_date': date_row['class_date'],
                'has_live': bool(date_row['has_live']),
                'can_work_n_prior': bool(date_row['can_work_n_prior']),
                'options': [{'id': o['id'],
                             'option_index': o['option_index'],
                             'location': o['location'] or '',
                             'start_time': o['start_time'],
                             'end_time': o['end_time'],
                             'capacity': o['capacity']} for o in option_rows],
            })
        return dates
    finally:
        if own_connection:
            conn.close()


def get_assigned_staff(class_id, db_path=DEFAULT_DB_PATH, conn=None):
    """Who is assigned to a class, alphabetically."""
    own_connection = conn is None
    conn = conn or _connect(db_path)
    try:
        rows = conn.execute(
            "SELECT staff_name FROM training_class_assignments WHERE class_id = ? "
            "ORDER BY staff_name", (class_id,)).fetchall()
        return [row['staff_name'] for row in rows]
    finally:
        if own_connection:
            conn.close()


def load_class_for_editing(training_year, class_name, db_path=DEFAULT_DB_PATH):
    """
    Everything the edit form needs for one class: its settings, dates and assignments,
    in the shape `save_class` takes back, so the editor round-trips without translating.
    """
    row = get_class_row(training_year, class_name, db_path=db_path)
    if not row:
        return None
    settings = {key: row[key] for key in CLASS_SETTING_COLUMNS if key in row}
    settings['is_staff_meeting'] = row['is_staff_meeting']
    settings['notes'] = row['notes']
    try:
        assignment_source = (json.loads(row['assignment_source'])
                             if row['assignment_source'] else None)
    except (TypeError, ValueError):
        assignment_source = None
    return {
        'id': row['id'],
        'class_name': row['class_name'],
        'source': row['source'],
        'settings': settings,
        'dates': get_dates_with_options(row['id'], db_path=db_path),
        'assigned_staff': get_assigned_staff(row['id'], db_path=db_path),
        'assignment_source': assignment_source,
    }


class ClassCatalog:
    """
    The reader the training module uses for classes, dates and assignments.

    Deliberately the same surface the workbook reader had — `get_all_classes`,
    `get_class_details`, `get_assigned_classes`, `get_class_dates`,
    `get_available_dates_with_options`, `is_staff_meeting`, `needs_educators`,
    `get_educator_requirement`, plus the staff-attribute lookups that already read the
    staff database. Call sites did not change when the source did.
    """

    def __init__(self, training_year, db_path=DEFAULT_DB_PATH):
        self.training_year = training_year
        self.db_path = db_path
        self.load_error = None
        self._details_cache = {}
        try:
            initialize_catalog_tables(db_path)
        except Exception as e:
            self.load_error = f"Could not open the class catalog: {e}"
            print(self.load_error)

    # -- cache ------------------------------------------------------------
    # A single enrollment screen asks for the same class's details many times over
    # (dates, then session options per date, then conflict checks per date). Reading
    # them once per render keeps that to one query instead of dozens.

    def invalidate(self, class_name=None):
        """Forget cached details, after an admin edits a class."""
        if class_name is None:
            self._details_cache.clear()
        else:
            self._details_cache.pop(str(class_name).strip().lower(), None)

    # -- classes ----------------------------------------------------------

    def get_all_classes(self):
        """Every class in this training year."""
        try:
            return get_class_names(self.training_year, db_path=self.db_path)
        except Exception as e:
            print(f"Error listing classes: {e}")
            return []

    def get_class_details(self, class_name):
        """
        One class's settings and schedule, as the flat dictionary the module expects:

            date_1, date_1_has_live, date_1_can_work_n_prior, date_1_location,
            date_1_options, date_2, … plus date_count and the class-level settings.

        `date_N_location` is the single location when a date has one, and the locations
        joined with " / " when it has several, which is what the read-only displays want.
        `date_N_options` carries them separately, for booking.

        A class that isn't in the catalog comes back as defaults flagged
        `_missing_sheet`, the same shape the workbook reader used for a class whose
        detail sheet was missing — the screens that warn "class not configured" key on
        that flag.
        """
        cache_key = str(class_name).strip().lower()
        if cache_key in self._details_cache:
            return dict(self._details_cache[cache_key])

        details = self._read_class_details(class_name)
        self._details_cache[cache_key] = details
        return dict(details)

    def _read_class_details(self, class_name):
        def unconfigured(flag):
            fallback = DEFAULT_CLASS_DETAILS.copy()
            fallback['class_name'] = class_name
            fallback[flag] = True
            return fallback

        try:
            conn = _connect(self.db_path)
        except Exception as e:
            print(f"Error opening the catalog for '{class_name}': {e}")
            return unconfigured('_error')

        try:
            row = conn.execute(
                "SELECT * FROM training_classes "
                "WHERE training_year = ? AND class_name = ? COLLATE NOCASE",
                (self.training_year, class_name)).fetchone()
            if not row:
                return unconfigured('_missing_sheet')

            dates = get_dates_with_options(row['id'], conn=conn)
            if not dates:
                return unconfigured('_missing_dates')

            details = {
                'class_name': row['class_name'],
                'has_ccemt': 'Yes' if row['has_ccemt'] else 'No',
                'is_multi_session': 'Yes' if row['is_multi_session'] else 'No',
                'session_length': row['session_length'],
                'is_count_exempt': bool(row['is_count_exempt']),
                'students_per_class': row['students_per_class'] or 21,
                'nurses_medic_separate': 'Yes' if row['nurses_medic_separate'] else 'No',
                'classes_per_day': row['classes_per_day'] or 1,
                'is_two_day_class': 'Yes' if row['is_two_day_class'] else 'No',
                'instructors_per_day': row['instructors_per_day'] or 0,
                'calendar_display': (row['calendar_display'] or ''
                                     if 'calendar_display' in row.keys() else ''),
                # NULL means no admin ever set it and the "SM" in the name rule that
                # predates the field decides, which is the answer `is_staff_meeting`
                # gives. Four displays read this key - the enrollment summary's facts,
                # its LIVE/Virtual enrollment breakdown, Class Details and the educator
                # screen - and while it was missing from here every one of them read a
                # staff meeting as an ordinary class.
                'is_staff_meeting': (bool(row['is_staff_meeting'])
                                     if row['is_staff_meeting'] is not None
                                     else 'SM' in str(row['class_name']).upper()),
                'date_count': len(dates),
                '_class_id': row['id'],
            }
            for slot in ('time_1_start', 'time_1_end', 'time_2_start', 'time_2_end',
                         'time_3_start', 'time_3_end', 'time_4_start', 'time_4_end'):
                details[slot] = row[slot]
            details['time_1_start'] = details['time_1_start'] or '08:00'
            details['time_1_end'] = details['time_1_end'] or '16:00'

            for index, date in enumerate(dates, start=1):
                details[f'date_{index}'] = date['class_date']
                details[f'date_{index}_has_live'] = date['has_live']
                details[f'date_{index}_can_work_n_prior'] = date['can_work_n_prior']
                details[f'date_{index}_options'] = date['options']
                locations = list(dict.fromkeys(
                    option['location'] for option in date['options']
                    if option['location']))
                details[f'date_{index}_location'] = ' / '.join(locations)

            return details
        except Exception as e:
            print(f"Error reading class details for '{class_name}': {e}")
            import traceback
            traceback.print_exc()
            return unconfigured('_error')
        finally:
            conn.close()

    def get_calendar_display(self, class_name):
        """The short label a class shows on the comprehensive schedule report.

        Empty when the admin never set one — the report then prints the full class
        name, which is what every class did before the field existed.
        """
        details = self.get_class_details(class_name)
        return (details.get('calendar_display') or '').strip()

    def has_class_data(self, class_name):
        """True when a class is configured with at least one date."""
        details = self.get_class_details(class_name)
        if (details.get('_missing_sheet') or details.get('_missing_dates')
                or details.get('_error')):
            return False
        return details.get('date_count', 0) > 0

    def get_class_dates(self, class_name):
        """A class's dates, chronologically."""
        details = self.get_class_details(class_name)
        if not self.has_class_data(class_name):
            return []
        return [details[f'date_{i}']
                for i in range(1, details.get('date_count', 0) + 1)
                if details.get(f'date_{i}')]

    def get_date_options(self, class_name, class_date):
        """
        The bookable options on one date — one per location the class runs at that day.

        Each option carries the location, the times and the seat count that apply to it,
        already resolved against the class-level settings, so a caller never has to know
        which of the two a value came from.
        """
        details = self.get_class_details(class_name)
        for index in range(1, details.get('date_count', 0) + 1):
            if details.get(f'date_{index}') != class_date:
                continue
            resolved = []
            for option in details.get(f'date_{index}_options') or []:
                resolved.append({
                    'location': option.get('location') or '',
                    'start_time': option.get('start_time') or details.get('time_1_start'),
                    'end_time': option.get('end_time') or details.get('time_1_end'),
                    'capacity': (option.get('capacity')
                                 or parse_int(details.get('students_per_class'), 21)),
                    'option_index': option.get('option_index', 1),
                })
            return resolved
        return []

    def get_date_attributes(self, class_name, class_date):
        """
        What is configured for one date of a class: its LIVE option, whether staff may
        work the night before, and its location text.

        Replaces the "walk date_1 … date_14 looking for a matching date" loop that was
        written out at each call site, and which is what capped a class at 14 dates.
        """
        details = self.get_class_details(class_name)
        for index in range(1, details.get('date_count', 0) + 1):
            if details.get(f'date_{index}') == class_date:
                return {
                    'has_live': details.get(f'date_{index}_has_live', False),
                    'can_work_n_prior': details.get(
                        f'date_{index}_can_work_n_prior', False),
                    'location': details.get(f'date_{index}_location', ''),
                    'options': details.get(f'date_{index}_options') or [],
                }
        return {'has_live': False, 'can_work_n_prior': False,
                'location': '', 'options': []}

    def is_staff_meeting(self, class_name):
        """
        True for a staff meeting.

        The workbook had no field for this and the module inferred it from "SM" in the
        name, which is why a class had to be named a certain way to behave like a
        meeting. The catalog stores the answer; the name rule stays as the fallback so
        every imported class keeps the behaviour it had.
        """
        details = self.get_class_details(class_name)
        if not details.get('_missing_sheet'):
            row = get_class_row(self.training_year, class_name, db_path=self.db_path)
            if row and row['is_staff_meeting'] is not None:
                return bool(row['is_staff_meeting'])
        return 'SM' in str(class_name).upper()

    def get_available_dates_with_options(self, class_name):
        """Dates paired with their LIVE/Virtual choice, for staff meetings."""
        if not self.has_class_data(class_name):
            return []

        date_options = []
        is_meeting = self.is_staff_meeting(class_name)
        details = self.get_class_details(class_name)
        for index in range(1, details.get('date_count', 0) + 1):
            date_str = details.get(f'date_{index}')
            if not date_str:
                continue
            if is_meeting:
                if details.get(f'date_{index}_has_live'):
                    date_options.append((date_str, 'LIVE', f"{date_str} (LIVE Option)"))
                    date_options.append((date_str, 'Virtual',
                                         f"{date_str} (Virtual Option)"))
                else:
                    date_options.append((date_str, 'Virtual',
                                         f"{date_str} (Virtual Only)"))
            else:
                date_options.append((date_str, None, date_str))
        return date_options

    # -- assignments ------------------------------------------------------

    def get_assigned_classes(self, staff_name):
        """The classes one staff member is assigned to this training year."""
        try:
            conn = _connect(self.db_path)
        except Exception as e:
            print(f"Error opening the catalog for '{staff_name}': {e}")
            return []
        try:
            rows = conn.execute(
                "SELECT c.class_name FROM training_class_assignments a "
                "JOIN training_classes c ON c.id = a.class_id "
                "WHERE a.staff_name = ? AND c.training_year = ? "
                "ORDER BY c.display_order, c.class_name",
                (str(staff_name).strip(), self.training_year)).fetchall()
            return [row['class_name'] for row in rows]
        except Exception as e:
            print(f"Error getting assigned classes for '{staff_name}': {e}")
            return []
        finally:
            conn.close()

    def get_staff_assigned_to_class(self, class_name):
        """Who is assigned to one class this training year."""
        row = get_class_row(self.training_year, class_name, db_path=self.db_path)
        if not row:
            return []
        return get_assigned_staff(row['id'], db_path=self.db_path)

    # -- educators --------------------------------------------------------

    def needs_educators(self, class_name):
        """True when a class needs anyone to teach it."""
        return self.get_educator_requirement(class_name) > 0

    def get_educator_requirement(self, class_name):
        """How many educators a class needs per day."""
        details = self.get_class_details(class_name)
        return parse_int(details.get('instructors_per_day'), 0) or 0

    # -- staff attributes -------------------------------------------------
    #
    # These already read the staff database — the workbook was only ever their fallback,
    # and now that it isn't read there is nothing to fall back to. A database with no
    # roster imported returns nothing here, which is the same answer the workbook gave
    # for a name it didn't know.

    def get_staff_list(self):
        """Every active staff member."""
        if staffdb is None:
            return []
        try:
            return staffdb.get_staff_names()
        except Exception as e:
            print(f"Error reading the staff list: {e}")
            return []

    def get_staff_role(self, staff_name):
        """A staff member's base role (NURSE, MEDIC, COMMS, CCEMT, ATP, AMT)."""
        if staffdb is None:
            return None
        try:
            return staffdb.get_role(staff_name) or None
        except Exception as e:
            print(f"Error reading the role for '{staff_name}': {e}")
            return None

    def is_management(self, staff_name):
        if staffdb is None:
            return False
        try:
            return staffdb.is_management(staff_name)
        except Exception as e:
            print(f"Error reading MGMT for '{staff_name}': {e}")
            return False

    def is_dual(self, staff_name):
        if staffdb is None:
            return False
        try:
            return staffdb.is_dual(staff_name)
        except Exception as e:
            print(f"Error reading DUAL for '{staff_name}': {e}")
            return False

    def is_educator_authorized(self, staff_name):
        """
        True when a staff member may sign up to teach.

        An unknown name stays permissive, as it was before: showing the signup to
        somebody the roster has not caught up with is recoverable, hiding it silently
        is not.
        """
        if staffdb is None:
            return True
        try:
            return staffdb.is_educator_at(staff_name, default=True)
        except Exception as e:
            print(f"Error reading educator authorization for '{staff_name}': {e}")
            return True


# ---------------------------------------------------------------------------
# Importing a roster workbook
#
# The workbook is no longer read to run the app — this is the one path that opens one,
# behind "Import roster workbook" in Training Admin. It exists so a fiscal year that was
# built in a spreadsheet can be brought in once, and so the FY26/FY27 workbooks already
# on disk became the starting content of the catalog.
# ---------------------------------------------------------------------------

# The class settings, and the cell each one occupied on a class's detail sheet. Fixed
# positions, never labelled in a way the reader could key on — which is why they are
# written down here rather than discovered.
_SETTING_CELLS = {
    'has_ccemt': (2, 6),             # F2
    'is_multi_session': (2, 7),      # G2
    'session_length': (2, 8),        # H2
    'is_count_exempt': (2, 9),       # I2
    'students_per_class': (16, 2),   # B16
    'nurses_medic_separate': (17, 2),
    'classes_per_day': (18, 2),
    'is_two_day_class': (19, 2),
    'time_1_start': (20, 2),
    'time_1_end': (21, 2),
    'time_2_start': (22, 2),
    'time_2_end': (23, 2),
    'time_3_start': (24, 2),
    'time_3_end': (25, 2),
    'time_4_start': (26, 2),
    'time_4_end': (27, 2),
    'instructors_per_day': (28, 2),
}

# Dates lived in rows 1-14 of a detail sheet, with row 15 reserved as the end-of-list
# marker. That ceiling is the reason a class could not have a fifteenth date.
_MAX_IMPORT_DATE_ROWS = 14


def import_workbook(workbook_path, training_year, overwrite=False,
                    db_path=DEFAULT_DB_PATH):
    """
    Read one roster workbook into the catalog for `training_year`.

    Every column on `Class_Enrollment` that isn't one of the fixed staff-attribute
    columns is a class; its detail sheet supplies the dates and settings, and the ticks
    down its column supply the assignments. A class whose sheet is missing still imports
    — as a class with no dates, which is what the app showed for it before — so the gap
    is visible in the editor rather than silently dropped.

    `overwrite=False` skips classes already in the catalog for this year, so a re-import
    can add what a workbook has gained without undoing an admin's edits. `overwrite=True`
    replaces them, dates and assignments included.

    Returns a report: {'imported': [...], 'skipped': [...], 'warnings': [...]}.
    """
    import openpyxl

    report = {'imported': [], 'skipped': [], 'warnings': []}

    if not os.path.exists(workbook_path):
        report['warnings'].append(f"No workbook at {workbook_path}")
        return report

    initialize_catalog_tables(db_path)

    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as e:
        report['warnings'].append(f"Could not open {workbook_path}: {e}")
        return report

    enrollment_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'class_enrollment' in sheet_name.lower().replace(' ', '_'):
            enrollment_sheet = workbook[sheet_name]
            break
    if enrollment_sheet is None:
        enrollment_sheet = workbook[workbook.sheetnames[0]]
        report['warnings'].append(
            f"No Class_Enrollment sheet; read '{enrollment_sheet.title}' instead")

    # Which column holds each class, and which row holds each staff member.
    class_columns = []
    for column_index, column in enumerate(
            enrollment_sheet.iter_cols(min_col=2, max_row=1), start=2):
        header = column[0].value
        if not header:
            continue
        header = str(header).strip()
        if header and header not in NON_CLASS_COLUMNS:
            class_columns.append((column_index, header))

    staff_rows = []
    for row_index, row in enumerate(enrollment_sheet.iter_rows(min_row=2, max_col=1),
                                    start=2):
        name = row[0].value
        if name and str(name).strip().upper() != 'STAFF NAME':
            staff_rows.append((row_index, str(name).strip()))

    existing = set(get_class_names(training_year, db_path=db_path))

    for column_index, class_name in class_columns:
        if class_name in existing and not overwrite:
            report['skipped'].append(class_name)
            continue

        sheet = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == class_name.lower():
                sheet = workbook[sheet_name]
                break

        settings = {}
        dates = []
        if sheet is None:
            report['warnings'].append(
                f"'{class_name}' has no detail sheet — imported with no dates")
        else:
            for key, (row_number, column_number) in _SETTING_CELLS.items():
                settings[key] = sheet.cell(row=row_number, column=column_number).value

            for row_number in range(1, _MAX_IMPORT_DATE_ROWS + 1):
                class_date = parse_date(sheet.cell(row=row_number, column=2).value)
                if not class_date:
                    continue
                location = sheet.cell(row=row_number, column=5).value
                dates.append({
                    'class_date': class_date,
                    'has_live': parse_checkbox(
                        sheet.cell(row=row_number, column=3).value),
                    'can_work_n_prior': parse_checkbox(
                        sheet.cell(row=row_number, column=4).value),
                    # One location per date is all the sheet could hold, so an imported
                    # date arrives with exactly one option. Adding a second is what the
                    # editor is for.
                    'options': [{'location': (str(location).strip()
                                              if location else None)}],
                })
            if not dates:
                report['warnings'].append(f"'{class_name}' has no dates on its sheet")

        # Frozen at import rather than left to be re-derived: the workbook had no field
        # for it and the module inferred it from "SM" in the name. Writing the inference
        # down keeps every imported class behaving exactly as it did, and puts the answer
        # somewhere an admin can correct it.
        settings['is_staff_meeting'] = 'SM' in class_name.upper()

        assigned = [name for row_index, name in staff_rows
                    if parse_checkbox(
                        enrollment_sheet.cell(row=row_index,
                                              column=column_index).value)]

        try:
            save_class(training_year, class_name, settings=settings, dates=dates,
                       assigned_staff=assigned, source='import', db_path=db_path)
            report['imported'].append(class_name)
        except Exception as e:
            report['warnings'].append(f"Could not import '{class_name}': {e}")

    return report
