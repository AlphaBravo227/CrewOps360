# training_modules/roster_export.py
"""
One fiscal year's training rosters, as an Excel workbook.

Everything about a training year is scoped to that year — its classes, who is assigned
to them, who enrolled and where — and this puts the whole of it in one file, for a year
chosen by name rather than whichever the dashboard happens to be showing.

The sheets, in the order they appear:

    Report Info      what this workbook is and when it was pulled
    Class Summary    one row per class: its span, seats, how full it is, educators
    Rosters          one row per enrollment — who, what, when, and at which location
    Educators        one row per educator signup
    Schedule         one row per class date and location, with times and seats
    Assignments      the staff-by-class grid the roster workbook used to hold
    per-class sheets optional, one printable roster each

`Assignments` is deliberately the shape of the old `Class_Enrollment` sheet: an X where
a staff member is assigned to a class. It is what people are used to reading, and it is
the sheet to hand to someone who wants the year on paper.
"""

import re
from datetime import datetime
from io import BytesIO

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from training_modules import class_catalog as catalog

try:
    from modules import staff_database as staffdb
except Exception as e:  # pragma: no cover
    staffdb = None
    print(f"Staff database unavailable to the roster export: {e}")

_eastern_tz = pytz.timezone('America/New_York')

HEADER_FILL = PatternFill('solid', fgColor='9C27B0')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style='thin', color='D0D0D0')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Excel refuses these in a sheet name, and caps the name at 31 characters.
_ILLEGAL_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')
_MAX_SHEET_NAME = 31


def _sheet_name(base, used):
    """A legal, unique sheet name for a class.

    Class names are free text and two of them can differ only past Excel's 31-character
    limit, so the truncated names would collide and the second sheet would be lost. A
    numeric suffix keeps them apart.
    """
    cleaned = _ILLEGAL_SHEET_CHARS.sub('-', str(base)).strip() or 'Class'
    name = cleaned[:_MAX_SHEET_NAME]
    if name.lower() not in used:
        used.add(name.lower())
        return name
    for suffix in range(2, 100):
        tail = f" ({suffix})"
        name = cleaned[:_MAX_SHEET_NAME - len(tail)] + tail
        if name.lower() not in used:
            used.add(name.lower())
            return name
    used.add(cleaned[:_MAX_SHEET_NAME].lower())
    return cleaned[:_MAX_SHEET_NAME]


def _sort_key(date_text):
    """Chronological order for a MM/DD/YYYY string, unparseable ones last."""
    try:
        return (0, datetime.strptime(str(date_text).strip(), '%m/%d/%Y'))
    except (TypeError, ValueError):
        return (1, datetime.max)


def _write_table(sheet, headers, rows, widths=None, start_row=1):
    """A header row and its data, styled the same way on every sheet."""
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    for offset, row in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=False)

    for column, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    # Freeze under the header so a long roster keeps its column names in view.
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
    if rows:
        sheet.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}")
    return start_row + len(rows) + 1


def _role_of(staff_name, roles):
    return roles.get(staff_name, '')


def collect_year(training_year, db_path=catalog.DEFAULT_DB_PATH, unified_db=None):
    """
    Everything the workbook needs for one year, read once.

    Returns a dict of classes (each with its dates, options and assignments),
    enrollments and educator signups, keyed so the sheet builders can look up without
    going back to the database per row.
    """
    classes = []
    for class_name in catalog.get_class_names(training_year, db_path=db_path):
        record = catalog.load_class_for_editing(training_year, class_name,
                                                db_path=db_path)
        if record:
            classes.append(record)

    enrollments, signups = [], []
    if unified_db is not None:
        try:
            enrollments, signups = unified_db.get_year_export_rows(training_year)
        except Exception as e:
            print(f"Could not read {training_year}'s enrollments: {e}")

    roles = {}
    if staffdb is not None:
        try:
            roles = staffdb.get_base_role_mapping(include_inactive=True)
        except Exception as e:
            print(f"Could not read staff roles: {e}")

    return {'year': training_year, 'classes': classes, 'enrollments': enrollments,
            'signups': signups, 'roles': roles}


def _report_info_sheet(workbook, data):
    sheet = workbook.active
    sheet.title = 'Report Info'
    year = data['year']
    rows = [
        ('Report', 'Training rosters'),
        ('Training year', year or 'not set'),
        ('Generated', datetime.now(_eastern_tz).strftime('%m/%d/%Y %I:%M %p %Z')),
        ('Classes', len(data['classes'])),
        ('Enrollments', len(data['enrollments'])),
        ('Educator signups', len(data['signups'])),
        ('Scope', f"Everything in this workbook belongs to {year or 'this year'} "
                  f"and to no other."),
    ]
    for index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=index, column=2, value=str(value))
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 70


def _summary_sheet(workbook, data):
    """One row per class: what it is, when it runs, and how full it is."""
    enrolled_by_class = {}
    for row in data['enrollments']:
        enrolled_by_class.setdefault(row['class_name'], []).append(row)
    signups_by_class = {}
    for row in data['signups']:
        signups_by_class.setdefault(row['class_name'], []).append(row)

    rows = []
    for record in data['classes']:
        name = record['class_name']
        settings = record['settings']
        dates = record['dates']
        ordered = sorted((entry['class_date'] for entry in dates), key=_sort_key)
        locations = sorted({option['location'] for entry in dates
                            for option in entry['options'] if option['location']})

        default_capacity = catalog.parse_int(settings.get('students_per_class'), 0) or 0
        # A date's seats are the sum of its locations' seats, each falling back to the
        # class capacity. A day at two sites genuinely holds two rooms of people.
        seats = 0
        for entry in dates:
            for option in entry['options']:
                seats += (catalog.parse_int(option.get('capacity'))
                          or default_capacity)

        enrolled = len(enrolled_by_class.get(name, []))
        instructors = catalog.parse_int(settings.get('instructors_per_day'), 0) or 0
        rows.append([
            name,
            record['source'] == 'import' and 'Imported' or 'Built in app',
            len(dates),
            ordered[0] if ordered else '',
            ordered[-1] if ordered else '',
            ', '.join(locations),
            default_capacity,
            seats,
            enrolled,
            max(seats - enrolled, 0),
            f"{(enrolled / seats * 100):.0f}%" if seats else '',
            len(record['assigned_staff']),
            instructors * len(dates),
            len(signups_by_class.get(name, [])),
        ])

    sheet = workbook.create_sheet('Class Summary')
    _write_table(
        sheet,
        ['Class', 'Origin', 'Dates', 'First date', 'Last date', 'Locations',
         'Seats per session', 'Seats offered', 'Enrolled', 'Seats left', 'Full',
         'Assigned staff', 'Educators needed', 'Educators signed up'],
        rows,
        widths=[30, 14, 7, 12, 12, 24, 10, 10, 10, 10, 8, 10, 10, 10])


def _rosters_sheet(workbook, data):
    """Every enrollment in the year, one row each."""
    roles = data['roles']
    rows = []
    for row in sorted(data['enrollments'],
                      key=lambda r: (str(r['class_name']).lower(),
                                     _sort_key(r['class_date']),
                                     str(r.get('location') or ''),
                                     str(r['staff_name']).lower())):
        rows.append([
            row['class_name'],
            row['class_date'],
            row.get('location') or '',
            row['staff_name'],
            _role_of(row['staff_name'], roles),
            row.get('role') or '',
            row.get('session_time') or '',
            row.get('meeting_type') or '',
            'Yes' if row.get('conflict_override') else '',
            row.get('enrollment_date') or '',
        ])

    sheet = workbook.create_sheet('Rosters')
    _write_table(
        sheet,
        ['Class', 'Date', 'Location', 'Staff Name', 'Staff Role', 'Enrolled As',
         'Session', 'Meeting Type', 'Conflict Override', 'Enrolled On'],
        rows,
        widths=[28, 12, 14, 24, 12, 12, 16, 13, 15, 22])


def _educators_sheet(workbook, data):
    roles = data['roles']
    rows = []
    for row in sorted(data['signups'],
                      key=lambda r: (str(r['class_name']).lower(),
                                     _sort_key(r['class_date']),
                                     str(r['staff_name']).lower())):
        rows.append([
            row['class_name'],
            row['class_date'],
            row['staff_name'],
            _role_of(row['staff_name'], roles),
            'Yes' if row.get('conflict_override') else '',
            row.get('signup_date') or '',
        ])

    sheet = workbook.create_sheet('Educators')
    _write_table(
        sheet,
        ['Class', 'Date', 'Educator', 'Staff Role', 'Conflict Override', 'Signed Up On'],
        rows,
        widths=[28, 12, 24, 12, 15, 22])


def _schedule_sheet(workbook, data):
    """Every class date, and every location it runs at, with its own times and seats."""
    enrolled_by_slot = {}
    for row in data['enrollments']:
        key = (row['class_name'], row['class_date'], row.get('location') or '')
        enrolled_by_slot[key] = enrolled_by_slot.get(key, 0) + 1

    rows = []
    for record in data['classes']:
        name = record['class_name']
        settings = record['settings']
        default_capacity = catalog.parse_int(settings.get('students_per_class'), 0) or 0
        default_start = settings.get('time_1_start') or ''
        default_end = settings.get('time_1_end') or ''
        for entry in sorted(record['dates'], key=lambda e: _sort_key(e['class_date'])):
            multi = len(entry['options']) > 1
            for option in entry['options']:
                location = option['location'] or ''
                capacity = catalog.parse_int(option.get('capacity')) or default_capacity
                # A single-location date's enrollments predate the location column and
                # may have none recorded, so they are counted under the blank key.
                enrolled = enrolled_by_slot.get((name, entry['class_date'], location), 0)
                if not multi and not enrolled:
                    enrolled = enrolled_by_slot.get((name, entry['class_date'], ''), 0)
                rows.append([
                    name,
                    entry['class_date'],
                    location,
                    'Yes' if multi else '',
                    option.get('start_time') or default_start,
                    option.get('end_time') or default_end,
                    capacity,
                    enrolled,
                    max(capacity - enrolled, 0),
                    'Yes' if entry['has_live'] else '',
                    'Yes' if entry['can_work_n_prior'] else '',
                ])

    sheet = workbook.create_sheet('Schedule')
    _write_table(
        sheet,
        ['Class', 'Date', 'Location', 'Multiple Locations', 'Start', 'End',
         'Seats', 'Enrolled', 'Seats Left', 'LIVE Option', 'Can Work Night Before'],
        rows,
        widths=[28, 12, 14, 12, 8, 8, 8, 9, 10, 11, 13])


def _assignments_sheet(workbook, data):
    """The staff-by-class grid, in the shape the roster workbook's own sheet had."""
    class_names = [record['class_name'] for record in data['classes']]
    assigned = {record['class_name']: set(record['assigned_staff'])
                for record in data['classes']}

    everyone = sorted({name for names in assigned.values() for name in names})
    roles = data['roles']

    rows = []
    for name in everyone:
        row = [name, _role_of(name, roles)]
        row.extend('X' if name in assigned[class_name] else ''
                   for class_name in class_names)
        row.append(sum(1 for class_name in class_names if name in assigned[class_name]))
        rows.append(row)

    sheet = workbook.create_sheet('Assignments')
    _write_table(
        sheet,
        ['STAFF NAME', 'Role'] + class_names + ['Total'],
        rows,
        widths=[24, 12] + [16] * len(class_names) + [8])

    # The class columns read as a wall of X without their names on their side; rotating
    # the headers keeps a year of classes on one screen.
    for column in range(3, 3 + len(class_names)):
        sheet.cell(row=1, column=column).alignment = Alignment(
            textRotation=90, horizontal='center', vertical='bottom', wrap_text=False)
    sheet.row_dimensions[1].height = 120
    sheet.freeze_panes = 'C2'


def _class_sheets(workbook, data):
    """One printable roster per class: its schedule, then who is on it."""
    used = {sheet.title.lower() for sheet in workbook.worksheets}
    roles = data['roles']

    enrolled_by_class = {}
    for row in data['enrollments']:
        enrolled_by_class.setdefault(row['class_name'], []).append(row)
    signups_by_class = {}
    for row in data['signups']:
        signups_by_class.setdefault(row['class_name'], []).append(row)

    for record in data['classes']:
        name = record['class_name']
        settings = record['settings']
        sheet = workbook.create_sheet(_sheet_name(name, used))

        sheet.cell(row=1, column=1, value=name).font = TITLE_FONT
        sheet.cell(row=2, column=1,
                   value=f"{data['year']}  ·  {len(record['dates'])} date(s)  ·  "
                         f"{len(record['assigned_staff'])} assigned  ·  "
                         f"{len(enrolled_by_class.get(name, []))} enrolled").font = (
                             Font(italic=True))

        default_capacity = catalog.parse_int(settings.get('students_per_class'), 0) or 0
        schedule_rows = []
        for entry in sorted(record['dates'], key=lambda e: _sort_key(e['class_date'])):
            for option in entry['options']:
                schedule_rows.append([
                    entry['class_date'],
                    option['location'] or '',
                    option.get('start_time') or settings.get('time_1_start') or '',
                    option.get('end_time') or settings.get('time_1_end') or '',
                    catalog.parse_int(option.get('capacity')) or default_capacity,
                ])
        next_row = _write_table(sheet, ['Date', 'Location', 'Start', 'End', 'Seats'],
                                schedule_rows, widths=[12, 16, 8, 8, 8], start_row=4)

        next_row += 1
        sheet.cell(row=next_row, column=1, value='Enrolled').font = Font(bold=True)
        roster_rows = [
            [row['class_date'], row.get('location') or '', row['staff_name'],
             _role_of(row['staff_name'], roles), row.get('session_time') or '',
             row.get('meeting_type') or '']
            for row in sorted(enrolled_by_class.get(name, []),
                              key=lambda r: (_sort_key(r['class_date']),
                                             str(r.get('location') or ''),
                                             str(r['staff_name']).lower()))
        ]
        next_row = _write_table(
            sheet, ['Date', 'Location', 'Staff Name', 'Role', 'Session', 'Meeting Type'],
            roster_rows, start_row=next_row + 1)

        signups = signups_by_class.get(name, [])
        if signups:
            next_row += 1
            sheet.cell(row=next_row, column=1, value='Educators').font = Font(bold=True)
            _write_table(
                sheet, ['Date', 'Educator', 'Role'],
                [[row['class_date'], row['staff_name'],
                  _role_of(row['staff_name'], roles)]
                 for row in sorted(signups, key=lambda r: (_sort_key(r['class_date']),
                                                           str(r['staff_name']).lower()))],
                start_row=next_row + 1)

        # Each block sets its own freeze pane as it is written; the last one wins and
        # would freeze the middle of the sheet. A per-class roster is short — leave it.
        sheet.freeze_panes = None


def build_roster_workbook(training_year, db_path=catalog.DEFAULT_DB_PATH,
                          unified_db=None, per_class_sheets=False):
    """
    One training year's rosters as .xlsx bytes, ready to hand to a download button.

    `per_class_sheets` adds a printable sheet per class on top of the year-wide ones.
    Off by default: a year with twenty-odd classes gets a workbook with twenty-odd
    extra tabs, which is the right thing when you are printing rosters and noise when
    you are not.
    """
    data = collect_year(training_year, db_path=db_path, unified_db=unified_db)

    workbook = Workbook()
    _report_info_sheet(workbook, data)
    _summary_sheet(workbook, data)
    _rosters_sheet(workbook, data)
    _educators_sheet(workbook, data)
    _schedule_sheet(workbook, data)
    _assignments_sheet(workbook, data)
    if per_class_sheets:
        _class_sheets(workbook, data)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
