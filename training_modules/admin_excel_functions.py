# training_modules/admin_excel_functions.py - Enhanced with comprehensive schedule report
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import streamlit as st
import calendar
from .config import NON_CLASS_COLUMNS, DEFAULT_CLASS_DETAILS

class ExcelAdminFunctions:
    def __init__(self, excel_handler, enrollment_manager, database, educator_manager=None):
        self.excel = excel_handler
        self.enrollment = enrollment_manager
        self.db = database
        self.educator = educator_manager
        
    def get_enrollment_compliance_report(self):
        """Generate compliance report showing enrollment status vs assignments - FIXED for cursor recursion"""
        staff_list = self.excel.get_staff_list()
        
        # NEW: Get manager information from direct_reports table
        manager_dict = {}
        try:
            self.db.connect()
            self.db.cursor.execute('SELECT staff_name, manager_initials FROM direct_reports')
            for row in self.db.cursor.fetchall():
                manager_dict[row['staff_name']] = row['manager_initials'] if row['manager_initials'] else ''
            self.db.disconnect()
        except Exception as e:
            print(f"Error loading manager data: {e}")
            if hasattr(self.db, 'disconnect'):
                self.db.disconnect()
        
        report_data = []
        
        for staff_name in staff_list:
            # Get assigned classes
            assigned_classes = self.excel.get_assigned_classes(staff_name)
            
            # Get enrolled classes - this returns a list of class names
            enrolled_classes = self.enrollment.get_enrolled_classes(staff_name)
            
            # Get ALL enrollment records (with full details) from database
            staff_enrollments = self.db.get_staff_enrollments(staff_name)
            
            # Get live meeting count - use a separate method call
            live_meeting_count = self.enrollment.get_live_staff_meeting_count(staff_name)
            
            # Get educator signups if available - call once and store
            educator_signups = 0
            if self.educator:
                try:
                    educator_signup_list = self.educator.get_staff_educator_signups(staff_name)
                    educator_signups = len(educator_signup_list)
                except Exception as e:
                    print(f"Error getting educator signups for {staff_name}: {e}")
                    educator_signups = 0
            
            # ===== UPDATED SECTION: Calculate completion metrics =====
            total_assigned = len(assigned_classes)
            
            # Count total enrollments including multiple SM enrollments
            # For staff meetings, count each enrollment separately
            # For regular classes, count unique enrollments only
            total_enrolled = 0
            for cls in assigned_classes:
                if self.excel.is_staff_meeting(cls):
                    # Count all SM enrollments for this class (can be multiple)
                    sm_enrollments = [e for e in staff_enrollments 
                                    if e['class_name'] == cls]
                    total_enrolled += len(sm_enrollments)
                else:
                    # For regular classes, count as 1 if enrolled
                    if cls in enrolled_classes:
                        total_enrolled += 1
            
            completion_rate = (total_enrolled / total_assigned * 100) if total_assigned > 0 else 0
            # ===== END OF UPDATED SECTION =====
            
            # Check staff meeting requirement
            staff_meetings_assigned = [cls for cls in assigned_classes 
                                    if self.excel.is_staff_meeting(cls)]
            staff_meeting_compliance = live_meeting_count >= 2 if staff_meetings_assigned else True

            # Calculate total staff meeting enrollments (LIVE + Virtual)
            # Count all SM enrollments including multiple enrollments in the same SM
            staff_enrollments = self.db.get_staff_enrollments(staff_name)
            total_sm_enrolled = len([enrollment for enrollment in staff_enrollments 
                                    if self.excel.is_staff_meeting(enrollment['class_name'])])
            total_sm_assigned = len(staff_meetings_assigned)

            # Format Total SM display
            total_sm_display = f"{total_sm_enrolled}/{total_sm_assigned}" if staff_meetings_assigned else ''

            # Determine Total SM Compliance
            total_sm_compliance = total_sm_enrolled >= total_sm_assigned if staff_meetings_assigned else True
            total_sm_compliance_display = "✅" if total_sm_compliance and staff_meetings_assigned else ("❌" if staff_meetings_assigned else '')            
            
            # Get conflict overrides - fetch all at once
            conflict_enrollments = []
            educator_conflicts = []
            try:
                conflict_enrollments = self.db.get_conflict_override_enrollments(staff_name)
            except Exception as e:
                print(f"Error getting conflict enrollments for {staff_name}: {e}")
            
            if self.educator:
                try:
                    educator_conflicts = self.db.get_conflict_override_educator_signups(staff_name)
                except Exception as e:
                    print(f"Error getting educator conflicts for {staff_name}: {e}")

            # Determine which classes are not enrolled
            classes_not_enrolled = [cls for cls in assigned_classes if cls not in enrolled_classes]

            # ===== NEW: Filter out SM classes if staff member is fully SM compliant =====
            if total_sm_compliance and staff_meeting_compliance:
                # Both SM requirements met - remove all SM classes from "not enrolled" list
                classes_not_enrolled = [cls for cls in classes_not_enrolled 
                                    if not self.excel.is_staff_meeting(cls)]

            classes_not_enrolled_str = ', '.join(classes_not_enrolled) if classes_not_enrolled else ''
            
            # ===== NEW: Filter classes starting within 90 days (or already started) =====
            classes_starting_soon = []
            current_date = datetime.now()
            ninety_days_out = current_date + timedelta(days=90)

            for cls in classes_not_enrolled:
                class_details = self.excel.get_class_details(cls)
                if class_details:
                    # Find the earliest date for this class (first date)
                    earliest_date = None
                    for i in range(1, 15):  # Check rows 1-14 for dates
                        date_key = f'date_{i}'
                        if date_key in class_details and class_details[date_key]:
                            try:
                                date_str = class_details[date_key]
                                date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                                
                                # For two-day classes, use Day 1 as the date
                                if earliest_date is None or date_obj < earliest_date:
                                    earliest_date = date_obj
                            except Exception as e:
                                print(f"Error parsing date for {cls}: {e}")
                                continue
                    
                    # Show if: first date is in the past OR first date is within 90 days
                    # Don't show if: first date is more than 90 days in the future
                    if earliest_date:
                        if earliest_date < current_date or earliest_date <= ninety_days_out:
                            classes_starting_soon.append(cls)

            classes_starting_soon_str = ', '.join(classes_starting_soon) if classes_starting_soon else ''

            report_data.append({
                'Staff Name': staff_name,
                'Manager': manager_dict.get(staff_name, 'N/A'),
                'Total Assigned': total_assigned,
                'Total Enrolled': total_enrolled,
                'Completion Rate': completion_rate / 100,
                'Classes Remaining': total_assigned - total_enrolled,
                'Unenrolled Classes Starting within 90 Days': classes_starting_soon_str,
                'All Classes Not Enrolled': classes_not_enrolled_str,
                'Total SM': total_sm_display,
                'Total SM Compliance': total_sm_compliance_display,
                'LIVE Meetings': f"{live_meeting_count}/2" if staff_meetings_assigned else "N/A",
                'Meeting Compliance': "✅" if staff_meeting_compliance else "❌",
                'LIVE SM Meetings': f"{live_meeting_count}/2" if staff_meetings_assigned else "N/A",
                'LIVE SM Meeting Compliance': "✅" if staff_meeting_compliance else "❌",
                'Conflict Overrides': len(conflict_enrollments),
                'Educator Signups': educator_signups,
                'Educator Conflicts': len(educator_conflicts),
                'Status': self._get_compliance_status(completion_rate, staff_meeting_compliance)
            })
        
        # Return sorted by Staff Name only
        return pd.DataFrame(report_data)


    def get_educator_coverage_report(self):
        """Generate report showing educator coverage for classes that need educators"""
        if not self.educator:
            return pd.DataFrame()  # Return empty DataFrame if educator manager not available
        
        opportunities = self.educator.get_educator_opportunities()
        coverage_data = []
        
        for opportunity in opportunities:
            class_name = opportunity['class_name']
            instructor_requirement = opportunity['instructor_count']
            available_dates = opportunity['available_dates']
            
            for date in available_dates:
                current_signups = self.db.get_educator_signup_count(class_name, date)
                coverage_rate = (current_signups / instructor_requirement * 100) if instructor_requirement > 0 else 0
                
                # Get educator names
                educator_roster = self.educator.get_class_educator_roster(class_name, date)
                educator_names = [e['staff_name'] for e in educator_roster if e['status'] == 'active']
                educator_conflicts = sum(1 for e in educator_roster if e.get('has_conflict', False))
                
                coverage_data.append({
                    'Class Name': class_name,
                    'Date': date,
                    'Required Educators': instructor_requirement,
                    'Current Signups': current_signups,
                    'Coverage Rate': f"{coverage_rate:.1f}%",
                    'Still Needed': max(0, instructor_requirement - current_signups),
                    'Educator Conflicts': educator_conflicts,
                    'Educators': ', '.join(educator_names) if educator_names else 'None',
                    'Status': self._get_coverage_status(coverage_rate, instructor_requirement, current_signups)
                })
        
        return pd.DataFrame(coverage_data)
    
    def get_educator_participation_report(self):
        """Generate report showing individual educator participation"""
        if not self.educator:
            return pd.DataFrame()
        
        all_staff = self.excel.get_staff_list()
        participation_data = []
        
        for staff_name in all_staff:
            educator_signups = self.educator.get_staff_educator_signups(staff_name)
            
            if educator_signups:  # Only include staff who have signed up as educators
                total_signups = len(educator_signups)
                conflict_overrides = sum(1 for signup in educator_signups 
                                       if signup.get('conflict_override', False))
                
                # Get unique classes
                unique_classes = set(signup['class_name'] for signup in educator_signups)
                
                # Get recent signups (last 30 days)
                thirty_days_ago = datetime.now().replace(day=1)  # Simplified for demo
                recent_signups = sum(1 for signup in educator_signups 
                                   if signup.get('signup_date_display'))  # Simplified check
                
                participation_data.append({
                    'Staff Name': staff_name,
                    'Total Educator Signups': total_signups,
                    'Unique Classes': len(unique_classes),
                    'Conflict Overrides': conflict_overrides,
                    'Recent Signups (30d)': recent_signups,
                    'Classes': ', '.join(sorted(unique_classes)),
                    'Participation Level': self._get_participation_level(total_signups)
                })
        
        return pd.DataFrame(participation_data)
    
    def get_classes_needing_educators_report(self):
        """Generate report showing classes that still need educator signups"""
        if not self.educator:
            return pd.DataFrame()
        
        needs_educators = self.educator.get_classes_needing_educators()
        
        if not needs_educators:
            return pd.DataFrame(columns=['Class Name', 'Date', 'Still Needed', 'Current', 'Required', 'Urgency'])
        
        report_data = []
        for need in needs_educators:
            urgency_level = self._get_urgency_level(need['current'], need['required'])
            
            report_data.append({
                'Class Name': need['class_name'],
                'Date': need['class_date'],
                'Still Needed': need['needed'],
                'Current': need['current'],
                'Required': need['required'],
                'Coverage %': f"{(need['current'] / need['required'] * 100):.1f}%" if need['required'] > 0 else "0%",
                'Urgency': urgency_level
            })
        
        # Sort by urgency and then by needed count
        df = pd.DataFrame(report_data)
        if not df.empty:
            urgency_order = {'🔴 Critical': 0, '🟡 Moderate': 1, '🟢 Low': 2}
            df['urgency_sort'] = df['Urgency'].map(urgency_order)
            df = df.sort_values(['urgency_sort', 'Still Needed'], ascending=[True, False])
            df = df.drop('urgency_sort', axis=1)
        
        return df
    
    def get_comprehensive_education_schedule_report(self, start_date, end_date):
        """Generate comprehensive education schedule report for a date range - UPDATED to include all staff with roles"""
        try:
            # Parse date strings to datetime objects for comparison
            start_dt = datetime.strptime(start_date, '%Y-%m-%d') if isinstance(start_date, str) else start_date
            end_dt = datetime.strptime(end_date, '%Y-%m-%d') if isinstance(end_date, str) else end_date
            
            # Generate date range with MM/DD/YY format
            date_range = []
            current_date = start_dt
            while current_date <= end_dt:
                date_range.append(current_date.strftime('%m/%d/%y'))
                current_date += timedelta(days=1)
            
            # Get all staff with their roles - this now includes ALL staff from Excel roster
            all_staff_with_roles = self._get_all_staff_from_database()
            
            # Initialize report data
            report_data = []
            
            for staff_name, role in all_staff_with_roles:  # Now we have both name and role
                # Initialize row data with staff name and role as separate columns
                row_data = {
                    'STAFF NAME': staff_name,
                    'ROLE': role,
                }
                
                # Get all enrollments for this staff member
                staff_enrollments = self.db.get_staff_enrollments(staff_name)
                
                # Get all educator signups for this staff member
                staff_educator_signups = []
                if self.educator:
                    staff_educator_signups = self.educator.get_staff_educator_signups(staff_name)
                
                # Process each date in the range
                for date_str in date_range:
                    activities = []
                    
                    # Convert MM/DD/YY back to MM/DD/YYYY for database comparison
                    date_parts = date_str.split('/')
                    full_date_str = f"{date_parts[0]}/{date_parts[1]}/20{date_parts[2]}"
                    
                    # Check for student enrollments on this date
                    enrollments_on_date = [e for e in staff_enrollments if e['class_date'] == full_date_str]
                    for enrollment in enrollments_on_date:
                        activities.append(enrollment['class_name'])
                    
                    # Check for educator signups on this date
                    educator_signups_on_date = [e for e in staff_educator_signups if e['class_date'] == full_date_str]
                    for signup in educator_signups_on_date:
                        activities.append(f"EDU:{signup['class_name']}")
                    
                    # Join activities with comma if multiple, otherwise leave blank
                    row_data[date_str] = ', '.join(activities) if activities else ''
                
                report_data.append(row_data)
            
            # Create DataFrame
            columns = ['STAFF NAME', 'ROLE'] + date_range
            df = pd.DataFrame(report_data, columns=columns)
            
            # Sort by staff name first, then by role
            df = df.sort_values(['STAFF NAME', 'ROLE'])
            
            return df
            
        except Exception as e:
            print(f"Error generating comprehensive schedule report: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def export_comprehensive_schedule_to_excel(self, schedule_df, start_date, end_date):
        """Export comprehensive schedule report to Excel format as a Table"""
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            if schedule_df.empty:
                return None
            
            # Create a new workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Education Schedule'
            
            # Write title and date range
            worksheet.merge_cells('A1:E1')
            title_cell = worksheet['A1']
            title_cell.value = f'Comprehensive Education Schedule Report'
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            worksheet.merge_cells('A2:E2')
            subtitle_cell = worksheet['A2']
            subtitle_cell.value = f'Date Range: {start_date} to {end_date}'
            subtitle_cell.font = Font(italic=True)
            subtitle_cell.alignment = Alignment(horizontal='center')
            
            # Start data at row 4
            start_row = 4
            
            # Write headers
            for col_num, column_name in enumerate(schedule_df.columns, 1):
                cell = worksheet.cell(row=start_row, column=col_num, value=column_name)
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 25  # Staff Name
            worksheet.column_dimensions['B'].width = 12  # Role column
            
            # Date columns
            date_columns = len(schedule_df.columns) - 2  # Subtract STAFF NAME and ROLE columns
            if date_columns > 0:
                for col_num in range(3, 3 + date_columns):  # Start from column C (3rd column)
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    worksheet.column_dimensions[col_letter].width = 15
            
            # Write data
            for row_num, (index, row) in enumerate(schedule_df.iterrows(), start_row + 1):
                for col_num, (column_name, value) in enumerate(row.items(), 1):
                    cell_value = str(value) if pd.notna(value) and value != '' else ''
                    cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
            
            # Create Excel Table
            last_row = start_row + len(schedule_df)
            last_col = len(schedule_df.columns)
            last_col_letter = openpyxl.utils.get_column_letter(last_col)
            
            table_ref = f"A{start_row}:{last_col_letter}{last_row}"
            table = Table(displayName="ScheduleTable", ref=table_ref)
            
            # Apply default blue table style
            style = TableStyleInfo(
                name="TableStyleMedium9",  # Default blue table style
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            worksheet.add_table(table)
            
            # Save to BytesIO buffer
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            print(f"Error exporting comprehensive schedule to Excel: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _generate_class_details_comment(self, activities, date_str, staff_name):
        """Generate comment text with class details for activities - IMPROVED formatting"""
        try:
            comment_lines = []
            
            # Convert MM/DD/YY back to MM/DD/YYYY for class detail lookup
            date_parts = date_str.split('/')
            if len(date_parts) == 3:
                full_date_str = f"{date_parts[0]}/{date_parts[1]}/20{date_parts[2]}"
            else:
                full_date_str = date_str
            
            for activity in activities:
                activity = activity.strip()
                if not activity:
                    continue
                
                # Extract class name (remove EDU: prefix if present)
                if activity.startswith('EDU:'):
                    class_name = activity[4:].strip()
                    activity_type = "Educator"
                    is_educator = True
                else:
                    class_name = activity
                    activity_type = "Student"
                    is_educator = False
                
                # Get class details
                class_details = self.excel.get_class_details(class_name)
                
                if class_details:
                    # Get specific enrollment details for this staff member
                    if is_educator:
                        # Get educator signup details
                        time_info, location_info = self._get_specific_educator_details(
                            staff_name, class_name, full_date_str, class_details
                        )
                    else:
                        # Get student enrollment details
                        time_info, location_info = self._get_specific_enrollment_details(
                            staff_name, class_name, full_date_str, class_details
                        )
                    
                    # Build comment for this activity with improved formatting
                    if len(activities) > 1:
                        # Multiple activities - include class name and type header
                        activity_comment = f"=== {class_name} ({activity_type}) ===\n"
                        activity_comment += f"Time: {time_info}\n"
                        activity_comment += f"Location: {location_info}"
                    else:
                        # Single activity - more concise format
                        activity_comment = f"{class_name} - {activity_type}\n"
                        activity_comment += f"Time: {time_info}\n"
                        activity_comment += f"Location: {location_info}"
                    
                    comment_lines.append(activity_comment)
                else:
                    # Fallback if class details not found
                    if len(activities) > 1:
                        activity_comment = f"=== {class_name} ({activity_type}) ===\n"
                        activity_comment += "Time: Not specified\n"
                        activity_comment += "Location: Not specified"
                    else:
                        activity_comment = f"{class_name} - {activity_type}\n"
                        activity_comment += "Time: Not specified\n"
                        activity_comment += "Location: Not specified"
                    
                    comment_lines.append(activity_comment)
            
            # Join all activity comments with separator for multiple activities
            if len(comment_lines) > 1:
                return '\n\n'.join(comment_lines)
            elif comment_lines:
                return comment_lines[0]
            else:
                return None
                
        except Exception as e:
            print(f"Error generating class details comment: {e}")
            return f"Class: {', '.join(activities)}\nDetails unavailable"

    def _get_specific_enrollment_details(self, staff_name, class_name, class_date, class_details):
        """Get time and location details for a specific staff enrollment"""
        try:
            # Get the actual enrollment record for this staff member
            enrollments = self.db.get_staff_enrollments(staff_name)
            matching_enrollment = None
            
            for enrollment in enrollments:
                if (enrollment['class_name'] == class_name and 
                    enrollment['class_date'] == class_date):
                    matching_enrollment = enrollment
                    break
            
            if matching_enrollment:
                # Get specific session time if enrolled in a specific session
                session_time = matching_enrollment.get('session_time')
                
                if session_time:
                    # Use the specific session time
                    time_info = session_time
                else:
                    # Fall back to general class time (first session only)
                    time_info = self._get_general_class_time(class_details)
                
                # Get location for this date
                location_info = self._get_class_location_for_date(class_details, class_date)
                
                return time_info, location_info
            else:
                # Fallback if enrollment not found
                time_info = self._get_general_class_time(class_details)
                location_info = self._get_class_location_for_date(class_details, class_date)
                return time_info, location_info
                
        except Exception as e:
            print(f"Error getting specific enrollment details: {e}")
            return "Not specified", "Not specified"

    def _get_specific_educator_details(self, staff_name, class_name, class_date, class_details):
        """Get time and location details for a specific educator signup"""
        try:
            # Get the actual educator signup record for this staff member
            if self.educator:
                educator_signups = self.educator.get_staff_educator_signups(staff_name)
                matching_signup = None
                
                for signup in educator_signups:
                    if (signup['class_name'] == class_name and 
                        signup['class_date'] == class_date):
                        matching_signup = signup
                        break
                
                if matching_signup:
                    # For educators, we typically use the general class time
                    time_info = self._get_general_class_time(class_details)
                    location_info = self._get_class_location_for_date(class_details, class_date)
                    return time_info, location_info
            
            # Fallback if educator manager not available or signup not found
            time_info = self._get_general_class_time(class_details)
            location_info = self._get_class_location_for_date(class_details, class_date)
            return time_info, location_info
            
        except Exception as e:
            print(f"Error getting specific educator details: {e}")
            return "Not specified", "Not specified"

    def _get_general_class_time(self, class_details):
        """Extract general time information from class details (first session only)"""
        try:
            # Get the first time slot only (not all sessions)
            start_time = class_details.get('time_1_start')
            end_time = class_details.get('time_1_end')
            
            if start_time and end_time:
                time_info = f"{start_time} - {end_time}"
            elif start_time:
                time_info = f"Starts: {start_time}"
            elif end_time:
                time_info = f"Ends: {end_time}"
            else:
                time_info = "Not specified"
            
            return time_info
            
        except Exception as e:
            print(f"Error getting class time: {e}")
            return "Not specified"

    def _get_class_location_for_date(self, class_details, target_date):
        """Get location information for a specific date"""
        try:
            # Search through date rows to find matching date and its location
            for i in range(1, 15):  # Check rows 1-14 for dates
                date_key = f'date_{i}'
                location_key = f'date_{i}_location'
                
                if date_key in class_details and class_details[date_key]:
                    stored_date = class_details[date_key]
                    
                    # Compare dates (handle different formats)
                    if stored_date == target_date:
                        location = class_details.get(location_key, '')
                        return location.strip() if location else "Not specified"
            
            # If no specific location found for the date
            return "Not specified"
            
        except Exception as e:
            print(f"Error getting class location: {e}")
            return "Not specified"

    def _get_all_staff_from_database(self):
        """Get all staff names with their roles - includes ALL staff from Excel roster plus any database-only staff"""
        try:
            all_staff_with_roles = {}  # Dictionary to store staff_name: role
            
            # FIRST: Get all staff from Excel roster (this ensures we include everyone)
            excel_staff = self.excel.get_staff_list()
            print(f"Found {len(excel_staff)} staff in Excel roster")
            
            for staff_name in excel_staff:
                if staff_name:
                    # Get role from Excel roster
                    role = self._get_staff_role_from_excel(staff_name)
                    all_staff_with_roles[staff_name] = role
                    print(f"DEBUG: Added from Excel: {staff_name} - {role}")
            
            # SECOND: Get any additional staff from database who might not be in Excel roster
            self.db.connect()
            
            # Get staff from training enrollments
            self.db.cursor.execute('''
                SELECT DISTINCT staff_name FROM training_enrollments 
                WHERE status = 'active'
            ''')
            enrollment_staff = self.db.cursor.fetchall()
            
            for row in enrollment_staff:
                staff_name = row['staff_name']
                if staff_name and staff_name not in all_staff_with_roles:
                    # This staff member is in database but not in Excel roster
                    role = self._get_staff_role_from_excel(staff_name) or "Unknown"
                    all_staff_with_roles[staff_name] = role
                    print(f"DEBUG: Added from enrollments DB: {staff_name} - {role}")
            
            # Get staff from educator signups
            if self.educator:
                self.db.cursor.execute('''
                    SELECT DISTINCT staff_name FROM training_educator_signups 
                    WHERE status = 'active'
                ''')
                educator_staff = self.db.cursor.fetchall()
                
                for row in educator_staff:
                    staff_name = row['staff_name']
                    if staff_name and staff_name not in all_staff_with_roles:
                        # This staff member is in database but not in Excel roster
                        role = self._get_staff_role_from_excel(staff_name) or "Unknown"
                        all_staff_with_roles[staff_name] = role
                        print(f"DEBUG: Added from educator DB: {staff_name} - {role}")
            
            self.db.disconnect()
            
            # Convert to list of tuples (name, role) and sort by name
            staff_with_roles = [(name, role) for name, role in all_staff_with_roles.items()]
            staff_with_roles.sort(key=lambda x: x[0])  # Sort by name
            
            print(f"Final count: {len(staff_with_roles)} staff members with roles")
            return staff_with_roles
            
        except Exception as e:
            print(f"Error getting staff from database: {e}")
            import traceback
            traceback.print_exc()
            if hasattr(self.db, 'disconnect'):
                self.db.disconnect()
            return []

    def _get_staff_role_from_excel(self, staff_name):
        """Get staff member's role from Excel enrollment sheet"""
        try:
            if not self.excel.enrollment_sheet:
                return "Unknown"
            
            # Find the staff member's row
            staff_row = None
            for row_idx, row in enumerate(self.excel.enrollment_sheet.iter_rows(min_row=2, max_col=1), start=2):
                if row[0].value and str(row[0].value).strip() == staff_name:
                    staff_row = row_idx
                    break
            
            if not staff_row:
                return "Unknown"
            
            # Find the Role column
            role_col = None
            for col_idx, col in enumerate(self.excel.enrollment_sheet.iter_cols(min_row=1, max_row=1), start=1):
                header_value = str(col[0].value).strip() if col[0].value else ""
                if header_value == "Role":
                    role_col = col_idx
                    break
            
            if not role_col:
                return "Unknown"
            
            # Get role value
            role_cell = self.excel.enrollment_sheet.cell(row=staff_row, column=role_col)
            role_value = str(role_cell.value).strip() if role_cell.value else "Unknown"
            
            return role_value
            
        except Exception as e:
            print(f"Error getting role for {staff_name}: {e}")
            return "Unknown"

    def _get_class_time_for_date(self, class_details):
        """Extract time information from class details"""
        try:
            # Get the first time slot (most common case)
            start_time = class_details.get('time_1_start')
            end_time = class_details.get('time_1_end')
            
            if start_time and end_time:
                time_info = f"{start_time} - {end_time}"
            elif start_time:
                time_info = f"Starts: {start_time}"
            elif end_time:
                time_info = f"Ends: {end_time}"
            else:
                time_info = "Not specified"
            
            # Check for multiple sessions
            classes_per_day = int(class_details.get('classes_per_day', 1))
            if classes_per_day > 1:
                sessions = []
                for i in range(1, min(classes_per_day + 1, 5)):  # Max 4 sessions
                    start_key = f'time_{i}_start'
                    end_key = f'time_{i}_end'
                    
                    session_start = class_details.get(start_key)
                    session_end = class_details.get(end_key)
                    
                    if session_start and session_end:
                        sessions.append(f"Session {i}: {session_start} - {session_end}")
                
                if sessions:
                    time_info = '; '.join(sessions)
            
            return time_info
            
        except Exception as e:
            print(f"Error getting class time: {e}")
            return "Not specified"

    def _get_coverage_status(self, coverage_rate, required, current):
        """Determine educator coverage status"""
        if current >= required:
            return "✅ Fully Covered"
        elif coverage_rate >= 75:
            return "🟡 Nearly Covered"
        elif coverage_rate >= 50:
            return "🟠 Partially Covered"
        elif current > 0:
            return "🔴 Under Covered"
        else:
            return "❌ No Coverage"
    
    def _get_participation_level(self, signup_count):
        """Determine educator participation level"""
        if signup_count >= 5:
            return "🌟 High"
        elif signup_count >= 3:
            return "📈 Moderate"
        elif signup_count >= 1:
            return "🟢 Active"
        else:
            return "⭕ None"
    
    def _get_urgency_level(self, current, required):
        """Determine urgency level for educator needs"""
        if current == 0:
            return "🔴 Critical"
        elif current < required * 0.5:
            return "🟡 Moderate"
        else:
            return "🟢 Low"
    
    # EXISTING METHODS (unchanged but enhanced with educator data)
    def _get_compliance_status(self, completion_rate, meeting_compliance):
        """Determine overall compliance status"""
        if completion_rate >= 100 and meeting_compliance:
            return "✅ Complete"
        elif completion_rate >= 80 and meeting_compliance:
            return "🟡 Nearly Complete"
        elif completion_rate >= 50:
            return "🟠 In Progress"
        else:
            return "🔴 Behind Schedule"
    
    def get_class_utilization_report(self):
        """Generate report showing class capacity utilization"""
        all_classes = self.excel.get_all_classes()
        utilization_data = []
        
        for class_name in all_classes:
            class_details = self.excel.get_class_details(class_name)
            enrollment_summary = self.enrollment.get_class_enrollment_summary(class_name)
            
            max_students = int(class_details.get('students_per_class', 21))
            instructor_requirement = class_details.get('instructors_per_day', 0)
            classes_per_day = int(class_details.get('classes_per_day', 1))
            is_two_day = class_details.get('is_two_day_class', 'No').lower() == 'yes'
            is_nurse_medic_separate = class_details.get('nurses_medic_separate', 'No').lower() == 'yes'
            
            total_capacity = 0
            total_enrolled = 0
            total_dates = 0
            total_educator_signups = 0
            educator_coverage = 0
            
            # Calculate capacity per date offering
            if is_nurse_medic_separate:
                # nurse/medic separate: slots per session × 2 roles × sessions per day
                capacity_per_date = max_students * classes_per_day
            else:
                # regular or single session: slots per session × sessions per day
                capacity_per_date = max_students * classes_per_day
            
            # Calculate across all dates
            for i in range(1, 15):  # Check rows 1-14
                date_key = f'date_{i}'
                if date_key in class_details and class_details[date_key]:
                    date_str = class_details[date_key]
                    
                    # For two-day classes, each configured date represents one offering
                    # but we need to count enrollments for both days
                    if is_two_day:
                        try:
                            date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                            day_1 = date_obj.strftime('%m/%d/%Y')
                            day_2 = (date_obj + timedelta(days=1)).strftime('%m/%d/%Y')
                            
                            # Count this as ONE date offering with ONE capacity
                            total_dates += 1
                            total_capacity += capacity_per_date
                            
                            # But count enrollments from both days
                            if day_1 in enrollment_summary:
                                total_enrolled += enrollment_summary[day_1]['total']
                            if day_2 in enrollment_summary:
                                total_enrolled += enrollment_summary[day_2]['total']
                        except:
                            # Fallback if date parsing fails
                            total_dates += 1
                            total_capacity += capacity_per_date
                            if date_str in enrollment_summary:
                                total_enrolled += enrollment_summary[date_str]['total']
                    else:
                        # Regular single-day class
                        total_dates += 1
                        total_capacity += capacity_per_date
                        
                        if date_str in enrollment_summary:
                            total_enrolled += enrollment_summary[date_str]['total']
                    
                    # Add educator data if available
                    if self.educator and instructor_requirement > 0:
                        educator_signups = self.db.get_educator_signup_count(class_name, date_str)
                        total_educator_signups += educator_signups
                        educator_coverage += instructor_requirement
            
            utilization_rate = (total_enrolled / total_capacity * 100) if total_capacity > 0 else 0
            educator_coverage_rate = (total_educator_signups / educator_coverage * 100) if educator_coverage > 0 else 0
            
            utilization_data.append({
                'Class Name': class_name,
                'Total Dates': total_dates,
                'Total Capacity': total_capacity,
                'Current Enrolled': total_enrolled,
                'Utilization Rate': f"{utilization_rate:.1f}%",
                'Available Slots': total_capacity - total_enrolled,
                'Educator Requirements': educator_coverage,
                'Educator Signups': total_educator_signups,
                'Educator Coverage': f"{educator_coverage_rate:.1f}%" if educator_coverage > 0 else "N/A",
                'Class Type': "Staff Meeting" if self.excel.is_staff_meeting(class_name) else "Training",
                'Status': self._get_utilization_status(utilization_rate)
            })
            
        return pd.DataFrame(utilization_data)

    def get_conflict_analysis_report(self):
        """Analyze schedule conflicts and overrides (including educator conflicts)"""
        all_staff = self.excel.get_staff_list()
        conflict_data = []
        
        for staff_name in all_staff:
            # Student enrollment conflicts
            conflict_enrollments = self.db.get_conflict_override_enrollments(staff_name)
            
            for enrollment in conflict_enrollments:
                conflict_data.append({
                    'Staff Name': staff_name,
                    'Type': 'Student Enrollment',
                    'Class Name': enrollment['class_name'],
                    'Class Date': enrollment['class_date'],
                    'Conflict Details': enrollment['conflict_details'],
                    'Override Date': enrollment.get('override_acknowledged_display', 'N/A'),
                    'Meeting Type': enrollment.get('meeting_type', 'N/A'),
                    'Session Time': enrollment.get('session_time', 'N/A')
                })
            
            # Educator signup conflicts
            if self.educator:
                educator_conflicts = self.db.get_conflict_override_educator_signups(staff_name)
                
                for signup in educator_conflicts:
                    conflict_data.append({
                        'Staff Name': staff_name,
                        'Type': 'Educator Signup',
                        'Class Name': signup['class_name'],
                        'Class Date': signup['class_date'],
                        'Conflict Details': signup['conflict_details'],
                        'Override Date': signup.get('override_acknowledged_display', 'N/A'),
                        'Meeting Type': 'N/A',
                        'Session Time': 'N/A'
                    })
        
        return pd.DataFrame(conflict_data)
    
    # ALL OTHER EXISTING METHODS remain unchanged...
    def get_staff_without_assignments(self):
        """Find staff members with no class assignments"""
        all_staff = self.excel.get_staff_list()
        unassigned_staff = []
        
        for staff_name in all_staff:
            assigned_classes = self.excel.get_assigned_classes(staff_name)
            if not assigned_classes:
                unassigned_staff.append(staff_name)
        
        return unassigned_staff
    
    def get_classes_without_enrollments(self):
        """Find classes with no current enrollments"""
        all_classes = self.excel.get_all_classes()
        unused_classes = []
        
        for class_name in all_classes:
            enrollments = self.db.get_class_enrollments(class_name)
            if not enrollments:
                unused_classes.append(class_name)
        
        return unused_classes
    
    def validate_excel_structure(self):
        """Validate Excel file structure and identify issues"""
        issues = []
        
        # Check if workbook loaded successfully
        if self.excel.load_error:
            issues.append(f"Excel Loading Error: {self.excel.load_error}")
            return issues
        
        # Check staff list
        staff_list = self.excel.get_staff_list()
        if not staff_list:
            issues.append("No staff members found in the roster")
        
        # Check for duplicate staff names
        duplicates = set([x for x in staff_list if staff_list.count(x) > 1])
        if duplicates:
            issues.append(f"Duplicate staff names found: {', '.join(duplicates)}")
        
        # Check class sheets
        all_classes = self.excel.get_all_classes()
        for class_name in all_classes:
            class_details = self.excel.get_class_details(class_name)
            
            # Check if class sheet exists and has valid data
            if not class_details or class_details == DEFAULT_CLASS_DETAILS:
                issues.append(f"Class '{class_name}' has no detailed configuration sheet")
                continue
            
            # Check for dates
            has_dates = any(class_details.get(f'date_{i}') for i in range(1, 9))
            if not has_dates:
                issues.append(f"Class '{class_name}' has no scheduled dates")
        
        return issues if issues else ["✅ Excel structure validation passed"]

    def get_individual_class_report(self, class_name):
        """Generate comprehensive report for a specific class"""
        class_details = self.excel.get_class_details(class_name)
        enrollment_summary = self.enrollment.get_class_enrollment_summary(class_name)
        
        if not class_details:
            return None
            
        # Basic class info
        report = {
            'class_name': class_name,
            'class_type': 'Staff Meeting' if self.excel.is_staff_meeting(class_name) else 'Training Class',
            'max_students_per_session': int(class_details.get('students_per_class', 21)),
            'classes_per_day': int(class_details.get('classes_per_day', 1)),
            'is_two_day_class': class_details.get('is_two_day_class', 'No'),
            'nurses_medic_separate': class_details.get('nurses_medic_separate', 'No'),
            'instructor_requirement': class_details.get('instructors_per_day', 0),
            'dates': [],
            'overall_stats': {},
            'staff_analysis': {},
            'educator_analysis': {}
        }
        
        # Check if this is a two-day class
        is_two_day = report['is_two_day_class'].lower() == 'yes'
        is_nurse_medic_separate = report['nurses_medic_separate'].lower() == 'yes'
        classes_per_day = report['classes_per_day']
        max_students_per_session = report['max_students_per_session']
        
        # Get all scheduled dates and their details
        total_capacity = 0
        total_enrolled = 0
        total_conflicts = 0
        total_educator_signups = 0
        total_educator_conflicts = 0
        
        for i in range(1, 15):  # Check rows 1-14 for dates
            date_key = f'date_{i}'
            if date_key in class_details and class_details[date_key]:
                date_str = class_details[date_key]
                location = class_details.get(f'date_{i}_location', '')
                has_live = class_details.get(f'date_{i}_has_live', False)
                can_work_n_prior = class_details.get(f'date_{i}_can_work_n_prior', False)
                
                # For two-day classes, expand to both days
                dates_to_process = []
                if is_two_day:
                    try:
                        date_obj = datetime.strptime(date_str, '%m/%d/%Y')
                        day_1 = date_obj.strftime('%m/%d/%Y')
                        day_2 = (date_obj + timedelta(days=1)).strftime('%m/%d/%Y')
                        dates_to_process = [
                            (day_1, 'Day 1'),
                            (day_2, 'Day 2')
                        ]
                    except:
                        dates_to_process = [(date_str, '')]
                else:
                    dates_to_process = [(date_str, '')]
                
                # Process each day (for two-day classes, this will be 2 iterations)
                for process_date, day_label in dates_to_process:
                    # Get enrollments for this specific date
                    date_enrollments = self.db.get_class_enrollments(class_name, process_date)
                    enrolled_count = len(date_enrollments)
                    conflict_count = sum(1 for e in date_enrollments if e.get('conflict_override', False))
                    
                    # Get educator signups for this specific date
                    educator_signups = 0
                    educator_conflicts = 0
                    educator_names = []
                    if self.educator:
                        educator_signups = self.db.get_educator_signup_count(class_name, process_date)
                        educator_roster = self.educator.get_class_educator_roster(class_name, process_date)
                        educator_conflicts = sum(1 for e in educator_roster if e.get('has_conflict', False))
                        educator_names = [e['staff_name'] for e in educator_roster if e['status'] == 'active']
                    
                    # ✅ CORRECTED CAPACITY CALCULATION
                    if is_nurse_medic_separate:
                        # nurse/medic separate: slots per session × 2 roles × sessions per day
                        capacity = max_students_per_session * classes_per_day
                    else:
                        # regular or single session: slots per session × sessions per day
                        capacity = max_students_per_session * classes_per_day
                    
                    # ✅ NEW: For two-day classes, only count Day 1 for capacity/enrollment tracking
                    # Day 2 is shown for display purposes but doesn't add to totals
                    should_count_in_totals = not is_two_day or day_label == 'Day 1'
                    
                    # Create display label for date
                    display_date = f"{process_date} ({day_label})" if day_label else process_date
                    
                    date_info = {
                        'date': process_date,
                        'display_date': display_date,
                        'location': location,
                        'has_live_option': has_live,
                        'night_shift_ok': can_work_n_prior,
                        'total_enrolled': enrolled_count,
                        'total_capacity': capacity,
                        'utilization_rate': (enrolled_count / capacity * 100) if capacity > 0 else 0,
                        'conflicts': conflict_count,
                        'available_slots': capacity - enrolled_count,
                        'enrollments': date_enrollments,
                        'educator_signups': educator_signups,
                        'educator_requirement': report['instructor_requirement'],
                        'educator_conflicts': educator_conflicts,
                        'educator_names': educator_names,
                        'educator_coverage_rate': (educator_signups / report['instructor_requirement'] * 100) if report['instructor_requirement'] > 0 else 0
                    }
                    
                    report['dates'].append(date_info)
                    
                    # ✅ Only add to totals for Day 1 of two-day classes
                    if should_count_in_totals:
                        total_capacity += capacity
                        total_enrolled += enrolled_count
                        total_conflicts += conflict_count
                        total_educator_signups += educator_signups
                        total_educator_conflicts += educator_conflicts                
        
        # Overall statistics
        total_educator_requirement = len(report['dates']) * report['instructor_requirement']
        
        report['overall_stats'] = {
            'total_dates': len(report['dates']),
            'total_capacity': total_capacity,
            'total_enrolled': total_enrolled,
            'overall_utilization': (total_enrolled / total_capacity * 100) if total_capacity > 0 else 0,
            'total_conflicts': total_conflicts,
            'total_available_slots': total_capacity - total_enrolled,
            'total_educator_requirement': total_educator_requirement,
            'total_educator_signups': total_educator_signups,
            'educator_coverage_rate': (total_educator_signups / total_educator_requirement * 100) if total_educator_requirement > 0 else 0,
            'total_educator_conflicts': total_educator_conflicts
        }
        
        # Get assigned vs enrolled staff
        assigned_staff = self._get_staff_assigned_to_class(class_name)
        all_enrollments = self.db.get_class_enrollments(class_name)
        enrolled_staff_names = list(set([e['staff_name'] for e in all_enrollments]))
        
        report['staff_analysis'] = {
            'total_assigned': len(assigned_staff),
            'total_enrolled': len(enrolled_staff_names),
            'enrollment_rate': (len(enrolled_staff_names) / len(assigned_staff) * 100) if assigned_staff else 0,
            'assigned_but_not_enrolled': [s for s in assigned_staff if s not in enrolled_staff_names],
            'enrolled_staff': enrolled_staff_names
        }
        
        # Educator analysis
        if self.educator:
            all_educator_signups = self.db.get_educator_signups_for_class(class_name)
            unique_educators = list(set([e['staff_name'] for e in all_educator_signups]))
            
            report['educator_analysis'] = {
                'unique_educators': len(unique_educators),
                'educator_names': unique_educators,
                'total_educator_signups': total_educator_signups,
                'dates_needing_educators': sum(1 for date_info in report['dates'] 
                                            if date_info['educator_signups'] < date_info['educator_requirement']),
                'fully_covered_dates': sum(1 for date_info in report['dates'] 
                                        if date_info['educator_signups'] >= date_info['educator_requirement'])
            }
        else:
            report['educator_analysis'] = {
                'unique_educators': 0,
                'educator_names': [],
                'total_educator_signups': 0,
                'dates_needing_educators': 0,
                'fully_covered_dates': 0
            }
        
        return report

    def _get_staff_assigned_to_class(self, class_name):
        """Get list of staff assigned to a specific class"""
        all_staff = self.excel.get_staff_list()
        assigned_staff = []
        
        for staff_name in all_staff:
            assigned_classes = self.excel.get_assigned_classes(staff_name)
            if class_name in assigned_classes:
                assigned_staff.append(staff_name)
        
        return assigned_staff
    
    def export_class_roster(self, class_name, date_str=None):
        """Export roster for a specific class/date in printable format"""
        if date_str:
            # Single date roster
            enrollments = self.db.get_class_enrollments(class_name, date_str)
            title = f"{class_name} - {date_str}"
        else:
            # All dates roster
            enrollments = self.db.get_class_enrollments(class_name)
            title = f"{class_name} - All Dates"
        
        roster_data = []
        for enrollment in enrollments:
            roster_data.append({
                'Staff Name': enrollment['staff_name'],
                'Date': enrollment['class_date'],
                'Role': enrollment.get('role', 'General'),
                'Meeting Type': enrollment.get('meeting_type', ''),
                'Session Time': enrollment.get('session_time', ''),
                'Has Conflict': '⚠️' if enrollment.get('conflict_override') else '',
                'Enrollment Date': enrollment.get('enrollment_date', ''),
                'Status': enrollment.get('status', 'active')
            })
        
        df = pd.DataFrame(roster_data)
        df = df.sort_values(['Date', 'Staff Name'])
        
        return df, title
    
    def export_educator_roster(self, class_name, date_str=None):
        """Export educator roster for a specific class/date"""
        if not self.educator:
            return pd.DataFrame(), "No educator data available"
        
        if date_str:
            # Single date educator roster
            educators = self.educator.get_class_educator_roster(class_name, date_str)
            title = f"{class_name} Educators - {date_str}"
        else:
            # All dates educator roster
            educators = self.educator.get_class_educator_roster(class_name)
            title = f"{class_name} Educators - All Dates"
        
        roster_data = []
        for educator in educators:
            roster_data.append({
                'Educator Name': educator['staff_name'],
                'Date': educator['class_date'],
                'Has Conflict': '⚠️' if educator.get('has_conflict') else '',
                'Conflict Details': educator.get('conflict_details', ''),
                'Signup Date': educator.get('signup_date', ''),
                'Status': educator.get('status', 'active')
            })
        
        df = pd.DataFrame(roster_data)
        if not df.empty:
            df = df.sort_values(['Date', 'Educator Name'])
        
        return df, title
    
    def get_class_completion_tracking(self, class_name):
        """Track completion status for all assigned staff"""
        assigned_staff = self._get_staff_assigned_to_class(class_name)
        class_details = self.excel.get_class_details(class_name)
        
        completion_data = []
        
        for staff_name in assigned_staff:
            staff_enrollments = [e for e in self.db.get_class_enrollments(class_name) 
                               if e['staff_name'] == staff_name]
            
            # Get all available dates for this class
            available_dates = []
            for i in range(1, 9):
                date_key = f'date_{i}'
                if date_key in class_details and class_details[date_key]:
                    available_dates.append(class_details[date_key])
            
            enrolled_dates = [e['class_date'] for e in staff_enrollments]
            
            # For staff meetings, check LIVE requirement
            live_count = sum(1 for e in staff_enrollments 
                           if e.get('meeting_type') == 'LIVE')
            
            is_staff_meeting = self.excel.is_staff_meeting(class_name)
            meeting_compliance = live_count >= 2 if is_staff_meeting else True
            
            # Get educator signup info if available
            educator_signups = 0
            if self.educator:
                educator_signups = len([e for e in self.educator.get_staff_educator_signups(staff_name)
                                      if e['class_name'] == class_name])
            
            completion_data.append({
                'Staff Name': staff_name,
                'Enrolled Dates': len(enrolled_dates),
                'Available Dates': len(available_dates),
                'Completion Status': 'Complete' if enrolled_dates else 'Not Started',
                'LIVE Meetings': f"{live_count}/2" if is_staff_meeting else 'N/A',
                'Meeting Compliance': '✅' if meeting_compliance else '❌',
                'Conflicts': sum(1 for e in staff_enrollments if e.get('conflict_override')),
                'Educator Signups': educator_signups,
                'Enrolled Dates List': ', '.join(enrolled_dates) if enrolled_dates else 'None'
            })
        
        return pd.DataFrame(completion_data)

    def _get_utilization_status(self, utilization_rate):
        """Determine utilization status"""
        if utilization_rate >= 90:
            return "🔴 Nearly Full"
        elif utilization_rate >= 70:
            return "🟡 Good Utilization"
        elif utilization_rate >= 40:
            return "🟠 Moderate"
        else:
            return "🟢 Low Utilization"


# Integration function for enhanced admin reports with educator functionality and comprehensive schedule report
def enhance_admin_reports(admin_access_instance, excel_admin_functions):
    """Add enhanced reporting to admin access including educator reports and comprehensive schedule"""
    
    def _show_enhanced_enrollment_reports():
        st.subheader("📈 Enhanced Enrollment Reports")
        
        # Add educator tab if educator functionality is available
        has_educator = excel_admin_functions.educator is not None
        
        if has_educator:
            tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
                "📊 Compliance", "🎯 Utilization", "⚠️ Conflicts", "👨‍🏫 Educators", 
                "📋 Individual Classes", "🔍 Validation", "📅 Schedule Report"
            ])
        else:
            tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
                "📊 Compliance", "🎯 Utilization", "⚠️ Conflicts", 
                "📋 Individual Classes", "🔍 Validation", "📅 Schedule Report"
            ])
        

        with tab1:
            st.write("### Staff Enrollment Compliance")
            try:
                compliance_df = excel_admin_functions.get_enrollment_compliance_report()
                
                if not compliance_df.empty:
                    # Summary metrics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        complete_count = len(compliance_df[compliance_df['Status'].str.contains('Complete', na=False)])
                        st.metric("Fully Compliant", f"{complete_count}/{len(compliance_df)}")
                    with col2:
                        avg_completion = compliance_df['Completion Rate'].mean() * 100
                        st.metric("Avg Completion", f"{avg_completion:.1f}%")
                    with col3:
                        total_conflicts = compliance_df['Conflict Overrides'].sum()
                        st.metric("Total Conflicts", total_conflicts)
                    with col4:
                        if has_educator:
                            total_educator_signups = compliance_df['Educator Signups'].sum()
                            st.metric("Educator Signups", total_educator_signups)
                        else:
                            behind_schedule = len(compliance_df[compliance_df['Status'] == 'ðŸ"´ Behind Schedule'])
                            st.metric("Behind Schedule", behind_schedule)
                    
                    # Detailed table
                    display_df = compliance_df.copy()
                    display_df['Completion Rate'] = display_df['Completion Rate'].apply(lambda x: f"{x*100:.1f}%")
                    st.dataframe(display_df, use_container_width=True)
                    
                    # Export functionality
                    if st.button("📥 Export Compliance Report"):
                        try:
                            from io import BytesIO
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.worksheet.table import Table, TableStyleInfo
                            
                            with st.spinner("Generating Excel report..."):
                                # Create Excel file
                                output = BytesIO()
                                
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    # Write the dataframe
                                    compliance_df.to_excel(writer, sheet_name='Compliance Report', index=False)
                                    
                                    # Get the worksheet to format it
                                    workbook = writer.book
                                    worksheet = workbook['Compliance Report']
                                    
                                    # Format Completion Rate column as percentage
                                    if 'Completion Rate' in compliance_df.columns:
                                        from openpyxl.styles import numbers
                                        completion_rate_col_idx = list(compliance_df.columns).index('Completion Rate') + 1
                                        
                                        for row_idx in range(2, worksheet.max_row + 1):
                                            cell = worksheet.cell(row=row_idx, column=completion_rate_col_idx)
                                            cell.number_format = numbers.FORMAT_PERCENTAGE

                                    # Auto-adjust column widths
                                    for idx, col in enumerate(compliance_df.columns):
                                        try:
                                            max_length = max(
                                                compliance_df[col].astype(str).apply(len).max(),
                                                len(col)
                                            ) + 2
                                            # Convert column index to letter (A, B, C, etc.)
                                            if idx < 26:
                                                column_letter = chr(65 + idx)
                                            else:
                                                # Handle columns beyond Z (AA, AB, etc.)
                                                column_letter = chr(64 + idx // 26) + chr(65 + idx % 26)
                                            
                                            worksheet.column_dimensions[column_letter].width = min(max_length, 50)
                                        except Exception as col_error:
                                            print(f"Error setting width for column {idx}: {col_error}")
                                            # Set default width if error
                                            if idx < 26:
                                                worksheet.column_dimensions[chr(65 + idx)].width = 15
                                    
                                    # ===== NEW: Create Excel Table =====
                                    try:
                                        # Define table range (from A1 to last column and row)
                                        last_col_letter = openpyxl.utils.get_column_letter(len(compliance_df.columns))
                                        last_row = len(compliance_df) + 1  # +1 for header row
                                        table_ref = f"A1:{last_col_letter}{last_row}"
                                        
                                        # Create table with a unique name
                                        table = Table(displayName="ComplianceTable", ref=table_ref)
                                        
                                        # Add table style with filter buttons and banded rows
                                        style = TableStyleInfo(
                                            name="TableStyleMedium9",  # Blue theme that matches existing header color
                                            showFirstColumn=False,
                                            showLastColumn=False,
                                            showRowStripes=True,  # Banded rows
                                            showColumnStripes=False
                                        )
                                        table.tableStyleInfo = style
                                        
                                        # Add the table to the worksheet
                                        worksheet.add_table(table)
                                        
                                        print(f"DEBUG: Added Excel table with range {table_ref}")
                                        
                                    except Exception as table_error:
                                        print(f"Error creating Excel table: {table_error}")
                                        import traceback
                                        traceback.print_exc()
                                        # Fall back to manual header formatting if table creation fails
                                        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                                        header_font = Font(color="FFFFFF", bold=True)
                                        
                                        for cell in worksheet[1]:
                                            cell.fill = header_fill
                                            cell.font = header_font
                                            cell.alignment = Alignment(horizontal='center', vertical='center')
                                    # ===== END NEW TABLE CODE =====
                                    
                                    # Add borders to all cells (still useful even with table)
                                    try:
                                        thin_border = Border(
                                            left=Side(style='thin'),
                                            right=Side(style='thin'),
                                            top=Side(style='thin'),
                                            bottom=Side(style='thin')
                                        )
                                        
                                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                                                    min_col=1, max_col=worksheet.max_column):
                                            for cell in row:
                                                cell.border = thin_border
                                    except Exception as border_error:
                                        print(f"Error adding borders: {border_error}")
                                    
                                    # Color code the Status column
                                    try:
                                        if 'Status' in compliance_df.columns:
                                            status_col_idx = list(compliance_df.columns).index('Status') + 1
                                            
                                            for row_idx in range(2, worksheet.max_row + 1):
                                                cell = worksheet.cell(row=row_idx, column=status_col_idx)
                                                status_value = str(cell.value) if cell.value else ''
                                                
                                                if '✅' in status_value or 'Complete' in status_value:
                                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                                elif '🟡' in status_value or 'Nearly' in status_value:
                                                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                                elif '🔴' in status_value or 'Behind' in status_value:
                                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    except Exception as color_error:
                                        print(f"Error color coding Status column: {color_error}")
                                
                                # Get the Excel data
                                excel_data = output.getvalue()
                                
                                # Provide download button
                                st.download_button(
                                    label="📊 Download Compliance Report (Excel)",
                                    data=excel_data,
                                    file_name=f"compliance_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                                st.success("✅ Excel report ready for download!")
                                
                        except Exception as e:
                            st.error(f"Error generating Excel export: {str(e)}")
                            print(f"Excel Export Error: {e}")
                            import traceback
                            traceback.print_exc()
                            
                            # Fallback to CSV export
                            st.warning("Falling back to CSV export...")
                            csv = compliance_df.to_csv(index=False)
                            st.download_button(
                                label="📄 Download as CSV (Fallback)",
                                data=csv,
                                file_name=f"compliance_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                                mime="text/csv",
                                use_container_width=True
                            )
                else:
                    st.info("No compliance data available")
                    
            except Exception as e:
                st.error(f"Error loading compliance data: {str(e)}")        

        with tab2:
            st.write("### Class Utilization Analysis")
            try:
                utilization_df = excel_admin_functions.get_class_utilization_report()
                
                if not utilization_df.empty:
                    # Utilization metrics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        avg_utilization = utilization_df['Utilization Rate'].str.rstrip('%').astype(float).mean()
                        st.metric("Avg Utilization", f"{avg_utilization:.1f}%")
                    with col2:
                        total_capacity = utilization_df['Total Capacity'].sum()
                        total_enrolled = utilization_df['Current Enrolled'].sum()
                        st.metric("Overall Utilization", f"{total_enrolled}/{total_capacity}")
                    with col3:
                        nearly_full = len(utilization_df[utilization_df['Status'] == '🔴 Nearly Full'])
                        st.metric("Nearly Full Classes", nearly_full)
                    
                    st.dataframe(utilization_df, use_container_width=True)
                else:
                    st.info("No utilization data available")
            except Exception as e:
                st.error(f"Error generating utilization report: {str(e)}")
        
        with tab3:
            st.write("### Schedule Conflict Analysis")
            try:
                conflict_df = excel_admin_functions.get_conflict_analysis_report()
                
                if not conflict_df.empty:
                    # Show conflict breakdown
                    if has_educator:
                        conflict_types = conflict_df['Type'].value_counts()
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Student Conflicts", conflict_types.get('Student Enrollment', 0))
                        with col2:
                            st.metric("Educator Conflicts", conflict_types.get('Educator Signup', 0))
                    
                    st.warning(f"Found {len(conflict_df)} schedule conflicts requiring manual resolution")
                    st.dataframe(conflict_df, use_container_width=True)
                else:
                    st.success("No schedule conflicts found!")
            except Exception as e:
                st.error(f"Error generating conflict report: {str(e)}")
        
        # EDUCATOR TAB (only if educator functionality is available)
        if has_educator:
            with tab4:
                st.write("### 👨‍🏫 Educator Coverage Analysis")
                
                try:
                    # Educator coverage report
                    coverage_df = excel_admin_functions.get_educator_coverage_report()
                    
                    if not coverage_df.empty:
                        # Coverage summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            fully_covered = len(coverage_df[coverage_df['Status'] == '✅ Fully Covered'])
                            st.metric("Fully Covered", fully_covered)
                        
                        with col2:
                            total_needed = coverage_df['Still Needed'].sum()
                            st.metric("Total Positions Needed", total_needed)
                        
                        with col3:
                            critical_classes = len(coverage_df[coverage_df['Status'] == '❌ No Coverage'])
                            st.metric("Classes w/o Educators", critical_classes)
                        
                        with col4:
                            avg_coverage = coverage_df['Coverage Rate'].str.rstrip('%').astype(float).mean()
                            st.metric("Avg Coverage", f"{avg_coverage:.1f}%")
                        
                        st.write("#### Educator Coverage by Class/Date")
                        st.dataframe(coverage_df, use_container_width=True)
                        
                        # Classes needing educators
                        st.write("#### 🚨 Priority - Classes Still Needing Educators")
                        needs_educators_df = excel_admin_functions.get_classes_needing_educators_report()
                        
                        if not needs_educators_df.empty:
                            st.dataframe(needs_educators_df, use_container_width=True)
                        else:
                            st.success("✅ All educator positions are filled!")
                        
                        # Individual educator participation
                        st.write("#### Individual Educator Participation")
                        participation_df = excel_admin_functions.get_educator_participation_report()
                        
                        if not participation_df.empty:
                            st.dataframe(participation_df, use_container_width=True)
                        else:
                            st.info("No educator signups found.")
                        
                        # Export functionality
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            if st.button("📥 Export Coverage Report"):
                                csv = coverage_df.to_csv(index=False)
                                st.download_button(
                                    "Download Coverage CSV",
                                    csv,
                                    f"educator_coverage_{datetime.now().strftime('%Y%m%d')}.csv",
                                    "text/csv"
                                )
                        
                        with col2:
                            if st.button("📥 Export Needs Report"):
                                csv = needs_educators_df.to_csv(index=False)
                                st.download_button(
                                    "Download Needs CSV",
                                    csv,
                                    f"educator_needs_{datetime.now().strftime('%Y%m%d')}.csv",
                                    "text/csv"
                                )
                        
                        with col3:
                            if st.button("📥 Export Participation Report"):
                                csv = participation_df.to_csv(index=False)
                                st.download_button(
                                    "Download Participation CSV",
                                    csv,
                                    f"educator_participation_{datetime.now().strftime('%Y%m%d')}.csv",
                                    "text/csv"
                                )
                    
                    else:
                        st.info("No educator data available - no classes require educators.")
                        
                except Exception as e:
                    st.error(f"Error generating educator reports: {str(e)}")
        
        # Individual classes tab
        individual_tab_idx = tab5 if has_educator else tab4
        with individual_tab_idx:
            st.write("### Enhanced Individual Class Reports")
            
            try:
                # Class selection
                all_classes = excel_admin_functions.excel.get_all_classes()
                selected_class = st.selectbox("Select Class for Detailed Report:", [""] + all_classes)
                
                if selected_class:
                    class_report = excel_admin_functions.get_individual_class_report(selected_class)
                    
                    if class_report:
                        # Class overview header
                        st.write(f"## 📚 {class_report['class_name']}")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Class Type", class_report['class_type'])
                        with col2:
                            st.metric("Total Capacity", class_report['overall_stats']['total_capacity'])
                        with col3:
                            st.metric("Total Enrolled", class_report['overall_stats']['total_enrolled'])
                        with col4:
                            utilization = class_report['overall_stats']['overall_utilization']
                            st.metric("Utilization", f"{utilization:.1f}%")
                        
                        # Educator overview (if applicable)
                        if has_educator and class_report['instructor_requirement'] > 0:
                            st.write("### 👨‍🏫 Educator Coverage Overview")
                            
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Required per Date", class_report['instructor_requirement'])
                            with col2:
                                st.metric("Total Required", class_report['overall_stats']['total_educator_requirement'])
                            with col3:
                                st.metric("Total Signups", class_report['overall_stats']['total_educator_signups'])
                            with col4:
                                coverage_rate = class_report['overall_stats']['educator_coverage_rate']
                                st.metric("Coverage Rate", f"{coverage_rate:.1f}%")
                        
                        # Enhanced detailed breakdown by date/session
                        st.write("### 📅 Detailed Breakdown by Date and Session")
                        
                        if class_report['dates']:
                            # Create tabs for each date
                            date_tabs = st.tabs([f"📅 {date_info['date']}" for date_info in class_report['dates']])
                            
                            for tab_idx, date_info in enumerate(class_report['dates']):
                                with date_tabs[tab_idx]:
                                    _display_enhanced_date_breakdown(
                                        date_info, selected_class, excel_admin_functions, 
                                        class_report, has_educator
                                    )
                        else:
                            st.warning("No dates configured for this class")
                        
                        # Export options
                        st.write("### 📥 Export Options")
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            if st.button("Export Class Roster"):
                                roster_df, title = excel_admin_functions.export_class_roster(selected_class)
                                
                                # Create Excel file instead of CSV
                                from io import BytesIO
                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    roster_df.to_excel(writer, sheet_name='Class Roster', index=False)
                                    
                                    # Auto-adjust column widths
                                    worksheet = writer.sheets['Class Roster']
                                    for idx, col in enumerate(roster_df.columns):
                                        max_length = max(
                                            roster_df[col].astype(str).apply(len).max(),
                                            len(col)
                                        ) + 2
                                        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
                                
                                excel_data = output.getvalue()
                                
                                st.download_button(
                                    "Download Roster Excel",
                                    excel_data,
                                    f"{selected_class.replace(' ', '_')}_roster_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )

                        with col2:
                            if st.button("Export Completion Tracking"):
                                completion_df = excel_admin_functions.get_class_completion_tracking(selected_class)
                                
                                # Create Excel file instead of CSV
                                from io import BytesIO
                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    completion_df.to_excel(writer, sheet_name='Completion Tracking', index=False)
                                    
                                    # Auto-adjust column widths
                                    worksheet = writer.sheets['Completion Tracking']
                                    for idx, col in enumerate(completion_df.columns):
                                        max_length = max(
                                            completion_df[col].astype(str).apply(len).max(),
                                            len(col)
                                        ) + 2
                                        worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
                                
                                excel_data = output.getvalue()
                                
                                st.download_button(
                                    "Download Completion Excel",
                                    excel_data,
                                    f"{selected_class.replace(' ', '_')}_completion_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )

                        with col3:
                            if has_educator and class_report['instructor_requirement'] > 0:
                                if st.button("Export Educator Roster"):
                                    educator_df, title = excel_admin_functions.export_educator_roster(selected_class)
                                    
                                    # Create Excel file instead of CSV
                                    from io import BytesIO
                                    output = BytesIO()
                                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                        educator_df.to_excel(writer, sheet_name='Educator Roster', index=False)
                                        
                                        # Auto-adjust column widths
                                        worksheet = writer.sheets['Educator Roster']
                                        for idx, col in enumerate(educator_df.columns):
                                            max_length = max(
                                                educator_df[col].astype(str).apply(len).max(),
                                                len(col)
                                            ) + 2
                                            worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
                                    
                                    excel_data = output.getvalue()
                                    
                                    st.download_button(
                                        "Download Educator Excel",
                                        excel_data,
                                        f"{selected_class.replace(' ', '_')}_educators_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )

                    else:
                        st.error(f"Could not generate report for {selected_class}")
            except Exception as e:
                st.error(f"Error with individual class reports: {str(e)}")
        
        # Validation tab
        validation_tab_idx = tab6 if has_educator else tab5
        with validation_tab_idx:
            st.write("### Excel Structure Validation")
            try:
                validation_issues = excel_admin_functions.validate_excel_structure()
                
                if "✅ Excel structure validation passed" in validation_issues:
                    st.success("Excel file structure is valid!")
                else:
                    st.error("Validation issues found:")
                    for issue in validation_issues:
                        st.write(f"• {issue}")
                
                # Additional checks
                unassigned_staff = excel_admin_functions.get_staff_without_assignments()
                if unassigned_staff:
                    st.warning(f"Staff without class assignments: {', '.join(unassigned_staff)}")
                
                unused_classes = excel_admin_functions.get_classes_without_enrollments()
                if unused_classes:
                    st.info(f"Classes with no enrollments: {', '.join(unused_classes)}")
            except Exception as e:
                st.error(f"Error with validation: {str(e)}")
        
        # COMPREHENSIVE SCHEDULE REPORT TAB
        schedule_report_tab = tab7 if has_educator else tab6
        with schedule_report_tab:
            _show_comprehensive_schedule_report()

    def _show_comprehensive_schedule_report():
        """Show comprehensive education schedule report with date range selection"""
        st.subheader("📅 Comprehensive Education Schedule Report")
        st.write("Generate a complete schedule showing all staff enrollments and educator signups across a date range.")
        
        # Date range selection
        col1, col2 = st.columns(2)
        
        with col1:
            start_date = st.date_input(
                "Start Date",
                value=datetime.now().replace(day=1),  # Default to first of current month
                key="schedule_report_start_date"
            )
        
        with col2:
            # Default end date to end of current month
            now = datetime.now()
            last_day = calendar.monthrange(now.year, now.month)[1]
            default_end = now.replace(day=last_day)
            
            end_date = st.date_input(
                "End Date",
                value=default_end,
                key="schedule_report_end_date"
            )
        
        # Validation
        if start_date > end_date:
            st.error("Start date must be before or equal to end date.")
            return
        
        # Calculate date range info
        date_diff = (end_date - start_date).days + 1
        if date_diff > 90:
            st.warning(f"⚠️ Large date range selected ({date_diff} days). This may take longer to generate.")
        
        st.info(f"📊 Report will cover {date_diff} days from {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}")
        
        # Generate report button
        if st.button("📊 Generate Comprehensive Schedule Report", type="primary", use_container_width=True):
            try:
                with st.spinner("Generating comprehensive education schedule report..."):
                    # Generate the report
                    schedule_df = excel_admin_functions.get_comprehensive_education_schedule_report(
                        start_date.strftime('%Y-%m-%d'),
                        end_date.strftime('%Y-%m-%d')
                    )
                
                if not schedule_df.empty:
                    st.success(f"✅ Report generated successfully! Found {len(schedule_df)} staff members.")
                    
                    # Display summary metrics
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        total_staff = len(schedule_df)
                        st.metric("Total Staff", total_staff)
                    
                    with col2:
                        # Count staff with any activities
                        date_columns = [col for col in schedule_df.columns if col not in ['STAFF NAME']]
                        staff_with_activities = 0
                        for _, row in schedule_df.iterrows():
                            if any(row[col] != '' for col in date_columns):
                                staff_with_activities += 1
                        st.metric("Staff with Activities", staff_with_activities)
                    
                    with col3:
                        # Count total activities
                        total_activities = 0
                        for _, row in schedule_df.iterrows():
                            for col in date_columns:
                                if row[col] != '':
                                    # Count comma-separated activities
                                    activities = row[col].split(',')
                                    total_activities += len([a.strip() for a in activities if a.strip()])
                        st.metric("Total Activities", total_activities)
                    
                    with col4:
                        # Count educator activities
                        educator_activities = 0
                        for _, row in schedule_df.iterrows():
                            for col in date_columns:
                                if 'EDU:' in row[col]:
                                    activities = row[col].split(',')
                                    educator_activities += len([a.strip() for a in activities if 'EDU:' in a.strip()])
                        st.metric("Educator Signups", educator_activities)
                    
                    st.markdown("---")
                    
                    # Display the report (with pagination for large reports)
                    st.write("### 📋 Schedule Report Preview")
                    
                    if len(schedule_df) > 50:
                        st.info(f"Large report with {len(schedule_df)} staff members. Showing first 50 rows in preview.")
                        st.dataframe(schedule_df.head(50), use_container_width=True)
                        st.info("Download the Excel file to see the complete report.")
                    else:
                        st.dataframe(schedule_df, use_container_width=True)
                    
                    # Export functionality
                    st.markdown("---")
                    st.write("### 📥 Export Options")
                    st.markdown("---")
                    
                    # Excel Export
                    excel_data = excel_admin_functions.export_comprehensive_schedule_to_excel(
                        schedule_df, 
                        start_date.strftime('%m/%d/%Y'), 
                        end_date.strftime('%m/%d/%Y')
                    )
                    
                    if excel_data:
                        st.download_button(
                            label="📊 Download as Excel",
                            type="primary",
                            data=excel_data,
                            file_name=f"education_schedule_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                            st.error("Failed to generate Excel file")
                    
                    # Legend
                    st.markdown("---")
                    st.write("### 📖 Legend")
                    st.write("• **Regular text** = Student enrollment in class")
                    st.write("• **EDU:** prefix = Educator signup for class")
                    st.write("• **Multiple activities** = Separated by commas")
                    st.write("• **Empty cells** = No scheduled activities")
                    
                else:
                    st.warning("No data found for the selected date range. Please check:")
                    st.write("• Date range includes dates with scheduled classes")
                    st.write("• Staff members have enrollments or educator signups")
                    st.write("• Database contains enrollment and educator signup data")
                    
            except Exception as e:
                st.error(f"Error generating report: {str(e)}")
                import traceback
                traceback.print_exc()

    def _display_enhanced_date_breakdown(date_info, class_name, excel_admin_functions, class_report, has_educator):
        """Display enhanced breakdown for a specific date with sessions and participants"""
        
        date_str = date_info['date']
        location = date_info.get('location', 'Not specified')
        has_live_option = date_info.get('has_live_option', False)
        night_shift_ok = date_info.get('night_shift_ok', False)
        is_two_day = class_report.get('is_two_day_class', 'No').lower() == 'yes'
        
        # Date header with metadata
        st.write(f"#### 📍 Location: {location}")
        
        if night_shift_ok:
            st.info("🌙 Night shift prior OK for this date")
        
        if is_two_day:
            st.info("📅 Two-Day Class: Enrollment covers consecutive days")
        
        # Get session options for this date
        session_options = excel_admin_functions.enrollment.get_available_session_options(class_name, date_str)
        
        if not session_options:
            st.warning("No session data available for this date")
            return
        
        # Display sessions
        for session_idx, session_option in enumerate(session_options):
            session_type = session_option.get('type', 'unknown')
            
            # Create session header
            if session_type == 'staff_meeting':
                _display_staff_meeting_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report)
            elif session_type in ['nurse_medic_separate', 'nurse_medic_separate_single']:
                _display_nurse_medic_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report)
            elif session_type in ['regular', 'regular_single']:
                _display_regular_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report)
            else:
                st.warning(f"Unknown session type: {session_type}")

    def _display_nurse_medic_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report):
        """Display nurse/medic separated session breakdown"""
        session_time = session_option.get('session_time')
        display_time = session_option.get('display_time', 'Session')
        
        with st.expander(f"⚕️ **{display_time}** (Role-Separated)", expanded=True):
            nurses = session_option.get('nurses', [])
            medics = session_option.get('medics', [])
            nurse_available = session_option.get('nurse_available', True)
            medic_available = session_option.get('medic_available', True)
            
            # FIXED: For nurse/medic separate, max_students_per_session is the TOTAL capacity
            # which is split equally between nurses and medics
            max_students_per_session = class_report['max_students_per_session']
            nurse_capacity = max_students_per_session // 2  # Half for nurses
            medic_capacity = max_students_per_session // 2  # Half for medics
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.write("**👩‍⚕️ Nurses**")
                nurse_spots_remaining = max(0, nurse_capacity - len(nurses))
                st.write(f"**Enrolled:** {len(nurses)}/{nurse_capacity}")
                st.write(f"**Spots Remaining:** {nurse_spots_remaining}")
                
                if nurses:
                    st.write("**Nurse Participants:**")
                    for nurse in sorted(nurses):
                        st.write(f"• {nurse}")
                else:
                    st.write("*No nurses enrolled*")
            
            with col2:
                st.write("**🚑 Medics**")
                medic_spots_remaining = max(0, medic_capacity - len(medics))
                st.write(f"**Enrolled:** {len(medics)}/{medic_capacity}")
                st.write(f"**Spots Remaining:** {medic_spots_remaining}")
                
                if medics:
                    st.write("**Medic Participants:**")
                    for medic in sorted(medics):
                        st.write(f"• {medic}")
                else:
                    st.write("*No medics enrolled*")
            
            with col3:
                if has_educator and class_report['instructor_requirement'] > 0:
                    _display_educator_breakdown(date_str, class_name, excel_admin_functions, class_report)

    def _display_regular_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report):
        """Display regular session breakdown"""
        session_time = session_option.get('session_time')
        display_time = session_option.get('display_time', 'Regular Session')
        enrolled_participants = session_option.get('enrolled', [])
        available_slots = session_option.get('available_slots', 0)
        max_capacity = len(enrolled_participants) + available_slots
        
        session_header = f"📚 **{display_time}**" if session_time else "📚 **Regular Class Session**"
        
        with st.expander(session_header, expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**👥 Attendees**")
                st.write(f"**Enrolled:** {len(enrolled_participants)}/{max_capacity}")
                st.write(f"**Spots Remaining:** {available_slots}")
                
                if enrolled_participants:
                    st.write("**Participant List:**")
                    for participant in sorted(enrolled_participants):
                        st.write(f"• {participant}")
                else:
                    st.write("*No participants enrolled*")
            
            with col2:
                if has_educator and class_report['instructor_requirement'] > 0:
                    _display_educator_breakdown(date_str, class_name, excel_admin_functions, class_report)

    def _display_educator_breakdown(date_str, class_name, excel_admin_functions, class_report):
        """Display educator breakdown for a specific date"""
        st.write("**👨‍🏫 Educators**")
        
        instructor_requirement = class_report['instructor_requirement']
        
        # Get educator roster for this date
        educator_roster = []
        if excel_admin_functions.educator:
            educator_roster = excel_admin_functions.educator.get_class_educator_roster(class_name, date_str)
        
        current_signups = len([e for e in educator_roster if e['status'] == 'active'])
        spots_remaining = max(0, instructor_requirement - current_signups)
        
        st.write(f"**Signed Up:** {current_signups}/{instructor_requirement}")
        st.write(f"**Spots Remaining:** {spots_remaining}")
        
        if educator_roster:
            active_educators = [e for e in educator_roster if e['status'] == 'active']
            if active_educators:
                st.write("**Educator List:**")
                for educator in sorted(active_educators, key=lambda x: x['staff_name']):
                    educator_display = f"• {educator['staff_name']}"
                    if educator.get('has_conflict'):
                        educator_display += " ⚠️"
                    st.write(educator_display)
            else:
                st.write("*No educators signed up*")
        else:
            st.write("*No educators signed up*")
        
        # Show status indicator
        if current_signups >= instructor_requirement:
            st.success("✅ Fully Staffed")
        elif current_signups >= instructor_requirement * 0.5:
            st.warning("🟡 Partially Staffed")
        else:
            st.error("🔴 Needs Educators")   

    def _display_staff_meeting_session(session_option, date_str, class_name, excel_admin_functions, has_educator, class_report):
        """Display staff meeting session breakdown"""
        meeting_type = session_option.get('meeting_type', 'Virtual')
        meeting_icon = "🔴" if meeting_type == 'LIVE' else "💻"
        
        with st.expander(f"{meeting_icon} **{meeting_type} Staff Meeting Session**", expanded=True):
            # Get enrolled participants
            enrolled_participants = session_option.get('enrolled', [])
            available_slots = session_option.get('available_slots', 0)
            max_capacity = len(enrolled_participants) + available_slots
            
            # Attendee breakdown
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**👥 Attendees**")
                st.write(f"**Enrolled:** {len(enrolled_participants)}/{max_capacity}")
                st.write(f"**Spots Remaining:** {available_slots}")
                
                if enrolled_participants:
                    st.write("**Participant List:**")
                    for participant in sorted(enrolled_participants):
                        st.write(f"• {participant}")
                else:
                    st.write("*No participants enrolled*")
            
            with col2:
                if has_educator and class_report['instructor_requirement'] > 0:
                    _display_educator_breakdown(date_str, class_name, excel_admin_functions, class_report)
                else:
                    st.write("**ℹ️ No educators required for staff meetings**")
                    
    admin_access_instance._show_enhanced_enrollment_reports = _show_enhanced_enrollment_reports
                    