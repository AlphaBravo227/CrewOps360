# training_modules/excel_handler.py - Updated for new Excel layout with up to 14 date rows
import pandas as pd
import openpyxl
from datetime import datetime, time
import os
from training_modules.config import NON_CLASS_COLUMNS

# Enhanced default class details
DEFAULT_CLASS_DETAILS = {
    'students_per_class': 21,
    'nurses_medic_separate': 'No',
    'classes_per_day': 1,
    'is_two_day_class': 'No',
    'time_1_start': '08:00',
    'time_1_end': '16:00',
    'instructors_per_day': 0  # Default to 0 instructors needed
}

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.workbook = None
        self.enrollment_sheet = None
        self.load_error = None
        self._load_workbook()
        
    def _load_workbook(self):
        """Load the Excel workbook"""
        try:
            # Check if file exists
            if not os.path.exists(self.excel_path):
                self.load_error = f"Excel file not found: {self.excel_path}"
                print(self.load_error)
                return
                
            # Try to load with pandas first to check structure
            try:
                # Read the first sheet to verify it exists
                df_test = pd.read_excel(self.excel_path, sheet_name=None)
                sheet_names = list(df_test.keys())
                print(f"Sheets found via pandas: {sheet_names}")
            except Exception as e:
                print(f"Warning: Could not read with pandas: {e}")
                
            # Load workbook with openpyxl
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
            print(f"Available sheets: {self.workbook.sheetnames}")
            
            # Look for Class_Enrollment sheet (case-insensitive)
            enrollment_sheet_name = None
            for sheet_name in self.workbook.sheetnames:
                if 'class_enrollment' in sheet_name.lower().replace(' ', '_'):
                    enrollment_sheet_name = sheet_name
                    break
                    
            if not enrollment_sheet_name:
                # If not found, use the first sheet
                enrollment_sheet_name = self.workbook.sheetnames[0]
                print(f"Warning: Class_Enrollment sheet not found, using first sheet: {enrollment_sheet_name}")
                
            self.enrollment_sheet = self.workbook[enrollment_sheet_name]
            print(f"Successfully loaded enrollment sheet: {enrollment_sheet_name}")
            
        except Exception as e:
            self.load_error = f"Error loading workbook: {str(e)}"
            print(self.load_error)
            import traceback
            traceback.print_exc()
            
    def get_staff_list(self):
        """Get list of all staff names"""
        if self.enrollment_sheet is None:
            print(f"Excel loading error: {self.load_error}")
            return []
            
        staff_list = []
        try:
            # Start from row 2 (assuming row 1 is header)
            for row in self.enrollment_sheet.iter_rows(min_row=2, max_col=1):
                if row[0].value:
                    # Clean up the staff name
                    staff_name = str(row[0].value).strip()
                    if staff_name and staff_name.upper() != 'STAFF NAME':
                        staff_list.append(staff_name)
                        
            print(f"Found {len(staff_list)} staff members")
        except Exception as e:
            print(f"Error reading staff list: {e}")
            import traceback
            traceback.print_exc()
            
        return staff_list
        
    def get_assigned_classes(self, staff_name):
        """Get list of classes assigned to a staff member"""
        if self.enrollment_sheet is None:
            return []
            
        assigned_classes = []
        
        try:
            # Find the staff member's row
            staff_row = None
            for row_idx, row in enumerate(self.enrollment_sheet.iter_rows(min_row=2, max_col=1), start=2):
                if row[0].value and str(row[0].value).strip() == staff_name:
                    staff_row = row_idx
                    break
                    
            if staff_row:
                # Get headers (class names) from row 1
                headers = []
                for col_idx, col in enumerate(self.enrollment_sheet.iter_cols(min_col=2, max_row=1), start=2):
                    if col[0].value:
                        header_value = str(col[0].value).strip()
                        # Only include columns that are NOT in the non-class list
                        if header_value not in NON_CLASS_COLUMNS:
                            headers.append((col_idx, header_value))
                
                # Check which classes are assigned (checkbox is True)
                for col_idx, class_name in headers:
                    cell = self.enrollment_sheet.cell(row=staff_row, column=col_idx)
                    cell_value = cell.value
                    
                    # Handle different representations of True
                    if (cell_value is True or 
                        (isinstance(cell_value, str) and cell_value.lower() in ['true', 'yes', '1', 'x', '✓']) or
                        (isinstance(cell_value, int) and cell_value == 1)):
                        assigned_classes.append(class_name)
                        
        except Exception as e:
            print(f"Error getting assigned classes: {e}")
            import traceback
            traceback.print_exc()
            
        return assigned_classes
        
    def is_staff_meeting(self, class_name):
        """Check if a class is a staff meeting (contains 'SM')"""
        return 'SM' in class_name.upper()
        
    def _parse_time_value(self, time_value):
        """Parse various time formats from Excel"""
        if time_value is None:
            return None
            
        # If it's already a string in HH:MM format
        if isinstance(time_value, str):
            # Remove any whitespace
            time_value = time_value.strip()
            # Handle common time formats
            if ':' in time_value:
                return time_value
            # Handle time without colon (e.g., "800" -> "8:00")
            elif time_value.isdigit():
                if len(time_value) == 3:
                    return f"{time_value[0]}:{time_value[1:3]}"
                elif len(time_value) == 4:
                    return f"{time_value[0:2]}:{time_value[2:4]}"
                else:
                    return time_value
                    
        # If it's a datetime object
        elif isinstance(time_value, datetime):
            return time_value.strftime('%H:%M')
            
        # If it's a time object
        elif isinstance(time_value, time):
            return time_value.strftime('%H:%M')
            
        # If it's a float (Excel time as decimal fraction of day)
        elif isinstance(time_value, (int, float)):
            # Excel stores time as fraction of day
            hours = int(time_value * 24)
            minutes = int((time_value * 24 - hours) * 60)
            return f"{hours:02d}:{minutes:02d}"
            
        # Try to convert to string and parse
        else:
            try:
                time_str = str(time_value)
                if ':' in time_str:
                    return time_str
            except:
                pass
                
        return None

    def _parse_checkbox_value(self, cell_value):
        """
        Parse checkbox values from Excel cells more reliably.
        Handles True/False booleans, Yes/No text, X marks, and 1/0 numbers.
        """
        if cell_value is None:
            return False
        
        # Handle boolean values (actual Excel checkboxes)
        if isinstance(cell_value, bool):
            return cell_value
        
        # Handle string values
        if isinstance(cell_value, str):
            cleaned = cell_value.strip().upper()
            # Check for common "true" representations
            if cleaned in ['YES', 'Y', 'TRUE', 'T', 'X', '✓', '1']:
                return True
            # Check for common "false" representations
            if cleaned in ['NO', 'N', 'FALSE', 'F', '', '0']:
                return False
        
        # Handle numeric values
        if isinstance(cell_value, (int, float)):
            return bool(cell_value)  # 0 = False, anything else = True
        
        # Default to False for unknown values
        print(f"Warning: Unknown checkbox value type {type(cell_value)}: {repr(cell_value)}")
        return False

    def get_class_details(self, class_name):
        """Get details for a specific class from its sheet - UPDATED with better checkbox handling"""
        try:
            # Try exact match first
            sheet = None
            if class_name in self.workbook.sheetnames:
                sheet = self.workbook[class_name]
            else:
                # Try case-insensitive match
                for sheet_name in self.workbook.sheetnames:
                    if sheet_name.lower() == class_name.lower():
                        sheet = self.workbook[sheet_name]
                        break
                        
            if sheet is None:
                print(f"Warning: Sheet '{class_name}' not found in workbook")
                # Return default values with the class name and a flag indicating missing data
                default_details = DEFAULT_CLASS_DETAILS.copy()
                default_details['class_name'] = class_name
                default_details['instructors_per_day'] = 0  # Default to 0 for missing sheets
                default_details['_missing_sheet'] = True  # Flag to indicate missing data
                return default_details
                
            details = {}
            
            # Extract dates from rows 1-14 (row 15 is blank)
            # Check columns B (date), C (LIVE option), D (Can work N prior), E (Location)
            has_any_dates = False
            for i in range(1, 15):  # Check rows 1-14 only (row 15 is blank)
                date_value = sheet.cell(row=i, column=2).value  # Column B
                live_option = sheet.cell(row=i, column=3).value  # Column C
                can_work_n_prior = sheet.cell(row=i, column=4).value  # Column D
                location = sheet.cell(row=i, column=5).value  # Column E
                
                if date_value:
                    has_any_dates = True
                    # Handle different date formats
                    if isinstance(date_value, datetime):
                        details[f'date_{i}'] = date_value.strftime('%m/%d/%Y')
                    elif isinstance(date_value, str) and date_value.strip():
                        details[f'date_{i}'] = date_value.strip()
                    else:
                        details[f'date_{i}'] = None
                        
                    # Use improved checkbox parsing
                    has_live = self._parse_checkbox_value(live_option)
                    can_n_prior = self._parse_checkbox_value(can_work_n_prior)
                    
                    details[f'date_{i}_has_live'] = has_live
                    details[f'date_{i}_can_work_n_prior'] = can_n_prior
                    
                    # Store location
                    details[f'date_{i}_location'] = str(location).strip() if location else ""
                    
                else:
                    # No date in this row - set defaults
                    details[f'date_{i}'] = None
                    details[f'date_{i}_has_live'] = False
                    details[f'date_{i}_can_work_n_prior'] = False
                    details[f'date_{i}_location'] = ""
            
            # If no dates were found, mark as missing data
            if not has_any_dates:
                print(f"Warning: No dates found for class '{class_name}' sheet")
                default_details = DEFAULT_CLASS_DETAILS.copy()
                default_details['class_name'] = class_name
                default_details['instructors_per_day'] = 0  # Default to 0 for classes with no dates
                default_details['_missing_dates'] = True  # Flag to indicate missing dates
                return default_details
            
            # Extract class configuration from fixed rows (updated positions)
            details['students_per_class'] = sheet.cell(row=16, column=2).value or 21  # Confirmed: Row 16
            
            # Use improved checkbox parsing for these boolean fields
            nurses_medic_separate_value = sheet.cell(row=17, column=2).value
            details['nurses_medic_separate'] = 'Yes' if self._parse_checkbox_value(nurses_medic_separate_value) else 'No'
            
            details['classes_per_day'] = sheet.cell(row=18, column=2).value or 1  # Row 18
            
            is_two_day_value = sheet.cell(row=19, column=2).value
            details['is_two_day_class'] = 'Yes' if self._parse_checkbox_value(is_two_day_value) else 'No'
                        
            # Extract time configurations (rows 20-27)
            time_labels = ['time_1_start', 'time_1_end', 'time_2_start', 'time_2_end',
                         'time_3_start', 'time_3_end', 'time_4_start', 'time_4_end']
            
            for idx, label in enumerate(time_labels):
                row_num = 20 + idx  # Rows 20-27 for time configurations
                time_value = sheet.cell(row=row_num, column=2).value
                print(f"  Row {row_num}, {label}: raw value = {repr(time_value)}, type = {type(time_value)}")
                
                parsed_time = self._parse_time_value(time_value)
                details[label] = parsed_time
                
                if parsed_time:
                    print(f"    -> Parsed as: {parsed_time}")
                else:
                    # Use default times if parsing fails
                    if label == 'time_1_start' and not parsed_time:
                        details[label] = '08:00'
                        print(f"    -> Using default: 08:00")
                    elif label == 'time_1_end' and not parsed_time:
                        details[label] = '16:00'
                        print(f"    -> Using default: 16:00")
            
            # Extract instructor count from row 28, column B (confirmed)
            instructors_per_day_value = sheet.cell(row=28, column=2).value
            
            # Convert instructor count to integer, defaulting to 0
            try:
                if instructors_per_day_value is not None:
                    details['instructors_per_day'] = int(float(instructors_per_day_value))
                else:
                    details['instructors_per_day'] = 0
                    print(f"  No instructor count found for {class_name} (cell B28 is empty)")
            except (ValueError, TypeError) as e:
                details['instructors_per_day'] = 0
                print(f"  Could not parse instructor count for {class_name}: {repr(instructors_per_day_value)}, error: {e}")
            
            # Add class name to details
            details['class_name'] = class_name
                        
            return details
            
        except Exception as e:
            print(f"Error getting class details for {class_name}: {e}")
            import traceback
            traceback.print_exc()
            # Return default values with the class name and error flag
            default_details = DEFAULT_CLASS_DETAILS.copy()
            default_details['class_name'] = class_name
            default_details['instructors_per_day'] = 0  # Default to 0 for error cases
            default_details['_error'] = str(e)  # Flag to indicate error
            return default_details

    def has_class_data(self, class_name):
        """Check if a class has proper configuration data"""
        class_details = self.get_class_details(class_name)
        
        # Check for flags indicating missing or problematic data
        if (class_details.get('_missing_sheet') or 
            class_details.get('_missing_dates') or 
            class_details.get('_error')):
            return False
        
        # Check if any dates are actually configured (checking rows 1-14)
        has_dates = any(class_details.get(f'date_{i}') for i in range(1, 15))
        return has_dates

    def get_class_dates(self, class_name):
        """Get all available dates for a specific class"""
        class_details = self.get_class_details(class_name)
        if not class_details or not self.has_class_data(class_name):
            return []
            
        dates = []
        
        # Check rows 1-14 for dates (only use the ones that exist)
        for i in range(1, 15):
            date_key = f'date_{i}'
            if date_key in class_details and class_details[date_key]:
                dates.append(class_details[date_key])
        
        return dates
    
    def get_available_dates_with_options(self, class_name):
        """Get available dates with LIVE/Virtual options for staff meetings"""
        class_details = self.get_class_details(class_name)
        if not class_details or not self.has_class_data(class_name):
            return []
            
        date_options = []
        
        # Check rows 1-14 for dates (only use the ones that exist)
        for i in range(1, 15):
            date_key = f'date_{i}'
            live_key = f'date_{i}_has_live'
            
            if date_key in class_details and class_details[date_key]:
                date_str = class_details[date_key]
                has_live = class_details.get(live_key, False)
                
                if self.is_staff_meeting(class_name):
                    # For staff meetings, provide separate options
                    if has_live:
                        date_options.append((date_str, 'LIVE', f"{date_str} (LIVE Option)"))
                        date_options.append((date_str, 'Virtual', f"{date_str} (Virtual Option)"))
                    else:
                        date_options.append((date_str, 'Virtual', f"{date_str} (Virtual Only)"))
                else:
                    # For regular classes, just return the date
                    date_options.append((date_str, None, date_str))
                    
        return date_options
    
    def needs_educators(self, class_name):
        """Check if a class needs educators"""
        class_details = self.get_class_details(class_name)
        instructor_count = class_details.get('instructors_per_day', 0)
        
        try:
            return int(float(instructor_count)) > 0 if instructor_count else False
        except (ValueError, TypeError):
            return False
    
    def get_educator_requirement(self, class_name):
        """Get the number of educators needed for a class"""
        class_details = self.get_class_details(class_name)
        instructor_count = class_details.get('instructors_per_day', 0)
        
        try:
            return int(float(instructor_count)) if instructor_count else 0
        except (ValueError, TypeError):
            return 0

    def is_educator_authorized(self, staff_name):
        """
        Check if a staff member is authorized to sign up as an educator
        based on the 'Educator AT' column in the Class_Enrollment sheet
        
        Args:
            staff_name (str): Name of the staff member to check
            
        Returns:
            bool: True if staff is authorized for educator signups, False otherwise
        """
        if self.enrollment_sheet is None:
            print(f"Excel loading error: {self.load_error}")
            return True  # Default to showing if there's an error
        
        try:
            # Find the staff member's row
            staff_row = None
            for row_idx, row in enumerate(self.enrollment_sheet.iter_rows(min_row=2, max_col=1), start=2):
                if row[0].value and str(row[0].value).strip() == staff_name:
                    staff_row = row_idx
                    break
            
            if not staff_row:
                print(f"Staff member '{staff_name}' not found in enrollment sheet")
                return True  # Default to showing if staff not found
            
            # Find the "Educator AT" column
            educator_col = None
            for col_idx, col in enumerate(self.enrollment_sheet.iter_cols(min_row=1, max_row=1), start=1):
                if col[0].value and str(col[0].value).strip() == "Educator AT":
                    educator_col = col_idx
                    break
            
            if not educator_col:
                print("Warning: 'Educator AT' column not found - defaulting to show educator signup")
                return True  # Default to showing if column not found
            
            # Check the value in the Educator AT column for this staff member
            cell = self.enrollment_sheet.cell(row=staff_row, column=educator_col)
            cell_value = cell.value
            
            # Use the improved checkbox parsing if available, otherwise fall back to original logic
            if hasattr(self, '_parse_checkbox_value'):
                is_authorized = self._parse_checkbox_value(cell_value)
            else:
                # Original logic as fallback
                is_authorized = (
                    cell_value is True or 
                    (isinstance(cell_value, str) and cell_value.lower() in ['true', 'yes', '1', 'x', '✓']) or
                    (isinstance(cell_value, int) and cell_value == 1)
                )
            
            print(f"Staff {staff_name} educator authorization: {is_authorized} (cell value: {cell_value})")
            return is_authorized
            
        except Exception as e:
            print(f"Error checking educator authorization for {staff_name}: {e}")
            import traceback
            traceback.print_exc()
            return True  # Default to showing if there's an error
            
    def get_all_classes(self):
        """Get list of all available classes"""
        if self.enrollment_sheet is None:
            return []
            
        classes = []
        try:
            # Get headers from Class_Enrollment sheet
            for col in self.enrollment_sheet.iter_cols(min_col=2, max_row=1):
                if col[0].value:
                    header_value = str(col[0].value).strip()
                    # Only include columns that are NOT in the non-class list
                    if header_value not in NON_CLASS_COLUMNS:
                        classes.append(header_value)
                        
            print(f"All available classes: {classes}")
        except Exception as e:
            print(f"Error getting all classes: {e}")
            
        return classes